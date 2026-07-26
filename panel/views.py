import json
import datetime
import os
import random
import secrets
import unicodedata
import base64
from decimal import Decimal, InvalidOperation
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .excel_logo import excel_logo

from django.conf import settings
from django.shortcuts import render, redirect
from django.urls import reverse
from django.contrib import messages
from django.http import HttpResponse, JsonResponse, FileResponse, Http404
from django.utils import timezone
from django.db.models import Sum, Q, Max, Prefetch
from django.core.files.base import ContentFile
from django.contrib.auth import authenticate, login as auth_login, logout as auth_logout
from django.contrib.auth.models import User

from .models import (Personel, KodKilit, Vardiya, Sube, Puantaj, Kalibrasyon, Irsaliye,
                     StokUrun, StokSayim, StokSayimKalem,
                     SevkiyatTalep, SevkiyatKalem, SevkiyatBirim, SevkiyatDurumu,
                     SevkiyatForm, Urun, SiparisHareket,
                     KahveSoru, GunlukSoru, SoruAyar, Bildirim, Duyuru,
                     GSosyalGonderi, GSosyalTepki,
                     EgitimDokuman, EgitimSoru, EgitimDurum, EgitimAyar, EgitimAcikCevap, PushAbonelik,
                     MolaQRAyar, SubeMolaToken, MolaOturum,
                     SubeMesaiToken, MesaiKayit, DogumGunuKutlama,
                     MutfakZayi, MutfakMaliyetKalemi, MutfakTarif, MutfakTarifKalemi, MaliyetBirim,
                     InsaatProje, InsaatMadde, InsaatMaddeDurum, InsaatKategori, InsaatSablonMadde,
                     LavaboDenetim,
                     DenetimBolum, DenetimMadde, Denetim, DenetimCevap,
                     Rol, OnayDurumu, VardiyaTipi)
from .hukuki_icerik import HUKUKI_SAYFALAR

MAX_DENEME = 5
KILIT_DK = 10
MODEL_BACKEND = 'django.contrib.auth.backends.ModelBackend'
GUN_ADLARI = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma', 'Cumartesi', 'Pazar']
UST_YONETIM = [Rol.GENEL_MUDUR, Rol.MUDUR, Rol.OPERATOR, Rol.YATIRIMCI]

TAM_YETKILI = [Rol.GENEL_MUDUR, Rol.OPERATOR, Rol.YATIRIMCI]

OFIS_ROLLERI = [Rol.GENEL_MUDUR, Rol.MUDUR, Rol.OPERATOR, Rol.YATIRIMCI, Rol.SATIN_ALMA, Rol.SEVKIYAT]
CALISMA_TIPLERI = [VardiyaTipi.SABAHCI, VardiyaTipi.ARACI, VardiyaTipi.AKSAMCI, VardiyaTipi.MUTFAK_GOREVI]

def _istemci_ip(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR')
    if xff:
        return xff.split(',')[0].strip()
    return request.META.get('REMOTE_ADDR', '') or 'bilinmiyor'

def _hafta_gunleri(secili):
    today = timezone.localdate()
    bu = today - datetime.timedelta(days=today.weekday())
    start = bu if secili == 'bu' else bu + datetime.timedelta(days=7)
    return start, start + datetime.timedelta(days=6), [start + datetime.timedelta(days=i) for i in range(7)]

def _ay_araligi(ay_str):
    try:
        ilk = datetime.datetime.strptime((ay_str or '') + '-01', '%Y-%m-%d').date()
    except ValueError:
        ilk = timezone.localdate().replace(day=1)
    sonraki = ilk.replace(year=ilk.year + 1, month=1) if ilk.month == 12 else ilk.replace(month=ilk.month + 1)
    return ilk, sonraki

def _gun_araligi(request, bas_param, bit_param):
    bas_str = (request.GET.get(bas_param) or '').strip()
    bit_str = (request.GET.get(bit_param) or '').strip()
    if not bas_str or not bit_str:
        return None
    try:
        bas = datetime.datetime.strptime(bas_str, '%Y-%m-%d').date()
        bit = datetime.datetime.strptime(bit_str, '%Y-%m-%d').date()
    except ValueError:
        return None
    if bit < bas:
        bas, bit = bit, bas
        bas_str, bit_str = bit_str, bas_str
    return bas, bit + datetime.timedelta(days=1), bas_str, bit_str

def _aktif_personel(request):
    return Personel.objects.filter(user=request.user).select_related('sube').first()

def _bildir(aliciler, mesaj, link='', tur=''):
    try:
        aliciler = [a for a in (aliciler or []) if a is not None]
        objs = [Bildirim(alici=a, mesaj=mesaj[:200], link=link, tur=tur)
                for a in aliciler]
        if objs:
            Bildirim.objects.bulk_create(objs)
    except Exception:
        aliciler = []
    try:
        _push_gonder(aliciler, mesaj, link)
    except Exception:
        pass


def _push_gonder(aliciler, mesaj, link=''):
    try:
        from pywebpush import webpush, WebPushException
    except Exception:
        return
    priv = getattr(settings, 'VAPID_PRIVATE_KEY', '')
    if not priv:
        return
    if not priv.lstrip().startswith('-----BEGIN') and not os.path.exists(priv):
        return
    ids = [a.id for a in aliciler if a is not None]
    if not ids:
        return
    claims = dict(getattr(settings, 'VAPID_CLAIMS', {'sub': 'mailto:info@geekcoffeeshop.com'}))
    payload = json.dumps({'baslik': 'Geek Panel', 'mesaj': (mesaj or '')[:150],
                          'link': link or '/bildirimler/'})
    for ab in PushAbonelik.objects.filter(personel_id__in=ids)[:400]:
        try:
            webpush(subscription_info=json.loads(ab.veri), data=payload,
                    vapid_private_key=priv, vapid_claims=dict(claims))
        except WebPushException as e:
            try:
                if getattr(e, 'response', None) is not None and e.response.status_code in (404, 410):
                    ab.delete()
            except Exception:
                pass
        except Exception:
            pass


def push_test(request):
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'mesaj': 'Oturum bulunamadı.'}, status=403)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in [Rol.OPERATOR, Rol.GENEL_MUDUR]:
        return JsonResponse({'ok': False, 'mesaj': 'Bu işlem için yetkiniz yok.'}, status=403)
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'mesaj': 'Geçersiz istek.'}, status=405)
    try:
        from pywebpush import webpush, WebPushException
    except Exception as e:
        return JsonResponse({'ok': False, 'mesaj': 'Sunucuda pywebpush kurulu değil. PythonAnywhere Bash: pip install pywebpush', 'detay': str(e)[:200]})
    priv = getattr(settings, 'VAPID_PRIVATE_KEY', '')
    if not priv:
        return JsonResponse({'ok': False, 'mesaj': 'VAPID_PRIVATE_KEY ayarı boş görünüyor (settings).'})
    if not priv.lstrip().startswith('-----BEGIN') and not os.path.exists(priv):
        return JsonResponse({'ok': False, 'mesaj': 'private_key.pem sunucuda bulunamadı. Anahtar dosyasını yükleyin ya da VAPID_PRIVATE_KEY_PEM ortam değişkenini ayarlayın.', 'detay': str(priv)})
    abonelikler = list(PushAbonelik.objects.filter(personel=personel))
    if not abonelikler:
        return JsonResponse({'ok': False, 'mesaj': 'Bu hesap için kayıtlı cihaz yok. Önce aşağıdaki "Bu cihazda bildirimleri aç" butonuna basın, sonra tekrar test edin.'})
    claims = dict(getattr(settings, 'VAPID_CLAIMS', {'sub': 'mailto:info@geekcoffeeshop.com'}))
    payload = json.dumps({'baslik': 'Geek Panel — Test', 'mesaj': 'Test bildirimi. Bunu gördüysen push çalışıyor ✔', 'link': '/bildirimler/'})
    basari, hatalar = 0, []
    for ab in abonelikler:
        try:
            webpush(subscription_info=json.loads(ab.veri), data=payload,
                    vapid_private_key=priv, vapid_claims=dict(claims))
            basari += 1
        except WebPushException as e:
            kod = None
            try:
                if getattr(e, 'response', None) is not None:
                    kod = e.response.status_code
                    if kod in (404, 410):
                        ab.delete()
            except Exception:
                pass
            hatalar.append(f'WebPush (kod {kod}): {str(e)[:160]}')
        except Exception as e:
            hatalar.append(str(e)[:160])
    if basari:
        return JsonResponse({'ok': True, 'mesaj': f'{basari} cihaza test bildirimi gönderildi. Birkaç saniyede gelmezse cihaz/tarayıcı bildirim iznini kontrol edin. (iPhone: uygulama ana ekrana eklenmiş olmalı.)'})
    return JsonResponse({'ok': False, 'mesaj': 'Gönderim başarısız oldu.', 'detay': ' | '.join(hatalar)[:400]})


def push_kaydet(request):
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False}, status=403)
    personel = _aktif_personel(request)
    if personel is None:
        return JsonResponse({'ok': False}, status=403)
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    try:
        data = json.loads(request.body.decode('utf-8'))
        endpoint = data.get('endpoint')
        if not endpoint:
            return JsonResponse({'ok': False}, status=400)
        PushAbonelik.objects.update_or_create(
            endpoint=endpoint, defaults={'personel': personel, 'veri': json.dumps(data)})
        return JsonResponse({'ok': True})
    except Exception:
        return JsonResponse({'ok': False}, status=400)

def _rol_personelleri(*roller):
    return list(Personel.objects.filter(rol__in=roller))

def _sube_sefleri(sube):
    if not sube:
        return []
    return list(Personel.objects.filter(sube=sube, rol=Rol.SEF))

def _sube_yoneticileri(sube):
    if not sube:
        return []
    return list(Personel.objects.filter(rol=Rol.MUDUR, sorumlu_subeler=sube).distinct())

def _yonetici_sube(request, subeler):
    sid = request.GET.get('sube_id')
    if sid:
        request.session['sel_sube_id'] = sid
    else:
        sid = request.session.get('sel_sube_id')
    sel = next((s for s in subeler if str(s.id) == str(sid)), None)
    if sel is None:
        sel = next((s for s in subeler if not s.depo_mu), subeler[0] if subeler else None)
    return sel

def _yon_subeler(personel):
    if personel and personel.rol == Rol.MUDUR:
        return list(personel.sorumlu_subeler.order_by('ad'))
    if personel and personel.rol == Rol.MAGAZA_MUDURU:
        if personel.sorumlu_subeler.exists():
            return list(personel.sorumlu_subeler.order_by('ad'))
        return list(Sube.objects.filter(id=personel.sube_id)) if personel.sube_id else []
    return list(Sube.objects.order_by('ad'))

def _vardiya_tablo(personeller, start, end, gunler):
    shifts = list(Vardiya.objects.filter(personel__in=personeller, tarih__range=[start, end])) if personeller else []
    smap = {(v.personel_id, v.tarih): v for v in shifts}
    rows = [{'personel': p, 'hucreler': [{'tarih': g, 'vardiya': smap.get((p.id, g))} for g in gunler]}
            for p in personeller]
    d = set(v.durum for v in shifts)
    if OnayDurumu.ONAY_BEKLIYOR in d:
        durum, sinif = 'Onay Bekliyor', 'badge-wait'
    elif OnayDurumu.REDDEDILDI in d:
        durum, sinif = 'Reddedildi', 'badge-no'
    elif OnayDurumu.ONAYLANDI in d:
        durum, sinif = 'Onaylandı', 'badge-ok'
    elif OnayDurumu.TASLAK in d:
        durum, sinif = 'Taslak', 'badge-info'
    else:
        durum, sinif = 'Planlanmadı', 'badge-info'
    red = next((v.red_notu for v in shifts if v.durum == OnayDurumu.REDDEDILDI and v.red_notu), None)
    return {
        'rows': rows, 'hafta_durum': durum, 'durum_sinif': sinif,
        'red_notu': red, 'onay_bekleyen': OnayDurumu.ONAY_BEKLIYOR in d,
        'onaya_gonderilebilir': bool(d & {OnayDurumu.TASLAK, OnayDurumu.REDDEDILDI}),
        'gun_basliklari': [{'gun_adi': GUN_ADLARI[i], 'tarih': g} for i, g in enumerate(gunler)],
    }

def _puantaj_hesapla(personel, bas, son, manuel_ay=None):
    if manuel_ay is not None:
        try:
            rec = Puantaj.objects.filter(personel=personel, ay=manuel_ay).first()
        except Exception:
            rec = None
        if rec and rec.manuel_duzenlendi:
            # Manuel kayıt, düzenleme anına kadarki tabanı temsil eder.
            # Düzenlemeden SONRA eklenen/değişen vardiyalar otomatik olarak üstüne eklenir.
            esik = rec.guncelleme.date() if rec.guncelleme else bas
            ek = personel.vardiyalar.filter(tarih__gt=esik, tarih__lt=son).exclude(durum=OnayDurumu.REDDEDILDI)
            return {
                'calisilan': rec.calisilan_gun + ek.filter(vardiya_tipi__in=CALISMA_TIPLERI).count(),
                'eksik': rec.eksik_gun + ek.filter(vardiya_tipi=VardiyaTipi.DEVAMSIZ).count(),
                'izinli': rec.izinli_gun + ek.filter(vardiya_tipi=VardiyaTipi.IZINLI).count(),
                'yillik': rec.yillik_gun + ek.filter(vardiya_tipi=VardiyaTipi.YILLIK_IZIN).count(),
                'raporlu': rec.raporlu_gun + ek.filter(vardiya_tipi=VardiyaTipi.RAPORLU).count(),
                'manuel': True,
            }

    s = personel.vardiyalar.filter(tarih__gte=bas, tarih__lt=son).exclude(durum=OnayDurumu.REDDEDILDI)
    return {
        'calisilan': s.filter(vardiya_tipi__in=CALISMA_TIPLERI).count(),
        'izinli': s.filter(vardiya_tipi=VardiyaTipi.IZINLI).count(),
        'yillik': s.filter(vardiya_tipi=VardiyaTipi.YILLIK_IZIN).count(),
        'raporlu': s.filter(vardiya_tipi=VardiyaTipi.RAPORLU).count(),
        'eksik': s.filter(vardiya_tipi=VardiyaTipi.DEVAMSIZ).count(),
        'manuel': False,
    }

def _cikis_mi(request):
    return request.method == 'POST' and request.POST.get('islem') == 'cikis'

def _logout(request):
    auth_logout(request)
    request.session.flush()
    return redirect('ana_sayfa')

def ana_sayfa(request):
    if not request.user.is_authenticated:
        if request.method == 'POST':
            islem = request.POST.get('islem')
            if islem == 'kod_giris':
                return _kod_giris(request)
            if islem == 'sifre_giris':
                return _sifre_giris(request)
            return redirect('ana_sayfa')
        return render(request, 'personel_giris.html')

    if _cikis_mi(request):
        return _logout(request)

    personel = _aktif_personel(request)
    if personel is None:
        return render(request, 'personel_panel.html', {'personel': None, 'yonetici_bekliyor': True, 'aktif': 'home'})
    if personel.rol in UST_YONETIM:
        return redirect('gosterge')
    if personel.rol in (Rol.SATIN_ALMA, Rol.SEVKIYAT):
        return redirect('sevkiyat')
    if personel.rol == Rol.EGITMEN:
        return redirect('soru_yonetimi')
    if personel.rol in SORU_ROLLERI and _soru_sistemi_aktif():
        _gs = _gunluk_soru_getir_veya_ata(personel, timezone.localdate())
        if _gs is not None:
            _gunluk_finalize(_gs)
            if not _gs.cevaplandi:
                return redirect('gunluk_soru')
    if personel.rol in (Rol.SEF, Rol.MAGAZA_MUDURU):
        return _sef_home(request, personel)
    if personel.rol == Rol.MUTFAK_SORUMLUSU:
        return redirect('mutfak_vardiya')
    return _personel_home(request, personel)

def _personel_home(request, personel):
    today = timezone.localdate()

    start, end, gunler = _hafta_gunleri('bu')
    vmap = {v.tarih: v for v in personel.vardiyalar.filter(tarih__range=[start, end], durum=OnayDurumu.ONAYLANDI)}
    hafta = [{'tarih': g, 'gun_adi': GUN_ADLARI[i], 'bugun': g == today, 'vardiya': vmap.get(g)}
             for i, g in enumerate(gunler)]
    g_start, g_end, g_gunler = _hafta_gunleri('gelecek')
    g_vmap = {v.tarih: v for v in personel.vardiyalar.filter(tarih__range=[g_start, g_end], durum=OnayDurumu.ONAYLANDI)}
    gelecek = [{'tarih': g, 'gun_adi': GUN_ADLARI[i], 'bugun': False, 'vardiya': g_vmap.get(g)}
               for i, g in enumerate(g_gunler)]
    ctx = {'personel': personel, 'aktif': 'home', 'is_gm': False,
           'hafta': hafta, 'haftabasi': start, 'haftasonu': end,
           'gelecek': gelecek, 'g_basi': g_start, 'g_sonu': g_end}
    return render(request, 'personel_panel.html', ctx)

def _sef_home(request, personel):
    magaza_subeler = None
    if personel.rol == Rol.MAGAZA_MUDURU and personel.sorumlu_subeler.exists():
        magaza_subeler = list(personel.sorumlu_subeler.order_by('ad'))
        sube = _yonetici_sube(request, magaza_subeler)
    else:
        sube = personel.sube
    secili = request.GET.get('hafta', 'gelecek')
    if secili not in ('bu', 'gelecek'):
        secili = 'gelecek'
    start, end, gunler = _hafta_gunleri(secili)

    if request.method == 'POST':
        islem = request.POST.get('islem')
        if not sube:
            messages.error(request, "Şubeniz tanımlı değil. Yöneticinize başvurun.")
            return redirect('ana_sayfa')
        if islem == 'vardiya_kaydet':
            _vardiya_kaydet(request, sube)
            return redirect(f'/?hafta={secili}')
        if islem == 'onaya_gonder':
            Vardiya.objects.filter(personel__sube=sube, tarih__range=[start, end],
                                   durum__in=[OnayDurumu.TASLAK, OnayDurumu.REDDEDILDI]
                                   ).update(durum=OnayDurumu.ONAY_BEKLIYOR, red_notu=None)
            _bildir(_sube_yoneticileri(sube),
                    "Vardiya planı onay bekliyor: %s" % (sube.ad if sube else ''), '/', 'vardiya')
            messages.success(request, "Vardiya programı yönetici onayına gönderildi.")
            return redirect(f'/?hafta={secili}')
        if islem == 'personel_ekle':
            ad = request.POST.get('ad_soyad', '').strip()
            if ad:
                yeni = Personel.objects.create(ad_soyad=ad, sube=sube, rol=Rol.PERSONEL)
                messages.success(request, f"{yeni.ad_soyad} eklendi. Giriş kodu: {yeni.giris_kodu}")
            else:
                messages.error(request, "Ad soyad boş olamaz.")
            return redirect(f'/?hafta={secili}')
        if islem == 'personel_cikar':
            Personel.objects.filter(id=request.POST.get('personel_id'), sube=sube, rol=Rol.PERSONEL).delete()
            messages.success(request, "Personel şubeden çıkarıldı.")
            return redirect(f'/?hafta={secili}')

    personeller = list(sube.personeller.order_by('ad_soyad')) if sube else []
    tablo = _vardiya_tablo(personeller, start, end, gunler)
    ctx = {'personel': personel, 'aktif': 'home', 'is_gm': False, 'sube': sube,
           'magaza_subeler': magaza_subeler,
           'vardiya_tipleri': VardiyaTipi.choices, 'secili': secili,
           'haftabasi': start, 'haftasonu': end, 'personeller': personeller}
    ctx.update(tablo)
    return render(request, 'sef_panel.html', ctx)

def _vardiya_kaydet(request, sube, durum=OnayDurumu.TASLAK):
    pid = request.POST.get('target_personel_id')
    tarih_str = request.POST.get('vardiya_tarihi', '')
    tip = request.POST.get('vardiya_tipi', '')
    try:
        t = datetime.datetime.strptime(tarih_str, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "Geçersiz tarih.")
        return
    hedef = Personel.objects.filter(id=pid, sube=sube).first()
    if not hedef:
        messages.error(request, "Yetkisiz işlem.")
        return
    if tip == 'Sil':
        Vardiya.objects.filter(personel=hedef, tarih=t).delete()
    elif tip in VardiyaTipi.values:
        Vardiya.objects.update_or_create(
            personel=hedef, tarih=t,
            defaults={'vardiya_tipi': tip, 'durum': durum, 'red_notu': None})

def vardiya_home(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in UST_YONETIM:
        return redirect('ana_sayfa')
    return _yonetici_vardiya(request, personel)


def _yonetici_vardiya(request, personel):
    is_gm = personel.rol == Rol.GENEL_MUDUR
    subeler = _yon_subeler(personel)
    sel_sube = _yonetici_sube(request, subeler)
    secili = request.GET.get('hafta', 'gelecek')
    if secili not in ('bu', 'gelecek'):
        secili = 'gelecek'
    start, end, gunler = _hafta_gunleri(secili)

    if request.method == 'POST' and sel_sube:
        islem = request.POST.get('islem')
        if islem == 'vardiya_kaydet':
            _vardiya_kaydet(request, sel_sube, durum=OnayDurumu.ONAYLANDI)
        elif islem == 'plan_onayla':
            Vardiya.objects.filter(personel__sube=sel_sube, tarih__range=[start, end],
                                   durum=OnayDurumu.ONAY_BEKLIYOR).update(durum=OnayDurumu.ONAYLANDI, red_notu=None)
            _bildir(_sube_sefleri(sel_sube),
                    "Vardiya planınız onaylandı: %s" % (sel_sube.ad if sel_sube else ''), '/', 'vardiya')
            if sel_sube:
                _bildir(list(Personel.objects.filter(sube=sel_sube, rol=Rol.PERSONEL)),
                        "Bu haftaki vardiya planınız yayınlandı.", '/', 'vardiya')
            messages.success(request, "Vardiya planı onaylandı.")
        elif islem == 'plan_reddet':
            neden = request.POST.get('red_notu', '').strip() or 'Neden belirtilmedi.'
            Vardiya.objects.filter(personel__sube=sel_sube, tarih__range=[start, end],
                                   durum=OnayDurumu.ONAY_BEKLIYOR).update(durum=OnayDurumu.REDDEDILDI, red_notu=neden)
            _bildir(_sube_sefleri(sel_sube),
                    "Vardiya planınız reddedildi: %s" % (sel_sube.ad if sel_sube else ''), '/', 'vardiya')
            messages.success(request, "Vardiya planı reddedildi ve şefe geri gönderildi.")
        return redirect(f'/?hafta={secili}')

    personeller = list(sel_sube.personeller.order_by('ad_soyad')) if sel_sube else []
    tablo = _vardiya_tablo(personeller, start, end, gunler)
    ctx = {'personel': personel, 'aktif': 'home', 'is_gm': is_gm, 'subeler': subeler, 'sel_sube': sel_sube,
           'vardiya_tipleri': VardiyaTipi.choices, 'secili': secili, 'haftabasi': start, 'haftasonu': end}
    ctx.update(tablo)
    return render(request, 'yonetici_vardiya.html', ctx)


MUTFAK_YONETEBILEN = [Rol.GENEL_MUDUR, Rol.MUDUR, Rol.OPERATOR, Rol.MUTFAK_SORUMLUSU]
MUTFAK_ZAYI_GORUNTULE = [Rol.GENEL_MUDUR, Rol.MUDUR, Rol.OPERATOR, Rol.YATIRIMCI]


def mutfak_vardiya_sayfa(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in MUTFAK_YONETEBILEN:
        return redirect('ana_sayfa')

    secili = request.GET.get('hafta', 'bu')
    if secili not in ('bu', 'gelecek'):
        secili = 'bu'
    start, end, gunler = _hafta_gunleri(secili)
    subeler = list(Sube.objects.filter(depo_mu=False).order_by('ad'))
    mutfak_personeli = list(Personel.objects.filter(
        rol__in=(Rol.MUTFAK_PERSONEL, Rol.MUTFAK_SORUMLUSU)).order_by('-rol', 'ad_soyad'))

    if request.method == 'POST':
        if request.POST.get('islem') == 'mutfak_personel_ekle':
            ad = (request.POST.get('ad_soyad') or '').strip()
            if ad:
                yeni = Personel.objects.create(ad_soyad=ad[:100], rol=Rol.MUTFAK_PERSONEL)
                messages.success(request, f"{ad} mutfak ekibine eklendi. Giriş kodu: {yeni.giris_kodu}")
            else:
                messages.error(request, "Ad soyad gerekli.")
            return redirect(f'/mutfak/vardiya/?hafta={secili}')
        if request.POST.get('islem') == 'mutfak_personel_cikar':
            k = Personel.objects.filter(id=request.POST.get('personel_id'), rol=Rol.MUTFAK_PERSONEL).first()
            if k:
                ad = k.ad_soyad
                (k.user or k).delete()
                messages.success(request, f"{ad} mutfak ekibinden çıkarıldı.")
            return redirect(f'/mutfak/vardiya/?hafta={secili}')
        pid = request.POST.get('personel_id')
        tarih_str = request.POST.get('tarih')
        secim = request.POST.get('secim', '')
        kisi = Personel.objects.filter(id=pid, rol__in=(Rol.MUTFAK_PERSONEL, Rol.MUTFAK_SORUMLUSU)).first()
        try:
            tarih = datetime.datetime.strptime(tarih_str, '%Y-%m-%d').date()
        except (TypeError, ValueError):
            kisi = None
        if kisi:
            if secim == '':
                Vardiya.objects.filter(personel=kisi, tarih=tarih).delete()
            elif secim.startswith('sube:'):
                sid = secim.split(':', 1)[1]
                sb = Sube.objects.filter(id=sid).first()
                if sb:
                    Vardiya.objects.update_or_create(
                        personel=kisi, tarih=tarih,
                        defaults={'vardiya_tipi': VardiyaTipi.MUTFAK_GOREVI, 'atanan_sube': sb,
                                 'durum': OnayDurumu.ONAYLANDI, 'red_notu': None})
            elif secim in (VardiyaTipi.IZINLI, VardiyaTipi.YILLIK_IZIN, VardiyaTipi.RAPORLU, VardiyaTipi.DEVAMSIZ):
                Vardiya.objects.update_or_create(
                    personel=kisi, tarih=tarih,
                    defaults={'vardiya_tipi': secim, 'atanan_sube': None,
                             'durum': OnayDurumu.ONAYLANDI, 'red_notu': None})
        return redirect(f'/mutfak/vardiya/?hafta={secili}')

    kayitlar = {(v.personel_id, v.tarih): v for v in
                Vardiya.objects.filter(personel__in=mutfak_personeli, tarih__range=[start, end])}
    tablo = []
    for k in mutfak_personeli:
        satir = {'personel': k, 'gunler': []}
        for g in gunler:
            v = kayitlar.get((k.id, g))
            satir['gunler'].append({'tarih': g, 'vardiya': v})
        tablo.append(satir)

    return render(request, 'mutfak_vardiya.html', {
        'personel': personel, 'aktif': 'mutfak_vardiya', 'tablo': tablo,
        'subeler': subeler, 'secili': secili, 'haftabasi': start, 'haftasonu': end, 'gunler': gunler,
        'ozel_tipler': [VardiyaTipi.IZINLI, VardiyaTipi.YILLIK_IZIN, VardiyaTipi.RAPORLU, VardiyaTipi.DEVAMSIZ],
    })


def mutfak_atamalari_sayfa(request):
    """Mağaza müdürü / şef: kendi şubesine bugün/bu hafta atanan mutfak personelini görür."""
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    if personel.rol in (Rol.SEF, Rol.MAGAZA_MUDURU):
        sube = personel.sube
    elif personel.rol in (Rol.GENEL_MUDUR, Rol.MUDUR, Rol.OPERATOR, Rol.YATIRIMCI):
        subeler = _yon_subeler(personel)
        sube = _yonetici_sube(request, subeler)
    else:
        return redirect('ana_sayfa')
    if not sube:
        return redirect('ana_sayfa')

    bugun = timezone.localdate()
    start = bugun - datetime.timedelta(days=bugun.weekday())
    gunler = [start + datetime.timedelta(days=i) for i in range(7)]
    atamalar = (Vardiya.objects.filter(atanan_sube=sube, tarih__range=[start, gunler[-1]],
                                       vardiya_tipi=VardiyaTipi.MUTFAK_GOREVI)
                .select_related('personel').order_by('tarih'))
    gunluk = {g: [] for g in gunler}
    for a in atamalar:
        if a.tarih in gunluk:
            gunluk[a.tarih].append(a.personel.ad_soyad if a.personel else '—')
    liste = [{'tarih': g, 'kisiler': gunluk[g]} for g in gunler]
    return render(request, 'mutfak_atamalari.html', {
        'personel': personel, 'aktif': 'mutfak_atamalari', 'sube': sube, 'liste': liste,
    })


MUTFAK_ZAYI_GUN = 35


def _mutfak_zayi_temizle():
    sinir = timezone.now() - datetime.timedelta(days=MUTFAK_ZAYI_GUN)
    for z in MutfakZayi.objects.filter(olusturma__lt=sinir):
        if z.foto:
            z.foto.delete(save=False)
        z.delete()


def mutfak_zayi_sayfa(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    ekleyebilir = personel.rol == Rol.MUTFAK_PERSONEL
    goruntuleyebilir = personel.rol in MUTFAK_ZAYI_GORUNTULE or personel.rol in (Rol.MAGAZA_MUDURU, Rol.MUTFAK_SORUMLUSU) or ekleyebilir
    if not goruntuleyebilir:
        return redirect('ana_sayfa')
    _mutfak_zayi_temizle()

    if request.method == 'POST' and ekleyebilir:
        if request.POST.get('islem') == 'mutfak_zayi_yukle':
            data = request.POST.get('foto_data', '')
            raw = None
            if data.startswith('data:image'):
                try:
                    raw = base64.b64decode(data.split(',', 1)[1])
                except (ValueError, IndexError):
                    raw = None
            aciklama = (request.POST.get('aciklama') or '').strip()
            bugun = timezone.localdate()
            aktif_gorev = Vardiya.objects.filter(personel=personel, tarih=bugun,
                                                 vardiya_tipi=VardiyaTipi.MUTFAK_GOREVI).first()
            sube = (aktif_gorev.atanan_sube if aktif_gorev else None) or personel.sube
            if raw and 100 < len(raw) <= 8 * 1024 * 1024:
                z = MutfakZayi(personel=personel, personel_ad_arsiv=personel.ad_soyad,
                               sube=sube, aciklama=aciklama[:2000])
                fname = f"mutfakzayi_{personel.id}_{timezone.now():%Y%m%d_%H%M%S}.jpg"
                z.foto.save(fname, ContentFile(raw), save=True)
                messages.success(request, "Zayi kaydı yüklendi.")
                sube_adi = sube.ad if sube else "şube tanımsız"
                alicilar = _rol_personelleri(Rol.MUTFAK_SORUMLUSU)
                if sube:
                    alicilar += Personel.objects.filter(sube=sube, rol=Rol.MAGAZA_MUDURU)
                _bildir(alicilar, "Mutfak zayi kaydı: %s (%s)" % (sube_adi, personel.ad_soyad),
                        '/mutfak/zayi/', 'mutfak_zayi')
            else:
                messages.error(request, "Görüntü alınamadı. Lütfen kameradan tekrar çekin.")
        return redirect('mutfak_zayi')

    # Şube bazlı görünürlük kapsamı
    subeler_secim = []
    if personel.rol == Rol.MAGAZA_MUDURU:
        sube_ids = [personel.sube_id] if personel.sube_id else []
    elif personel.rol == Rol.MUDUR:
        sube_ids = list(personel.sorumlu_subeler.values_list('id', flat=True))
        subeler_secim = list(Sube.objects.filter(id__in=sube_ids).order_by('ad')) if len(sube_ids) > 1 else []
    elif personel.rol in (Rol.GENEL_MUDUR, Rol.OPERATOR, Rol.YATIRIMCI, Rol.MUTFAK_SORUMLUSU):
        sube_ids = list(Sube.objects.values_list('id', flat=True))
        subeler_secim = list(Sube.objects.order_by('ad'))
    else:
        sube_ids = []

    sec = request.GET.get('sube')
    if sec and sec.isdigit() and int(sec) in sube_ids:
        filtre_ids = [int(sec)]
        secili = int(sec)
    else:
        filtre_ids = sube_ids
        secili = 0

    today = timezone.localdate()
    try:
        secili_tarih = datetime.datetime.strptime(request.GET.get('mz_tarih', ''), '%Y-%m-%d').date()
    except ValueError:
        secili_tarih = today
    secili_tarih = min(secili_tarih, today)
    gun_bas = datetime.datetime.combine(secili_tarih, datetime.time.min, tzinfo=timezone.get_current_timezone())
    gun_son = gun_bas + datetime.timedelta(days=1)

    if ekleyebilir:
        gorseller = list(MutfakZayi.objects.filter(personel=personel, olusturma__gte=gun_bas, olusturma__lt=gun_son)
                         .select_related('personel', 'sube').order_by('-olusturma'))
    else:
        gorseller = list(MutfakZayi.objects.filter(
            Q(sube_id__in=filtre_ids) | Q(sube__isnull=True, personel__sube_id__in=filtre_ids),
            olusturma__gte=gun_bas, olusturma__lt=gun_son)
                         .select_related('personel', 'sube').order_by('-olusturma'))

    return render(request, 'mutfak_zayi.html', {
        'personel': personel, 'aktif': 'mutfak_zayi', 'ekleyebilir': ekleyebilir, 'gorseller': gorseller,
        'subeler': subeler_secim, 'secili': secili,
        'secili_tarih': secili_tarih.strftime('%Y-%m-%d'), 'bugun': today.strftime('%Y-%m-%d'),
        'saklama_gun': MUTFAK_ZAYI_GUN,
    })


def mutfak_maliyet_sayfa(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in MUTFAK_YONETEBILEN:
        return redirect('ana_sayfa')

    if request.method == 'POST':
        islem = request.POST.get('islem')
        if islem == 'urun_ekle':
            ad = (request.POST.get('ad') or '').strip()
            birim = request.POST.get('birim') if request.POST.get('birim') in MaliyetBirim.values else MaliyetBirim.KG
            try:
                fiyat = float((request.POST.get('fiyat') or '0').replace(',', '.'))
            except ValueError:
                fiyat = 0
            if ad and fiyat > 0:
                MutfakMaliyetKalemi.objects.update_or_create(ad=ad, defaults={'fiyat': fiyat, 'birim': birim})
                messages.success(request, "Ürün fiyatı kaydedildi.")
            else:
                messages.error(request, "Ürün adı ve geçerli bir birim fiyatı gerekli.")
        elif islem == 'urun_sil':
            MutfakMaliyetKalemi.objects.filter(id=request.POST.get('urun_id')).delete()
        elif islem == 'tarif_ekle':
            ad = (request.POST.get('tarif_ad') or '').strip()
            if ad:
                MutfakTarif.objects.create(ad=ad, olusturan=personel)
                messages.success(request, "Tarif oluşturuldu.")
        elif islem == 'tarif_sil':
            MutfakTarif.objects.filter(id=request.POST.get('tarif_id')).delete()
        elif islem == 'kalem_ekle':
            tarif = MutfakTarif.objects.filter(id=request.POST.get('tarif_id')).first()
            urun = MutfakMaliyetKalemi.objects.filter(id=request.POST.get('urun_id')).first()
            try:
                miktar = float((request.POST.get('miktar') or '0').replace(',', '.'))
            except ValueError:
                miktar = 0
            if tarif and urun and miktar > 0:
                MutfakTarifKalemi.objects.create(tarif=tarif, urun=urun, miktar=miktar)
        elif islem == 'kalem_sil':
            MutfakTarifKalemi.objects.filter(id=request.POST.get('kalem_id')).delete()
        return redirect('mutfak_maliyet')

    urunler = list(MutfakMaliyetKalemi.objects.order_by('ad'))
    tarifler = list(MutfakTarif.objects.prefetch_related('kalemler__urun').order_by('ad'))
    return render(request, 'mutfak_maliyet.html', {
        'personel': personel, 'aktif': 'mutfak_maliyet', 'urunler': urunler, 'tarifler': tarifler,
        'birimler': MaliyetBirim.choices,
    })


def mutfak_puantaj_sayfa(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in MUTFAK_YONETEBILEN:
        return redirect('ana_sayfa')

    puantaj_duzenleyebilir = personel.rol in (Rol.GENEL_MUDUR, Rol.MUDUR, Rol.OPERATOR)
    ay_str = request.GET.get('puantaj_ay') or timezone.localdate().strftime('%Y-%m')
    ay_ilk, ay_son = _ay_araligi(ay_str)
    mutfak_personeli = list(Personel.objects.filter(rol=Rol.MUTFAK_PERSONEL).order_by('ad_soyad'))

    if request.method == 'POST':
        if not puantaj_duzenleyebilir:
            return redirect(f'/mutfak/puantaj/?puantaj_ay={ay_str}')
        islem = request.POST.get('islem')
        hedef = Personel.objects.filter(id=request.POST.get('target_personel_id'), rol=Rol.MUTFAK_PERSONEL).first()
        if hedef and islem == 'puantaj_kaydet':
            def _say(ad):
                try:
                    return max(0, int(request.POST.get(ad, 0)))
                except (TypeError, ValueError):
                    return 0
            Puantaj.objects.update_or_create(
                personel=hedef, ay=ay_ilk,
                defaults={'calisilan_gun': _say('calisilan_gun'), 'eksik_gun': _say('eksik_gun'),
                          'izinli_gun': _say('izinli_gun'), 'yillik_gun': _say('yillik_gun'),
                          'raporlu_gun': _say('raporlu_gun'),
                          'manuel_duzenlendi': True})
            messages.success(request, f"{hedef.ad_soyad} puantajı elle güncellendi.")
        elif hedef and islem == 'puantaj_sifirla':
            Puantaj.objects.filter(personel=hedef, ay=ay_ilk).delete()
            messages.success(request, "Puantaj otomatik hesaplamaya döndürüldü.")
        return redirect(f'/mutfak/puantaj/?puantaj_ay={ay_str}')

    liste = []
    for p in mutfak_personeli:
        d = _puantaj_hesapla(p, ay_ilk, ay_son, manuel_ay=ay_ilk)
        d['personel'] = p
        d['ayrilan'] = False
        d['yillik'] = d.get('yillik', 0)
        d['hakedis'] = d['calisilan'] + d['izinli'] + d['yillik']
        liste.append(d)

    return render(request, 'mutfak_puantaj.html', {
        'personel': personel, 'aktif': 'mutfak_puantaj',
        'puantaj_duzenleyebilir': puantaj_duzenleyebilir,
        'puantaj_listesi': liste, 'selected_ay_str': ay_str,
    })


def puantaj_sayfa(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol == Rol.PERSONEL:
        return redirect('ana_sayfa')

    is_gm = personel.rol == Rol.GENEL_MUDUR
    is_yon = personel.rol in UST_YONETIM
    puantaj_duzenleyebilir = personel.rol in (Rol.GENEL_MUDUR, Rol.MUDUR, Rol.OPERATOR)
    subeler = _yon_subeler(personel) if is_yon else []
    sel_sube = _yonetici_sube(request, subeler) if is_yon else personel.sube
    ay_str = request.GET.get('puantaj_ay') or timezone.localdate().strftime('%Y-%m')
    ay_ilk, ay_son = _ay_araligi(ay_str)

    if request.method == 'POST' and sel_sube:
        if not puantaj_duzenleyebilir:
            return redirect(f'/puantaj/?puantaj_ay={ay_str}')
        islem = request.POST.get('islem')
        hedef = Personel.objects.filter(id=request.POST.get('target_personel_id'), sube=sel_sube).first()
        if hedef and islem == 'puantaj_kaydet':
            def _say(ad):
                try:
                    return max(0, int(request.POST.get(ad, 0)))
                except (TypeError, ValueError):
                    return 0
            Puantaj.objects.update_or_create(
                personel=hedef, ay=ay_ilk,
                defaults={'calisilan_gun': _say('calisilan_gun'), 'eksik_gun': _say('eksik_gun'),
                          'izinli_gun': _say('izinli_gun'), 'yillik_gun': _say('yillik_gun'),
                          'raporlu_gun': _say('raporlu_gun'),
                          'manuel_duzenlendi': True})
            messages.success(request, f"{hedef.ad_soyad} puantajı elle güncellendi.")
        elif hedef and islem == 'puantaj_sifirla':
            Puantaj.objects.filter(personel=hedef, ay=ay_ilk).delete()
            messages.success(request, "Puantaj otomatik hesaplamaya döndürüldü.")
        return redirect(f'/puantaj/?puantaj_ay={ay_str}')

    personeller = list(sel_sube.personeller.order_by('ad_soyad')) if sel_sube else []
    aralik = _gun_araligi(request, 'puantaj_bas', 'puantaj_bit')
    aralik_mod = aralik is not None
    if aralik_mod:
        bas, son, bas_str, bit_str = aralik
        hesap_bas, hesap_son, manuel_ay = bas, son, None
    else:
        hesap_bas, hesap_son, manuel_ay = ay_ilk, ay_son, ay_ilk
        bas_str = bit_str = ''

    liste = []
    for p in personeller:
        d = _puantaj_hesapla(p, hesap_bas, hesap_son, manuel_ay=manuel_ay)
        d['personel'] = p
        d['ayrilan'] = False

        d['yillik'] = d.get('yillik', 0)
        d['hakedis'] = d['calisilan'] + d['izinli'] + d['yillik']
        liste.append(d)

    if not aralik_mod and sel_sube:
        try:
            ayrilanlar = list(Puantaj.objects.filter(personel__isnull=True, sube_arsiv=sel_sube, ay=ay_ilk))
        except Exception:
            ayrilanlar = []
        for r in ayrilanlar:
            liste.append({
                'calisilan': r.calisilan_gun, 'eksik': r.eksik_gun, 'izinli': r.izinli_gun,
                'yillik': r.yillik_gun, 'raporlu': r.raporlu_gun, 'manuel': True,
                'hakedis': r.calisilan_gun + r.izinli_gun + r.yillik_gun,
                'personel': None, 'ayrilan': True, 'ayrilan_ad': r.personel_ad_soyad_arsiv,
            })
    return render(request, 'puantaj.html', {
        'personel': personel, 'aktif': 'puantaj', 'is_gm': is_gm, 'is_yon': is_yon,
        'puantaj_duzenleyebilir': puantaj_duzenleyebilir,
        'subeler': subeler, 'sel_sube': sel_sube, 'puantaj_listesi': liste,
        'selected_ay_str': ay_str, 'aralik_mod': aralik_mod,
        'puantaj_bas': bas_str, 'puantaj_bit': bit_str,
    })

def _kullanici_adi_uret(ad):
    t = ad.lower().strip()
    for k, v in {'ı': 'i', 'ş': 's', 'ğ': 'g', 'ü': 'u', 'ö': 'o', 'ç': 'c', 'İ': 'i'}.items():
        t = t.replace(k, v)
    t = unicodedata.normalize('NFKD', t).encode('ascii', 'ignore').decode()
    t = '.'.join(t.split())
    base = ''.join(ch for ch in t if ch.isalnum() or ch == '.') or 'kullanici'
    uname, i = base, 1
    while User.objects.filter(username=uname).exists():
        i += 1
        uname = f"{base}{i}"
    return uname

def _yeni_sifre():
    return secrets.token_urlsafe(6)

def ekip_sayfa(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in TAM_YETKILI:
        return redirect('ana_sayfa')

    subeler = _yon_subeler(personel)
    is_tam = personel.rol in TAM_YETKILI
    if request.method == 'POST':
        islem = request.POST.get('islem')
        if islem in ('magaza_muduru_ekle', 'magaza_muduru_cikar') and personel.rol in (Rol.GENEL_MUDUR, Rol.OPERATOR):
            sube_id_list = [s.id for s in subeler]
            if islem == 'magaza_muduru_ekle':
                ad = request.POST.get('ad_soyad', '').strip()
                sid = request.POST.get('mm_sube_id') or ''
                if ad and sid.isdigit() and int(sid) in sube_id_list:
                    yeni = Personel.objects.create(ad_soyad=ad, sube_id=int(sid), rol=Rol.MAGAZA_MUDURU)
                    messages.success(request, f"{ad} mağaza müdürü olarak eklendi. Giriş kodu: {yeni.giris_kodu} (kodla giriş yapar).")
                else:
                    messages.error(request, "Ad soyad ve yetkili olduğunuz bir şube zorunlu.")
            else:
                s = Personel.objects.filter(id=request.POST.get('personel_id'), rol=Rol.MAGAZA_MUDURU).first()
                if s and s.sube_id in sube_id_list:
                    (s.user or s).delete()
                    messages.success(request, "Mağaza müdürü çıkarıldı.")
            return redirect('ekip')
        if islem in ('mutfak_sorumlusu_ekle', 'mutfak_sorumlusu_cikar') and personel.rol in (Rol.GENEL_MUDUR, Rol.OPERATOR):
            if islem == 'mutfak_sorumlusu_ekle':
                ad = request.POST.get('ms_ad_soyad', '').strip()
                if ad:
                    yeni = Personel.objects.create(ad_soyad=ad[:100], rol=Rol.MUTFAK_SORUMLUSU)
                    messages.success(request, f"{ad} mutfak sorumlusu olarak eklendi. Giriş kodu: {yeni.giris_kodu} (kodla giriş yapar).")
                else:
                    messages.error(request, "Ad soyad zorunlu.")
            else:
                s = Personel.objects.filter(id=request.POST.get('personel_id'), rol=Rol.MUTFAK_SORUMLUSU).first()
                if s:
                    ad = s.ad_soyad
                    (s.user or s).delete()
                    messages.success(request, f"{ad} mutfak sorumluluğundan çıkarıldı.")
            return redirect('ekip')
        if not is_tam:
            return redirect('ekip')

        if islem == 'yonetici_ekle':
            ad = request.POST.get('ad_soyad', '').strip()
            rol = request.POST.get('rol', '')
            if ad and rol in OFIS_ROLLERI:
                uname = _kullanici_adi_uret(ad)
                sifre = _yeni_sifre()
                u = User.objects.create(username=uname)
                u.set_password(sifre)
                u.save()
                Personel.objects.create(user=u, ad_soyad=ad, rol=rol)
                messages.success(request, f"{ad} ({rol}) eklendi. Kullanıcı adı: {uname} · Şifre: {sifre} — bu şifre yalnızca şimdi görünür, not alın.")
            else:
                messages.error(request, "Ad soyad ve geçerli bir yönetici rolü zorunlu.")
            return redirect('ekip')

        if islem == 'yonetici_sifre_yenile':
            s = Personel.objects.filter(id=request.POST.get('personel_id'), rol__in=OFIS_ROLLERI).select_related('user').first()
            if s and s.user:
                sifre = _yeni_sifre()
                s.user.set_password(sifre)
                s.user.save()
                messages.success(request, f"{s.ad_soyad} için yeni şifre: {sifre} · Kullanıcı adı: {s.user.username} — şimdi not alın.")
            return redirect('ekip')

        if islem == 'yonetici_cikar':
            s = Personel.objects.filter(id=request.POST.get('personel_id'), rol__in=OFIS_ROLLERI).select_related('user').first()
            if s and s.id == personel.id:
                messages.error(request, "Kendi hesabınızı buradan çıkaramazsınız.")
            elif s:
                (s.user or s).delete()
                messages.success(request, "Yönetici hesabı çıkarıldı.")
            return redirect('ekip')

        if islem == 'bolge_sube_ata':
            m = Personel.objects.filter(id=request.POST.get('mudur_id'), rol=Rol.MUDUR).first()
            if m:
                ids = request.POST.getlist('sube_idler')
                m.sorumlu_subeler.set(Sube.objects.filter(id__in=ids))
                messages.success(request, f"{m.ad_soyad} için sorumlu şubeler güncellendi ({m.sorumlu_subeler.count()} şube).")
            return redirect('ekip')

        if islem == 'magaza_sube_ata':
            m = Personel.objects.filter(id=request.POST.get('magaza_id'), rol=Rol.MAGAZA_MUDURU).first()
            if m:
                ids = request.POST.getlist('sube_idler')
                m.sorumlu_subeler.set(Sube.objects.filter(id__in=ids))
                if m.sorumlu_subeler.count() == 1:
                    m.sube = m.sorumlu_subeler.first()
                    m.save(update_fields=['sube'])
                messages.success(request, f"{m.ad_soyad} için sorumlu olduğu şubeler güncellendi ({m.sorumlu_subeler.count()} şube).")
            return redirect('ekip')

        if islem == 'sef_ekle':
            ad = request.POST.get('ad_soyad', '').strip()
            sid = request.POST.get('sef_sube_id')
            if ad and sid:
                yeni = Personel.objects.create(ad_soyad=ad, sube_id=sid, rol=Rol.SEF)
                messages.success(request, f"{ad} şef olarak eklendi. Giriş kodu: {yeni.giris_kodu} (şef de kodla girer).")
            else:
                messages.error(request, "Ad soyad ve şube zorunlu.")
            return redirect('ekip')

        if islem == 'sef_cikar':
            s = Personel.objects.filter(id=request.POST.get('sef_id'), rol=Rol.SEF).select_related('user').first()
            if s:
                (s.user or s).delete()
                messages.success(request, "Şef sistemden çıkarıldı.")
            return redirect('ekip')

        if islem == 'sef_degistir':
            s = Personel.objects.filter(id=request.POST.get('sef_id'), rol=Rol.SEF).first()
            if s:
                ad = request.POST.get('yeni_ad_soyad', '').strip()
                sb = request.POST.get('yeni_sube_id')
                if ad:
                    s.ad_soyad = ad
                if sb:
                    s.sube_id = sb
                s.save()
                messages.success(request, "Şef bilgileri güncellendi.")
            return redirect('ekip')

        if islem == 'egitmen_ekle':
            s = Personel.objects.filter(id=request.POST.get('personel_id'),
                                        rol__in=[Rol.PERSONEL, Rol.SEF]).first()
            if s:
                s.egitmen = True
                s.save(update_fields=['egitmen'])
                messages.success(request, f"{s.ad_soyad} artık eğitmen (kodla girip soru yönetimine erişir).")
            return redirect('ekip')

        if islem == 'egitmen_cikar':
            s = Personel.objects.filter(id=request.POST.get('personel_id')).first()
            if s:
                s.egitmen = False
                s.save(update_fields=['egitmen'])
                messages.success(request, f"{s.ad_soyad} eğitmen yetkisinden çıkarıldı.")
            return redirect('ekip')

        if islem == 'manuel_yetki_ekle':
            s = Personel.objects.filter(id=request.POST.get('personel_id'),
                                        rol__in=[Rol.PERSONEL, Rol.SEF, Rol.MAGAZA_MUDURU]).first()
            if s:
                s.manuel_giris_yetkisi = True
                s.save(update_fields=['manuel_giris_yetkisi'])
                messages.success(request, f"{s.ad_soyad} artık kamera olmadan manuel mola/mesai girişi yapabilir.")
            return redirect('ekip')

        if islem == 'manuel_yetki_cikar':
            s = Personel.objects.filter(id=request.POST.get('personel_id')).first()
            if s:
                s.manuel_giris_yetkisi = False
                s.save(update_fields=['manuel_giris_yetkisi'])
                messages.success(request, f"{s.ad_soyad} için manuel giriş yetkisi kaldırıldı.")
            return redirect('ekip')

    yoneticiler = list(Personel.objects.filter(rol__in=OFIS_ROLLERI).select_related('user').order_by('ad_soyad'))
    sefler_qs = Personel.objects.filter(rol=Rol.SEF).select_related('sube')
    if personel.rol == Rol.MUDUR:
        sefler_qs = sefler_qs.filter(sube_id__in=[s.id for s in subeler])
    sefler = list(sefler_qs.order_by('ad_soyad'))

    is_atayabilir = is_tam
    bolge_mudurleri = []
    tum_subeler = []
    egitmenler = []
    egitmen_adaylari = []
    if is_atayabilir:
        tum_subeler = list(Sube.objects.order_by('ad'))
        bolge_mudurleri = list(Personel.objects.filter(rol=Rol.MUDUR)
                               .prefetch_related('sorumlu_subeler').order_by('ad_soyad'))
        for m in bolge_mudurleri:
            m.atanan_ids = set(m.sorumlu_subeler.values_list('id', flat=True))
        egitmenler = list(Personel.objects.filter(egitmen=True)
                          .select_related('sube').order_by('ad_soyad'))
        egitmen_adaylari = list(Personel.objects.filter(rol__in=[Rol.PERSONEL, Rol.SEF], egitmen=False)
                                .select_related('sube').order_by('ad_soyad'))
        manuel_yetkili_liste = list(Personel.objects.filter(manuel_giris_yetkisi=True)
                                   .select_related('sube').order_by('ad_soyad'))
        manuel_yetki_adaylari = list(Personel.objects.filter(
            rol__in=[Rol.PERSONEL, Rol.SEF, Rol.MAGAZA_MUDURU], manuel_giris_yetkisi=False)
            .select_related('sube').order_by('ad_soyad'))
        magaza_mudurleri_liste = list(Personel.objects.filter(rol=Rol.MAGAZA_MUDURU)
                                      .select_related('sube').prefetch_related('sorumlu_subeler')
                                      .order_by('ad_soyad'))
        for m in magaza_mudurleri_liste:
            m.atanan_ids = set(m.sorumlu_subeler.values_list('id', flat=True))
    return render(request, 'ekip.html', {
        'personel': personel, 'aktif': 'ekip', 'subeler': subeler,
        'yoneticiler': yoneticiler, 'sefler': sefler,
        'yonetici_rolleri': OFIS_ROLLERI,
        'is_tam': is_tam,
        'magaza_atayabilir': personel.rol in (Rol.GENEL_MUDUR, Rol.OPERATOR),
        'mutfak_sorumlulari': list(Personel.objects.filter(rol=Rol.MUTFAK_SORUMLUSU).order_by('ad_soyad')),
        'magaza_mudurleri': magaza_mudurleri_liste if is_atayabilir else list(Personel.objects.filter(rol=Rol.MAGAZA_MUDURU, sube__in=subeler).select_related('sube').order_by('ad_soyad')),
        'is_atayabilir': is_atayabilir, 'bolge_mudurleri': bolge_mudurleri, 'tum_subeler': tum_subeler,
        'egitmenler': egitmenler, 'egitmen_adaylari': egitmen_adaylari,
        'manuel_yetkili_liste': manuel_yetkili_liste, 'manuel_yetki_adaylari': manuel_yetki_adaylari,
    })

def puantaj_excel_export(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    p = _aktif_personel(request)
    if not p or p.rol == Rol.PERSONEL:
        return redirect('ana_sayfa')
    ay_str = request.GET.get('puantaj_ay') or timezone.localdate().strftime('%Y-%m')
    ay_ilk, ay_son = _ay_araligi(ay_str)
    aralik = _gun_araligi(request, 'puantaj_bas', 'puantaj_bit')
    if aralik:
        hbas, hson, bas_str, bit_str = aralik
        manuel_ay = None
        donem_etiket = f"{bas_str} – {bit_str}"
    else:
        hbas, hson, manuel_ay = ay_ilk, ay_son, ay_ilk
        donem_etiket = ay_str
    if p.rol == Rol.SEF:
        subeler = Sube.objects.filter(id=p.sube_id) if p.sube_id else Sube.objects.none()
    else:
        sid = request.GET.get('sube_id') or request.session.get('sel_sube_id')
        izinli = _yon_subeler(p)
        izinli_ids = [s.id for s in izinli]
        if sid and (int(sid) in izinli_ids if str(sid).isdigit() else False):
            subeler = Sube.objects.filter(id=sid)
        else:
            subeler = Sube.objects.filter(id__in=izinli_ids).order_by('ad')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Puantaj"
    navy = PatternFill("solid", fgColor="162AA3")
    branch_fill = PatternFill("solid", fgColor="E1E8F0")
    head_font = Font(size=10, bold=True, color="FFFFFF")
    branch_font = Font(size=11, bold=True, color="162AA3")
    bold = Font(size=10, bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(border_style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:I1")
    ws["A1"] = f"GEEK PANEL — Puantaj ({donem_etiket})"
    ws["A1"].font = Font(size=14, bold=True, color="162AA3")
    basliklar = ["Ad Soyad", "Görev", "Çalışılan", "Eksik (Devamsız)", "Haftalık İzin", "Yıllık İzin", "Raporlu", "Hakediş", "Kaynak"]
    NKOL = len(basliklar)
    for c, t in enumerate(basliklar, 1):
        cell = ws.cell(row=3, column=c, value=t)
        cell.font = head_font
        cell.fill = navy
        cell.alignment = center
        cell.border = border
    r = 4
    for s in subeler:
        plist = Personel.objects.filter(sube=s).order_by('ad_soyad')
        if not plist.exists():
            continue
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NKOL)
        bc = ws.cell(row=r, column=1, value=f"Şube: {s.ad}")
        bc.font = branch_font
        bc.alignment = left
        for c in range(1, NKOL + 1):
            ws.cell(row=r, column=c).fill = branch_fill
            ws.cell(row=r, column=c).border = border
        r += 1
        toplam = {'calisilan': 0, 'eksik': 0, 'izinli': 0, 'yillik': 0, 'raporlu': 0, 'hakedis': 0}
        for pp in plist:
            try:
                d = _puantaj_hesapla(pp, hbas, hson, manuel_ay=manuel_ay)
            except Exception:
                d = {}
            calisilan = int(d.get('calisilan', 0) or 0)
            eksik = int(d.get('eksik', 0) or 0)
            izinli = int(d.get('izinli', 0) or 0)
            yillik = int(d.get('yillik', 0) or 0)
            raporlu = int(d.get('raporlu', 0) or 0)
            hakedis = calisilan + izinli + yillik
            ws.cell(row=r, column=1, value=pp.ad_soyad).alignment = left
            ws.cell(row=r, column=2, value=pp.rol).alignment = center
            ws.cell(row=r, column=3, value=calisilan).alignment = center
            ws.cell(row=r, column=4, value=eksik).alignment = center
            ws.cell(row=r, column=5, value=izinli).alignment = center
            ws.cell(row=r, column=6, value=yillik).alignment = center
            ws.cell(row=r, column=7, value=raporlu).alignment = center
            hc = ws.cell(row=r, column=8, value=hakedis)
            hc.alignment = center
            hc.font = bold
            ws.cell(row=r, column=9, value=("Manuel" if d.get('manuel') else "Otomatik")).alignment = center
            for c in range(1, NKOL + 1):
                ws.cell(row=r, column=c).border = border
            toplam['calisilan'] += calisilan
            toplam['eksik'] += eksik
            toplam['izinli'] += izinli
            toplam['yillik'] += yillik
            toplam['raporlu'] += raporlu
            toplam['hakedis'] += hakedis
            r += 1
        ws.cell(row=r, column=1, value=f"{s.ad} Toplam").font = bold
        for c, key in [(3, 'calisilan'), (4, 'eksik'), (5, 'izinli'), (6, 'yillik'), (7, 'raporlu'), (8, 'hakedis')]:
            tc = ws.cell(row=r, column=c, value=toplam[key])
            tc.font = bold
            tc.alignment = center
        for c in range(1, NKOL + 1):
            ws.cell(row=r, column=c).border = border
        r += 2
    for i, c in enumerate(basliklar, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(12, len(c) + 4)

    resp = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="GeekPanel_Puantaj_{ay_str}.xlsx"'
    excel_logo(ws)
    wb.save(resp)
    return resp


KALIBRASYON_GUN = 35

def _kalibrasyon_temizle():
    sinir = timezone.now() - datetime.timedelta(days=KALIBRASYON_GUN)
    for k in Kalibrasyon.objects.filter(olusturma__lt=sinir):
        if k.foto:
            k.foto.delete(save=False)
        k.delete()


def kalibrasyon_sayfa(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')

    ekleyebilir = personel.rol in (Rol.PERSONEL, Rol.SEF, Rol.MAGAZA_MUDURU)
    is_yon = personel.rol in UST_YONETIM
    subeler = _yon_subeler(personel) if is_yon else []
    sel_sube = _yonetici_sube(request, subeler) if is_yon else personel.sube
    _kalibrasyon_temizle()

    today = timezone.localdate()
    en_eski = today - datetime.timedelta(days=KALIBRASYON_GUN)
    if is_yon:
        try:
            ref = datetime.datetime.strptime(request.GET.get('kalibrasyon_tarih'), '%Y-%m-%d').date()
        except (TypeError, ValueError):
            ref = today
        ref = min(max(ref, en_eski), today)
    else:
        ref = today

    if request.method == 'POST' and ekleyebilir:
        if not sel_sube:
            messages.error(request, "Şubeniz tanımlı değil. Yöneticinize başvurun.")
            return redirect('kalibrasyon')
        if request.POST.get('islem') == 'kalibrasyon_yukle':
            data = request.POST.get('foto_data', '')
            raw = None
            if data.startswith('data:image'):
                try:
                    raw = base64.b64decode(data.split(',', 1)[1])
                except (ValueError, IndexError):
                    raw = None
            if raw and 100 < len(raw) <= 8 * 1024 * 1024:
                _kalibrasyon_temizle()
                k = Kalibrasyon(sube=sel_sube, giren=personel, giren_ad=personel.ad_soyad)
                fname = f"kal_{sel_sube.id}_{personel.id}_{timezone.now():%Y%m%d_%H%M%S}.jpg"
                k.foto.save(fname, ContentFile(raw), save=True)
                messages.success(request, "Kalibrasyon görüntüsü yüklendi.")
                _bildir(_sube_yoneticileri(sel_sube),
                        "%s şubesi kalibrasyon görüntüsü yükledi (%s)." % (sel_sube.ad, personel.ad_soyad),
                        '/kalibrasyon/', 'kalibrasyon')
            else:
                messages.error(request, "Görüntü alınamadı. Lütfen kameradan tekrar çekin.")
        return redirect('kalibrasyon')

    gorseller = []
    if sel_sube:
        gorseller = list(Kalibrasyon.objects.filter(sube=sel_sube, olusturma__date=ref).select_related('giren'))

    return render(request, 'kalibrasyon.html', {
        'personel': personel, 'aktif': 'kalibrasyon', 'ekleyebilir': ekleyebilir, 'is_yon': is_yon,
        'subeler': subeler, 'sel_sube': sel_sube, 'gorseller': gorseller,
        'secili_tarih': ref.strftime('%Y-%m-%d'), 'en_eski_tarih': en_eski.strftime('%Y-%m-%d'),
        'bugun': today.strftime('%Y-%m-%d'), 'saklama_gun': KALIBRASYON_GUN,
    })

IRSALIYE_GUN = 180

LAVABO_GUN = 2

def _lavabo_temizle():
    sinir = timezone.now() - datetime.timedelta(days=LAVABO_GUN)
    for k in LavaboDenetim.objects.filter(olusturma__lt=sinir):
        if k.foto:
            k.foto.delete(save=False)
        k.delete()


def lavabo_sayfa(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')

    ekleyebilir = personel.rol in (Rol.PERSONEL, Rol.SEF, Rol.MAGAZA_MUDURU)
    is_yon = personel.rol in UST_YONETIM
    subeler = _yon_subeler(personel) if is_yon else []
    sel_sube = _yonetici_sube(request, subeler) if is_yon else personel.sube
    _lavabo_temizle()

    if request.method == 'POST' and ekleyebilir:
        if not sel_sube:
            messages.error(request, "Şubeniz tanımlı değil. Yöneticinize başvurun.")
            return redirect('lavabo')
        if request.POST.get('islem') == 'lavabo_yukle':
            data = request.POST.get('foto_data', '')
            raw = None
            if data.startswith('data:image'):
                try:
                    raw = base64.b64decode(data.split(',', 1)[1])
                except (ValueError, IndexError):
                    raw = None
            if raw and 100 < len(raw) <= 8 * 1024 * 1024:
                _lavabo_temizle()
                k = LavaboDenetim(sube=sel_sube, giren=personel, giren_ad=personel.ad_soyad)
                fname = f"lavabo_{sel_sube.id}_{personel.id}_{timezone.now():%Y%m%d_%H%M%S}.jpg"
                k.foto.save(fname, ContentFile(raw), save=True)
                messages.success(request, "Lavabo denetim görüntüsü yüklendi.")
                _bildir(_sube_yoneticileri(sel_sube),
                        "%s şubesi lavabo denetim görüntüsü yükledi (%s)." % (sel_sube.ad, personel.ad_soyad),
                        '/lavabo/', 'lavabo')
            else:
                messages.error(request, "Görüntü alınamadı. Lütfen kameradan tekrar çekin.")
        return redirect('lavabo')

    gorseller = []
    if sel_sube:
        gorseller = list(LavaboDenetim.objects.filter(sube=sel_sube).select_related('giren')[:60])

    return render(request, 'lavabo.html', {
        'personel': personel, 'aktif': 'lavabo', 'ekleyebilir': ekleyebilir, 'is_yon': is_yon,
        'subeler': subeler, 'sel_sube': sel_sube, 'gorseller': gorseller,
        'saklama_gun': LAVABO_GUN,
    })


def lavabo_json(request):
    if not request.user.is_authenticated:
        return JsonResponse({'gorseller': []}, status=403)
    personel = _aktif_personel(request)
    if personel is None:
        return JsonResponse({'gorseller': []}, status=403)
    is_yon = personel.rol in UST_YONETIM
    if is_yon:
        subeler = _yon_subeler(personel)
        sel_sube = _yonetici_sube(request, subeler)
    else:
        sel_sube = personel.sube
    if not sel_sube:
        return JsonResponse({'gorseller': []})
    qs = LavaboDenetim.objects.filter(sube=sel_sube).select_related('giren')[:60]
    out = [{'url': g.foto.url, 'ad': g.giren_ad or '—',
            'zaman': timezone.localtime(g.olusturma).strftime('%d.%m.%Y %H:%M')} for g in qs]
    return JsonResponse({'gorseller': out})


def _irsaliye_temizle():
    sinir = timezone.now() - datetime.timedelta(days=IRSALIYE_GUN)
    for k in Irsaliye.objects.filter(olusturma__lt=sinir):
        if k.foto:
            k.foto.delete(save=False)
        k.delete()

def irsaliye_sayfa(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')

    ekleyebilir = personel.rol == Rol.SEVKIYAT
    goruntuleyebilir = personel.rol in (Rol.SATIN_ALMA, Rol.GENEL_MUDUR, Rol.MUDUR, Rol.OPERATOR, Rol.YATIRIMCI)
    if not (ekleyebilir or goruntuleyebilir):
        return redirect('ana_sayfa')

    today = timezone.localdate()
    en_eski = today - datetime.timedelta(days=IRSALIYE_GUN)
    if goruntuleyebilir:
        try:
            ref = datetime.datetime.strptime(request.GET.get('irsaliye_tarih'), '%Y-%m-%d').date()
        except (TypeError, ValueError):
            ref = today
        ref = min(max(ref, en_eski), today)
    else:
        ref = today

    if request.method == 'POST' and ekleyebilir:
        if request.POST.get('islem') == 'irsaliye_yukle':
            data = request.POST.get('foto_data', '')
            aciklama = request.POST.get('aciklama', '').strip()
            raw = None
            if data.startswith('data:image'):
                try:
                    raw = base64.b64decode(data.split(',', 1)[1])
                except (ValueError, IndexError):
                    raw = None
            if not aciklama:
                messages.error(request, "Açıklama zorunlu. Lütfen transfer bilgisini yazın.")
            elif raw and 100 < len(raw) <= 8 * 1024 * 1024:
                _irsaliye_temizle()
                k = Irsaliye(giren=personel, giren_ad=personel.ad_soyad, aciklama=aciklama)
                fname = f"irs_{personel.id}_{timezone.now():%Y%m%d_%H%M%S}.jpg"
                k.foto.save(fname, ContentFile(raw), save=True)
                messages.success(request, "İrsaliye görüntüsü yüklendi.")
            else:
                messages.error(request, "Görüntü alınamadı. Lütfen kameradan tekrar çekin.")
        return redirect('irsaliye')

    kayitlar = list(Irsaliye.objects.filter(olusturma__date=ref).select_related('giren'))
    return render(request, 'irsaliye.html', {
        'personel': personel, 'aktif': 'irsaliye', 'ekleyebilir': ekleyebilir,
        'goruntuleyebilir': goruntuleyebilir, 'kayitlar': kayitlar,
        'secili_tarih': ref.strftime('%Y-%m-%d'), 'en_eski_tarih': en_eski.strftime('%Y-%m-%d'),
        'bugun': today.strftime('%Y-%m-%d'),
    })

def stok_sayimi(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')

    is_sef = personel.rol in (Rol.SEF, Rol.MAGAZA_MUDURU)
    is_viewer = personel.rol in (Rol.MUDUR, Rol.GENEL_MUDUR, Rol.SATIN_ALMA, Rol.OPERATOR, Rol.YATIRIMCI)
    if not (is_sef or is_viewer):
        return redirect('ana_sayfa')

    ay_str = request.GET.get('stok_ay') or timezone.localdate().strftime('%Y-%m')
    ay_ilk, _ = _ay_araligi(ay_str)

    if is_sef:
        subeler = []
        sel_sube = personel.sube
    else:
        subeler = _yon_subeler(personel)
        sel_sube = _yonetici_sube(request, subeler)

    if request.method == 'POST' and is_sef and sel_sube:
        if request.POST.get('islem') == 'stok_kaydet':
            sayim, _olusan = StokSayim.objects.get_or_create(sube=sel_sube, ay=ay_ilk)
            sayim.giren = personel
            sayim.giren_ad = personel.ad_soyad
            sayim.save()
            sayim.kalemler.all().delete()
            def _say(v):
                try:
                    return Decimal((v or '').strip().replace(',', '.') or '0')
                except Exception:
                    return Decimal('0')
            for u in StokUrun.objects.filter(aktif=True):
                kap = _say(request.POST.get('kapali_%s' % u.id))
                ack = _say(request.POST.get('acik_%s' % u.id))
                note = (request.POST.get('aciklama_%s' % u.id) or '').strip()[:300]
                if kap == 0 and ack == 0 and not note:
                    continue
                StokSayimKalem.objects.create(
                    sayim=sayim, urun=u, urun_ad=u.ad, kategori=u.kategori,
                    kapali_icerik=u.kapali_icerik, acik_carpan=u.acik_carpan,
                    kapali_adet=kap, acik_miktar=ack, aciklama=note)

            ek_adlar = request.POST.getlist('ek_ad')
            ek_miktarlar = request.POST.getlist('ek_miktar')
            ek_notlar = request.POST.getlist('ek_not')
            for idx, ad in enumerate(ek_adlar):
                ad = (ad or '').strip()
                if not ad:
                    continue
                mik = _say(ek_miktarlar[idx]) if idx < len(ek_miktarlar) else Decimal('0')
                note = (ek_notlar[idx].strip()[:300] if idx < len(ek_notlar) else '')
                StokSayimKalem.objects.create(
                    sayim=sayim, urun=None, urun_ad=ad[:200], kategori='EK ÜRÜNLER',
                    kapali_icerik=1, acik_carpan=1, kapali_adet=mik, acik_miktar=0, aciklama=note)
            messages.success(request, f"{ay_ilk:%m.%Y} stok sayımı kaydedildi.")
        return redirect(f"{reverse('stok')}?stok_ay={ay_str}")

    sayim = StokSayim.objects.filter(sube=sel_sube, ay=ay_ilk).first() if sel_sube else None
    girilen = {}
    ek_kalemler = []
    if sayim:
        for k in sayim.kalemler.all():
            if k.urun_id:
                girilen[k.urun_id] = k
            else:
                ek_kalemler.append(k)

    gruplar = []
    if is_sef:
        kat_map = {}
        for u in StokUrun.objects.filter(aktif=True).order_by('sira', 'ad'):
            g = girilen.get(u.id)
            kat_map.setdefault(u.kategori or 'Diğer', []).append({
                'urun': u,
                'kapali': (g.kapali_adet if g else None),
                'acik': (g.acik_miktar if g else None),
                'aciklama': (g.aciklama if g else ''),
            })
        gruplar = [{'kategori': k, 'urunler': v} for k, v in kat_map.items()]
    elif sayim:
        kat_map = {}
        for k in sayim.kalemler.all():
            kat_map.setdefault(k.kategori or 'Diğer', []).append(k)
        gruplar = [{'kategori': k, 'urunler': v} for k, v in kat_map.items()]

    return render(request, 'stok.html', {
        'personel': personel, 'aktif': 'stok', 'is_sef': is_sef, 'is_viewer': is_viewer,
        'subeler': subeler, 'sel_sube': sel_sube, 'gruplar': gruplar, 'sayim': sayim,
        'ek_kalemler': ek_kalemler,
        'selected_ay_str': ay_str,
        'katalog_var': StokUrun.objects.filter(aktif=True).exists(),
    })

def stok_excel(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    rol = personel.rol
    if rol not in (Rol.SEF, Rol.MAGAZA_MUDURU, Rol.MUDUR, Rol.GENEL_MUDUR, Rol.SATIN_ALMA, Rol.OPERATOR, Rol.YATIRIMCI):
        return redirect('ana_sayfa')

    ay_str = request.GET.get('stok_ay') or timezone.localdate().strftime('%Y-%m')
    ay_ilk, _ = _ay_araligi(ay_str)
    sid = request.GET.get('sube_id') or request.session.get('sel_sube_id')
    if rol == Rol.SEF:
        sube = personel.sube
    else:
        sube = Sube.objects.filter(id=sid).first()
        if rol == Rol.MUDUR:
            izin_ids = [s.id for s in _yon_subeler(personel)]
            if not sube or sube.id not in izin_ids:
                return redirect('stok')
    if not sube:
        return redirect('stok')

    sayim = StokSayim.objects.filter(sube=sube, ay=ay_ilk).first()
    kalemler = list(sayim.kalemler.all()) if sayim else []

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sayım Raporu"
    navy = PatternFill("solid", fgColor="162AA3")
    gray = PatternFill("solid", fgColor="EEF1FB")
    wf = Font(bold=True, color="FFFFFF", name="Arial")
    bf = Font(bold=True, name="Arial")
    nf = Font(name="Arial")
    ctr = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center")
    thin = Side(style="thin", color="C9CEE8")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:H1")
    ws["A1"] = f"GEEK COFFEE SHOP — AYLIK SAYIM RAPORU"
    ws["A1"].font = Font(bold=True, size=13, name="Arial")
    ws["A1"].alignment = ctr
    ws["A2"] = "TARİH:"; ws["B2"] = ay_ilk.strftime("%m.%Y")
    ws["D2"] = "ŞUBE:"; ws["E2"] = sube.ad
    for c in ("A2", "D2"):
        ws[c].font = bf

    ws.merge_cells("C3:D3"); ws["C3"] = "KAPALI KUTU"
    ws.merge_cells("E3:F3"); ws["E3"] = "AÇIK KUTU"
    hdr = ["GRUP", "ÜRÜN ADI", "MİKTAR", "ADET/ML/KG", "MİKTAR", "ADET/ML/KG", "TOPLAM", "AÇIKLAMA"]
    for j, h in enumerate(hdr, 1):
        cell = ws.cell(row=4, column=j, value=h)
        cell.font = wf; cell.fill = navy; cell.alignment = ctr; cell.border = border
    for c in ("C3", "E3"):
        ws[c].font = wf; ws[c].fill = navy; ws[c].alignment = ctr
        ws[c].border = border

    girilen = {k.urun_ad: k for k in kalemler if k.urun_id}
    ekler = [k for k in kalemler if not k.urun_id]

    def _yaz(r, kategori, ad, kap, ic, ack, carp, note):
        ws.cell(row=r, column=1, value=kategori).font = nf
        ws.cell(row=r, column=2, value=ad).font = nf
        ws.cell(row=r, column=3, value=(float(kap) if kap is not None else None)).font = nf
        ws.cell(row=r, column=4, value=float(ic)).font = nf
        ws.cell(row=r, column=5, value=(float(ack) if ack is not None else None)).font = nf
        ws.cell(row=r, column=6, value=float(carp)).font = nf

        toplam = float((kap or 0)) * float(ic) + float((ack or 0)) * float(carp)
        ws.cell(row=r, column=7, value=toplam).font = bf
        ws.cell(row=r, column=8, value=note).font = nf
        for j in range(1, 9):
            ws.cell(row=r, column=j).border = border
            if j != 2 and j != 8:
                ws.cell(row=r, column=j).alignment = ctr
        if r % 2 == 0:
            for j in range(1, 9):
                ws.cell(row=r, column=j).fill = gray

    r = 5
    for u in StokUrun.objects.filter(aktif=True).order_by('sira', 'ad'):
        k = girilen.get(u.ad)
        _yaz(r, u.kategori, u.ad,
             (k.kapali_adet if k else None), u.kapali_icerik,
             (k.acik_miktar if k else None), u.acik_carpan,
             (k.aciklama if k else ""))
        r += 1
    for k in ekler:
        _yaz(r, k.kategori or "EK ÜRÜNLER", k.urun_ad,
             k.kapali_adet, k.kapali_icerik, k.acik_miktar, k.acik_carpan, k.aciklama)
        r += 1

    widths = [20, 42, 9, 12, 9, 12, 12, 28]
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + j)].width = w
    ws.freeze_panes = "A5"

    from django.http import HttpResponse
    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    fn = f"stok_{sube.ad}_{ay_str}.xlsx".replace(' ', '_')
    resp["Content-Disposition"] = f'attachment; filename="{fn}"'
    excel_logo(ws)
    wb.save(resp)
    return resp

def _katalog_gruplu():
    form_map = {}
    for u in Urun.objects.filter(aktif=True).order_by('form', 'sira', 'ad'):
        form_map.setdefault(u.form, {}).setdefault(u.kategori, []).append(u)
    sonuc = []
    for form_kod, etiket in SevkiyatForm.choices:
        if form_kod in form_map:
            kats = [{'ad': k, 'urunler': v} for k, v in form_map[form_kod].items()]
            sonuc.append({'kod': form_kod, 'etiket': etiket, 'kategoriler': kats})
    return sonuc

def _birim_secenek(urun):
    secs = [urun.birim]
    if urun.ust_birim and urun.ust_birim != urun.birim:
        secs.append(urun.ust_birim)
    return secs

SEVKIYAT_DUZENLE_ROLLERI = [Rol.SATIN_ALMA, Rol.OPERATOR, Rol.GENEL_MUDUR]

def sevkiyat_duzenle(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in SEVKIYAT_DUZENLE_ROLLERI:
        return redirect('ana_sayfa')

    birimler = [b for b, _ in SevkiyatBirim.choices]
    formlar = [(f, e) for f, e in SevkiyatForm.choices]

    if request.method == 'POST':
        islem = request.POST.get('islem')
        geri_grup = (request.POST.get('geri_grup') or '').strip()
        geri = f"{reverse('sevkiyat_duzenle')}?grup={geri_grup}" if geri_grup else reverse('sevkiyat_duzenle')

        if islem == 'urun_guncelle':
            u = Urun.objects.filter(id=request.POST.get('urun_id')).first()
            if u:
                ad = (request.POST.get('ad') or '').strip()
                try:
                    koli = max(1, int(float((request.POST.get('koli_icerigi') or '1').replace(',', '.'))))
                except Exception:
                    koli = u.koli_icerigi
                birim = request.POST.get('birim')
                if birim not in birimler:
                    birim = u.birim
                ust = (request.POST.get('ust_birim') or '').strip()
                if ust and ust not in birimler:
                    ust = u.ust_birim
                if ad:
                    u.ad = ad[:160]
                u.koli_icerigi = koli
                u.birim = birim
                u.ust_birim = ust
                u.save(update_fields=['ad', 'koli_icerigi', 'birim', 'ust_birim'])
                messages.success(request, f"{u.ad} güncellendi.")
            return redirect(geri)

        if islem == 'urun_cikar':
            u = Urun.objects.filter(id=request.POST.get('urun_id')).first()
            if u:
                u.aktif = False
                u.save(update_fields=['aktif'])
                messages.success(request, f"{u.ad} listeden çıkarıldı.")
            return redirect(geri)

        if islem == 'urun_ekle':
            ad = (request.POST.get('ad') or '').strip()
            kategori = (request.POST.get('kategori') or '').strip() or geri_grup
            form = request.POST.get('form')
            if form not in [f for f, _ in SevkiyatForm.choices]:
                form = SevkiyatForm.HAMMADDE
            try:
                koli = max(1, int(float((request.POST.get('koli_icerigi') or '1').replace(',', '.'))))
            except Exception:
                koli = 1
            birim = request.POST.get('birim')
            if birim not in birimler:
                birim = SevkiyatBirim.ADET
            ust = (request.POST.get('ust_birim') or '').strip()
            if ust and ust not in birimler:
                ust = ''
            if ad and kategori:
                var = Urun.objects.filter(form=form, ad=ad).first()
                if var:
                    var.kategori = kategori
                    var.koli_icerigi = koli
                    var.birim = birim
                    var.ust_birim = ust
                    var.aktif = True
                    var.save()
                    messages.success(request, f"{ad} güncellendi (zaten vardı, yeniden eklendi).")
                else:
                    son = Urun.objects.filter(form=form).order_by('-sira').first()
                    Urun.objects.create(form=form, kategori=kategori, ad=ad[:160],
                                        koli_icerigi=koli, birim=birim, ust_birim=ust,
                                        sira=(son.sira + 1 if son else 0), aktif=True)
                    messages.success(request, f"{ad} eklendi.")
                geri = f"{reverse('sevkiyat_duzenle')}?grup={kategori}"
            else:
                messages.error(request, "Ürün adı ve grup zorunlu.")
            return redirect(geri)

        return redirect(geri)

    aktif_urunler = list(Urun.objects.filter(aktif=True).order_by('form', 'sira', 'ad'))
    grup_sira = []
    grup_form = {}
    for u in aktif_urunler:
        if u.kategori not in grup_form:
            grup_form[u.kategori] = u.form
            grup_sira.append(u.kategori)
    sel_grup = request.GET.get('grup') or (grup_sira[0] if grup_sira else '')
    urunler = [u for u in aktif_urunler if u.kategori == sel_grup]
    sel_form = grup_form.get(sel_grup, SevkiyatForm.HAMMADDE)

    return render(request, 'sevkiyat_duzenle.html', {
        'personel': personel, 'aktif': 'sevkiyat_duzenle',
        'gruplar': grup_sira, 'sel_grup': sel_grup, 'sel_form': sel_form,
        'urunler': urunler, 'birimler': birimler, 'formlar': formlar,
    })

import calendar as _calmod
_AY_ADLARI = ['', 'Ocak', 'Şubat', 'Mart', 'Nisan', 'Mayıs', 'Haziran',
              'Temmuz', 'Ağustos', 'Eylül', 'Ekim', 'Kasım', 'Aralık']
_GUN_KISA = ['Pzt', 'Sal', 'Çar', 'Per', 'Cum', 'Cmt', 'Paz']

def _takvim_kur(yil, ay, sel_gun, sayac):
    cal = _calmod.Calendar(firstweekday=0)
    bugun = timezone.localdate()
    haftalar = []
    for week in cal.monthdatescalendar(yil, ay):
        satir = []
        for d in week:
            satir.append({
                'no': d.day, 'gun': d.isoformat(), 'bu_ay': (d.month == ay),
                'sayi': sayac.get(d, 0), 'secili': (sel_gun == d), 'bugun': (d == bugun),
            })
        haftalar.append(satir)
    onceki = datetime.date(yil, ay, 1) - datetime.timedelta(days=1)
    sonraki = (datetime.date(yil, ay, 28) + datetime.timedelta(days=10)).replace(day=1)
    return {
        'yil': yil, 'ay': ay, 'ay_adi': _AY_ADLARI[ay], 'gun_basliklari': _GUN_KISA,
        'haftalar': haftalar,
        'prev': '%04d-%02d' % (onceki.year, onceki.month),
        'next': '%04d-%02d' % (sonraki.year, sonraki.month),
    }

def _gecmis_hazirla(request, sel_id, izin_ids=None):
    bugun = timezone.localdate()
    yil, ay = bugun.year, bugun.month
    sel_gun = None
    gun_str = request.GET.get('gun')
    if gun_str:
        try:
            sel_gun = datetime.date.fromisoformat(gun_str)
            yil, ay = sel_gun.year, sel_gun.month
        except ValueError:
            sel_gun = None
    if sel_gun is None:
        ay_str = request.GET.get('ay')
        if ay_str:
            try:
                yil, ay = (int(x) for x in ay_str.split('-'))
            except (ValueError, TypeError):
                yil, ay = bugun.year, bugun.month
    ay_basi = datetime.date(yil, ay, 1)
    son_gun = _calmod.monthrange(yil, ay)[1]
    ay_sonu = datetime.date(yil, ay, son_gun)

    qs = SevkiyatTalep.objects.filter(olusturma__date__gte=ay_basi, olusturma__date__lte=ay_sonu)
    if izin_ids is not None:
        qs = qs.filter(sube_id__in=izin_ids)
    if sel_id:
        qs = qs.filter(sube_id=sel_id)
    ay_listesi = list(qs.select_related('sube').prefetch_related('kalemler').order_by('-olusturma')[:500])

    sayac = {}
    for t in ay_listesi:
        g = timezone.localtime(t.olusturma).date()
        sayac[g] = sayac.get(g, 0) + 1

    if sel_gun and sel_id:
        liste = [t for t in ay_listesi if timezone.localtime(t.olusturma).date() == sel_gun]
    else:

        liste = []
    return _takvim_kur(yil, ay, sel_gun, sayac), liste, sel_gun

def sevkiyat_sayfa(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    rol = personel.rol
    is_sef = rol in (Rol.SEF, Rol.MAGAZA_MUDURU)
    is_satinalma = rol == Rol.SATIN_ALMA
    is_sevkiyat = rol == Rol.SEVKIYAT
    is_yon = rol in UST_YONETIM
    cikis_yetkili = (rol in (Rol.GENEL_MUDUR, Rol.YATIRIMCI)) or is_satinalma
    if not (is_sef or is_satinalma or is_sevkiyat or is_yon):
        return redirect('ana_sayfa')

    sef_sube = personel.sube
    sef_coklu_sube = None
    if is_sef and rol == Rol.MAGAZA_MUDURU and personel.sorumlu_subeler.exists():
        sef_coklu_sube = list(personel.sorumlu_subeler.order_by('ad'))
        sef_sube = _yonetici_sube(request, sef_coklu_sube)

    if request.method == 'POST' and is_sef and request.POST.get('islem') == 'siparis_olustur':
        if not sef_sube:
            messages.error(request, "Şubeniz tanımlı değil. Yöneticinize başvurun.")
            return redirect('sevkiyat')
        not_metni = request.POST.get('not_metni', '').strip()[:400]
        secilen = []
        for u in Urun.objects.filter(aktif=True):
            raw = request.POST.get('miktar_%s' % u.id, '').strip().replace(',', '.')
            if not raw:
                continue
            try:
                miktar = Decimal(raw)
            except Exception:
                continue
            if miktar <= 0:
                continue
            birim = request.POST.get('birim_%s' % u.id, u.birim)
            if birim not in _birim_secenek(u):
                birim = u.birim
            secilen.append((u, miktar, birim))

        ek_urunler = request.POST.getlist('ek_urun')
        ek_miktarlar = request.POST.getlist('ek_miktar')
        ek_birimler = request.POST.getlist('ek_birim')
        gecerli_birimler = [b for b, _ in SevkiyatBirim.choices]
        ekstra = []
        for ad, mik, bir in zip(ek_urunler, ek_miktarlar, ek_birimler):
            ad = ad.strip()[:160]
            if not ad:
                continue
            try:
                miktar = Decimal((mik or '').strip().replace(',', '.'))
            except Exception:
                continue
            if miktar <= 0:
                continue
            bir = bir if bir in gecerli_birimler else SevkiyatBirim.ADET
            ekstra.append((ad, miktar, bir))
        if secilen or ekstra:
            # Çift tıklama / yavaş bağlantı yüzünden aynı siparişin birden fazla
            # oluşturulmasını önlemek için: aynı kişi aynı şube için son 20 saniye
            # içinde zaten bir sipariş oluşturduysa, yeni bir kayıt AÇMA.
            yakin_zamanda = SevkiyatTalep.objects.filter(
                sube=sef_sube, olusturan=personel,
                olusturma__gte=timezone.now() - datetime.timedelta(seconds=20)
            ).order_by('-olusturma').first()
            if yakin_zamanda:
                messages.info(request, "Bu sipariş az önce zaten gönderildi (#%s) — tekrar oluşturulmadı." % yakin_zamanda.id)
                return redirect('sevkiyat')
            talep = SevkiyatTalep.objects.create(
                sube=sef_sube, olusturan=personel,
                olusturan_ad=personel.ad_soyad, not_metni=not_metni)
            for u, miktar, birim in secilen:
                SevkiyatKalem.objects.create(
                    talep=talep, urun=u, urun_ad=u.ad, kategori=u.kategori, form=u.form,
                    koli_icerigi=u.koli_icerigi, istenen_miktar=miktar, istenen_birim=birim)
            for ad, miktar, bir in ekstra:
                SevkiyatKalem.objects.create(
                    talep=talep, urun=None, urun_ad=ad, kategori='DİĞER', form='',
                    koli_icerigi=1, istenen_miktar=miktar, istenen_birim=bir)
            SiparisHareket.objects.create(talep=talep, mesaj="Sipariş oluşturuldu",
                                          yapan_ad=personel.ad_soyad)
            messages.success(request, "Sipariş oluşturuldu (#%s, %s kalem)." % (
                talep.id, len(secilen) + len(ekstra)))
            _bildir(_rol_personelleri(Rol.SATIN_ALMA),
                    "Yeni sevkiyat talebi: %s" % talep.sube.ad, '/sevkiyat/', 'sevkiyat')
        else:
            messages.error(request, "En az bir ürüne miktar girin.")
        return redirect('sevkiyat')

    if request.method == 'POST' and is_satinalma and request.POST.get('islem') == 'talep_iptal':
        talep = SevkiyatTalep.objects.filter(id=request.POST.get('talep_id'), durum=SevkiyatDurumu.TALEP).first()
        if talep:
            sube_adi = talep.sube.ad
            olusturan = talep.olusturan
            talep.delete()
            messages.success(request, "#%s numaralı sipariş iptal edildi." % request.POST.get('talep_id'))
            if olusturan:
                _bildir([olusturan], "%s şubesi için oluşturduğun sipariş satın alma tarafından iptal edildi."
                        % sube_adi, '/sevkiyat/', 'sevkiyat')
        else:
            messages.error(request, "Bu sipariş artık iptal edilemez (satın alma zaten işleme almış olabilir).")
        return redirect('sevkiyat')

    if request.method == 'POST' and is_satinalma and request.POST.get('islem') == 'satinalma_sevkiyat_olustur':
        sube = Sube.objects.filter(id=request.POST.get('sube_id'), depo_mu=False).first()
        if not sube:
            messages.error(request, "Geçerli bir şube seçmelisiniz.")
            return redirect('sevkiyat')
        not_metni = request.POST.get('not_metni', '').strip()[:400]
        secilen = []
        for u in Urun.objects.filter(aktif=True):
            raw = request.POST.get('miktar_%s' % u.id, '').strip().replace(',', '.')
            if not raw:
                continue
            try:
                miktar = Decimal(raw)
            except Exception:
                continue
            if miktar <= 0:
                continue
            birim = request.POST.get('birim_%s' % u.id, u.birim)
            if birim not in _birim_secenek(u):
                birim = u.birim
            secilen.append((u, miktar, birim))

        ek_urunler = request.POST.getlist('ek_urun')
        ek_miktarlar = request.POST.getlist('ek_miktar')
        ek_birimler = request.POST.getlist('ek_birim')
        gecerli_birimler = [b for b, _ in SevkiyatBirim.choices]
        ekstra = []
        for ad, mik, bir in zip(ek_urunler, ek_miktarlar, ek_birimler):
            ad = ad.strip()[:160]
            if not ad:
                continue
            try:
                miktar = Decimal((mik or '').strip().replace(',', '.'))
            except Exception:
                continue
            if miktar <= 0:
                continue
            bir = bir if bir in gecerli_birimler else SevkiyatBirim.ADET
            ekstra.append((ad, miktar, bir))

        if secilen or ekstra:
            yakin_zamanda = SevkiyatTalep.objects.filter(
                sube=sube, olusturan=personel,
                olusturma__gte=timezone.now() - datetime.timedelta(seconds=20)
            ).order_by('-olusturma').first()
            if yakin_zamanda:
                messages.info(request, "Bu sipariş az önce zaten oluşturuldu (#%s) — tekrar oluşturulmadı." % yakin_zamanda.id)
                return redirect('sevkiyat')
            talep = SevkiyatTalep.objects.create(
                sube=sube, olusturan=personel, olusturan_ad=personel.ad_soyad, not_metni=not_metni,
                durum=SevkiyatDurumu.SEVKIYATTA, satin_alan_ad=personel.ad_soyad, satin_alma_tarih=timezone.now())
            for u, miktar, birim in secilen:
                SevkiyatKalem.objects.create(
                    talep=talep, urun=u, urun_ad=u.ad, kategori=u.kategori, form=u.form,
                    koli_icerigi=u.koli_icerigi, istenen_miktar=miktar, istenen_birim=birim,
                    satinalma_miktar=miktar, satinalma_birim=birim)
            for ad, miktar, bir in ekstra:
                SevkiyatKalem.objects.create(
                    talep=talep, urun=None, urun_ad=ad, kategori='DİĞER', form='',
                    koli_icerigi=1, istenen_miktar=0, istenen_birim=bir,
                    satinalma_miktar=miktar, satinalma_birim=bir)
            SiparisHareket.objects.create(talep=talep, mesaj="Satın alma tarafından oluşturuldu, sevkiyata iletildi",
                                          yapan_ad=personel.ad_soyad)
            messages.success(request, "Sipariş oluşturuldu ve sevkiyata iletildi (#%s, %s kalem)." % (
                talep.id, len(secilen) + len(ekstra)))
            _bildir(_rol_personelleri(Rol.SEVKIYAT),
                    "Sevkiyata hazır: %s" % talep.sube.ad, '/sevkiyat/', 'sevkiyat')
        else:
            messages.error(request, "En az bir ürüne miktar girin.")
        return redirect('sevkiyat')

    if request.method == 'POST' and is_satinalma and request.POST.get('islem') == 'satinalma_tamamla':
        talep = (SevkiyatTalep.objects.filter(id=request.POST.get('talep_id'), durum=SevkiyatDurumu.TALEP)
                 .prefetch_related('kalemler').first())
        if talep:
            gecerli = [b for b, _ in SevkiyatBirim.choices]
            for k in talep.kalemler.all():
                raw = request.POST.get('sa_miktar_%s' % k.id, '').strip().replace(',', '.')
                try:
                    miktar = Decimal(raw)
                except Exception:
                    miktar = k.istenen_miktar
                if miktar < 0:
                    miktar = Decimal(0)
                birim = request.POST.get('sa_birim_%s' % k.id, k.istenen_birim)
                if birim not in gecerli:
                    birim = k.istenen_birim
                k.satinalma_miktar = miktar
                k.satinalma_birim = birim
                k.save()
            ek_adlar = request.POST.getlist('sa_ek_ad')
            ek_miktarlar = request.POST.getlist('sa_ek_miktar')
            ek_birimler = request.POST.getlist('sa_ek_birim')
            for i, ad in enumerate(ek_adlar):
                ad = (ad or '').strip()
                if not ad:
                    continue
                raw = (ek_miktarlar[i] if i < len(ek_miktarlar) else '').strip().replace(',', '.')
                try:
                    mik = Decimal(raw)
                except Exception:
                    continue
                if mik <= 0:
                    continue
                bir = ek_birimler[i] if i < len(ek_birimler) else SevkiyatBirim.ADET
                if bir not in gecerli:
                    bir = SevkiyatBirim.ADET
                SevkiyatKalem.objects.create(
                    talep=talep, urun=None, urun_ad=ad[:160], kategori='DİĞER', form='',
                    koli_icerigi=1, istenen_miktar=0, istenen_birim=bir,
                    satinalma_miktar=mik, satinalma_birim=bir)
            talep.durum = SevkiyatDurumu.SEVKIYATTA
            talep.satin_alan_ad = personel.ad_soyad
            talep.satin_alma_tarih = timezone.now()
            talep.save()
            SiparisHareket.objects.create(talep=talep, mesaj="Satın alma tamamlandı, depoya iletildi",
                                          yapan_ad=personel.ad_soyad)
            messages.success(request, "#%s sevkiyata iletildi." % talep.id)
            _bildir(_rol_personelleri(Rol.SEVKIYAT),
                    "Sevkiyata hazır: %s" % talep.sube.ad, '/sevkiyat/', 'sevkiyat')
        return redirect('sevkiyat')

    if request.method == 'POST' and is_sevkiyat and request.POST.get('islem') == 'sevkiyat_onayla':
        talep = (SevkiyatTalep.objects.filter(id=request.POST.get('talep_id'),
                 durum__in=[SevkiyatDurumu.SEVKIYATTA, SevkiyatDurumu.REDDEDILDI])
                 .prefetch_related('kalemler').first())
        if talep:
            gecerli = [b for b, _ in SevkiyatBirim.choices]
            for k in talep.kalemler.all():
                varsayilan = k.satinalma_miktar if k.satinalma_miktar is not None else k.istenen_miktar
                raw = request.POST.get('sv_miktar_%s' % k.id, '').strip().replace(',', '.')
                try:
                    miktar = Decimal(raw)
                except Exception:
                    miktar = varsayilan
                if miktar < 0:
                    miktar = Decimal(0)
                birim = request.POST.get('sv_birim_%s' % k.id) or (k.satinalma_birim or k.istenen_birim)
                if birim not in gecerli:
                    birim = k.satinalma_birim or k.istenen_birim
                k.sevkiyat_miktar = miktar
                k.sevkiyat_birim = birim
                k.save()
            ek_adlar = request.POST.getlist('sv_ek_ad')
            ek_miktarlar = request.POST.getlist('sv_ek_miktar')
            ek_birimler = request.POST.getlist('sv_ek_birim')
            for i, ad in enumerate(ek_adlar):
                ad = (ad or '').strip()
                if not ad:
                    continue
                raw = (ek_miktarlar[i] if i < len(ek_miktarlar) else '').strip().replace(',', '.')
                try:
                    mik = Decimal(raw)
                except Exception:
                    continue
                if mik <= 0:
                    continue
                bir = ek_birimler[i] if i < len(ek_birimler) else SevkiyatBirim.ADET
                if bir not in gecerli:
                    bir = SevkiyatBirim.ADET
                SevkiyatKalem.objects.create(
                    talep=talep, urun=None, urun_ad=ad[:160], kategori='DİĞER', form='',
                    koli_icerigi=1, istenen_miktar=0, istenen_birim=bir,
                    sevkiyat_miktar=mik, sevkiyat_birim=bir)
            talep.durum = SevkiyatDurumu.ONAY_BEKLIYOR
            talep.sevkiyatci_ad = personel.ad_soyad
            talep.sevkiyat_tarih = timezone.now()
            talep.red_notu = ''
            talep.save()
            SiparisHareket.objects.create(talep=talep, mesaj="Sevkiyat hazırlandı, çıkış onayına gönderildi",
                                          yapan_ad=personel.ad_soyad)
            messages.success(request, "#%s çıkış onayına gönderildi." % talep.id)
            _bildir(_rol_personelleri(Rol.GENEL_MUDUR, Rol.YATIRIMCI),
                    "Sevkiyat çıkış onayı bekliyor: %s" % talep.sube.ad, '/sevkiyat/', 'sevkiyat')
        return redirect('sevkiyat')

    if request.method == 'POST' and cikis_yetkili and request.POST.get('islem') == 'cikis_onayla':
        talep = SevkiyatTalep.objects.filter(id=request.POST.get('talep_id'),
                                             durum=SevkiyatDurumu.ONAY_BEKLIYOR).first()
        if talep:
            talep.durum = SevkiyatDurumu.ONAYLANDI
            talep.onaylayan_ad = personel.ad_soyad
            talep.onay_tarih = timezone.now()
            talep.save()
            SiparisHareket.objects.create(talep=talep, mesaj="Çıkış onaylandı", yapan_ad=personel.ad_soyad)
            _bildir(_sube_sefleri(talep.sube),
                    "Sevkiyatınız onaylandı: %s" % talep.sube.ad, '/sevkiyat/', 'sevkiyat')
            messages.success(request, "#%s onaylandı." % talep.id)
        return redirect('sevkiyat')

    if request.method == 'POST' and cikis_yetkili and request.POST.get('islem') == 'cikis_reddet':
        talep = SevkiyatTalep.objects.filter(id=request.POST.get('talep_id'),
                                             durum=SevkiyatDurumu.ONAY_BEKLIYOR).first()
        if talep:
            aciklama = request.POST.get('red_notu', '').strip()[:400]
            talep.durum = SevkiyatDurumu.REDDEDILDI
            talep.red_notu = aciklama
            talep.save()
            SiparisHareket.objects.create(talep=talep, mesaj="Çıkış reddedildi", aciklama=aciklama,
                                          yapan_ad=personel.ad_soyad)
            _bildir(_sube_sefleri(talep.sube),
                    "Sevkiyatınız reddedildi: %s" % talep.sube.ad, '/sevkiyat/', 'sevkiyat')
            messages.success(request, "#%s reddedildi, sevkiyata geri gönderildi." % talep.id)
        return redirect('sevkiyat')

    ctx = {
        'personel': personel, 'aktif': 'sevkiyat',
        'is_sef': is_sef, 'is_satinalma': is_satinalma, 'is_sevkiyat': is_sevkiyat, 'is_yon': is_yon,
        'belge_yetkili': (not is_sef) and rol != Rol.MUDUR,
    }

    ctx['cikis_yetkili'] = cikis_yetkili
    ctx['tum_birimler'] = [b for b, _ in SevkiyatBirim.choices]

    if is_sef:
        katalog = _katalog_gruplu()
        for f in katalog:
            for kat in f['kategoriler']:
                kat['urunler'] = [{'u': u, 'birimler': _birim_secenek(u)} for u in kat['urunler']]
        ctx['katalog'] = katalog
        ctx['ek_oneri'] = list(SevkiyatKalem.objects.filter(urun__isnull=True)
                               .values_list('urun_ad', flat=True).distinct().order_by('urun_ad')[:200])
        sefler = list(SevkiyatTalep.objects.filter(sube=sef_sube)
                      .prefetch_related('kalemler', 'hareketler')[:50])
        for t in sefler:
            t.mode = 'read'
        ctx['talepler'] = sefler
        ctx['sef_coklu_sube'] = sef_coklu_sube
        ctx['sef_sube'] = sef_sube
    elif is_satinalma:
        subeler = _yon_subeler(personel)
        sel_id = request.GET.get('sube')
        try:
            sel_id = int(sel_id) if sel_id else None
        except (TypeError, ValueError):
            sel_id = None
        bekleyen = SevkiyatTalep.objects.filter(durum=SevkiyatDurumu.TALEP)
        if sel_id:
            bekleyen = bekleyen.filter(sube_id=sel_id)
        bekleyen = list(bekleyen.select_related('sube').prefetch_related('kalemler')[:100])
        for t in bekleyen:
            t.mode = 'sa'
        onaylar = list(SevkiyatTalep.objects.filter(durum=SevkiyatDurumu.ONAY_BEKLIYOR)
                       .select_related('sube').prefetch_related('kalemler')[:100])
        for t in onaylar:
            t.mode = 'cikis'
        ctx['talepler'] = onaylar + bekleyen
        takvim, gecmis, sel_gun = _gecmis_hazirla(request, sel_id)
        for t in gecmis:
            t.mode = 'read'
        ctx['subeler'] = subeler
        ctx['sel_id'] = sel_id
        ctx['takvim'] = takvim
        ctx['gecmis_talepler'] = gecmis
        ctx['sel_gun'] = sel_gun
        ctx['gecmis'] = True
        katalog = _katalog_gruplu()
        for f in katalog:
            for kat in f['kategoriler']:
                kat['urunler'] = [{'u': u, 'birimler': _birim_secenek(u)} for u in kat['urunler']]
        ctx['katalog'] = katalog
        ctx['ek_oneri'] = list(SevkiyatKalem.objects.filter(urun__isnull=True)
                               .values_list('urun_ad', flat=True).distinct().order_by('urun_ad')[:200])
    elif is_sevkiyat:
        sel_id = request.GET.get('sube')
        try:
            sel_id = int(sel_id) if sel_id else None
        except (TypeError, ValueError):
            sel_id = None
        svt_q = SevkiyatTalep.objects.filter(durum__in=[SevkiyatDurumu.SEVKIYATTA, SevkiyatDurumu.REDDEDILDI])
        if sel_id:
            svt_q = svt_q.filter(sube_id=sel_id)
        svt = list(svt_q.select_related('sube').prefetch_related('kalemler')[:100])
        gosterilecek = []
        for t in svt:
            t.mode = 'sv'
            kalemler = []
            for k in t.kalemler.all():
                if k.satinalma_miktar == 0:
                    continue
                k.sv_def_miktar = (k.sevkiyat_miktar if k.sevkiyat_miktar is not None
                                   else (k.satinalma_miktar if k.satinalma_miktar is not None else k.istenen_miktar))
                k.sv_def_birim = k.sevkiyat_birim or k.satinalma_birim or k.istenen_birim
                kalemler.append(k)
            t.sv_kalemler = kalemler
            if kalemler:
                gosterilecek.append(t)
        ctx['talepler'] = gosterilecek
        ctx['subeler'] = list(Sube.objects.filter(depo_mu=False).order_by('ad'))
        ctx['sel_id'] = sel_id
    else:
        subeler = _yon_subeler(personel)
        izin_ids = [s.id for s in subeler] if personel.rol == Rol.MUDUR else None
        sel_id = request.GET.get('sube')
        try:
            sel_id = int(sel_id) if sel_id else None
        except (TypeError, ValueError):
            sel_id = None
        if izin_ids is not None and sel_id not in izin_ids:
            sel_id = None
        aktif = []
        if cikis_yetkili:
            aq = SevkiyatTalep.objects.filter(durum=SevkiyatDurumu.ONAY_BEKLIYOR)
            if sel_id:
                aq = aq.filter(sube_id=sel_id)
            aktif = list(aq.select_related('sube').prefetch_related('kalemler')[:100])
            for t in aktif:
                t.mode = 'cikis'
        ctx['talepler'] = aktif
        takvim, gecmis, sel_gun = _gecmis_hazirla(request, sel_id, izin_ids=izin_ids)
        for t in gecmis:
            t.mode = 'read'
        ctx['subeler'] = subeler
        ctx['sel_id'] = sel_id
        ctx['takvim'] = takvim
        ctx['gecmis_talepler'] = gecmis
        ctx['sel_gun'] = sel_gun
        ctx['gecmis'] = True

    return render(request, 'sevkiyat.html', ctx)

def sevkiyat_belge(request, talep_id, tip):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    if personel.rol in (Rol.SEF, Rol.MAGAZA_MUDURU, Rol.MUDUR):
        return redirect('sevkiyat')
    if tip not in ('yukleme', 'fis'):
        return redirect('sevkiyat')
    talep = (SevkiyatTalep.objects.filter(id=talep_id)
             .select_related('sube').prefetch_related('kalemler').first())
    if talep is None:
        return redirect('sevkiyat')
    rol = personel.rol
    yetkili = rol in OFIS_ROLLERI or (rol in (Rol.SEF, Rol.MAGAZA_MUDURU) and personel.sube_id == talep.sube_id)
    if not yetkili:
        return redirect('ana_sayfa')
    if tip == 'yukleme' and talep.durum not in (SevkiyatDurumu.SEVKIYATTA,
                                                SevkiyatDurumu.ONAY_BEKLIYOR, SevkiyatDurumu.ONAYLANDI):
        return redirect('sevkiyat')
    if tip == 'fis' and talep.durum != SevkiyatDurumu.ONAYLANDI:
        return redirect('sevkiyat')
    from .sevkiyat_pdf import sevkiyat_pdf_bytes
    pdf = sevkiyat_pdf_bytes(talep, tip)
    adi = 'yukleme_belgesi' if tip == 'yukleme' else 'teslim_fisi'
    resp = HttpResponse(pdf, content_type='application/pdf')
    resp['Content-Disposition'] = 'inline; filename="sevkiyat_%s_%s.pdf"' % (talep.id, adi)
    return resp

def sevkiyat_excel(request, talep_id):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    if personel.rol in (Rol.SEF, Rol.MAGAZA_MUDURU, Rol.MUDUR):
        return redirect('sevkiyat')
    talep = (SevkiyatTalep.objects.filter(id=talep_id, durum=SevkiyatDurumu.ONAYLANDI)
             .select_related('sube').prefetch_related('kalemler').first())
    if talep is None:
        return redirect('sevkiyat')
    if personel.rol not in OFIS_ROLLERI:
        return redirect('ana_sayfa')
    from .sevkiyat_excel import siparis_excel_bytes
    data = siparis_excel_bytes(talep)
    resp = HttpResponse(data, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="siparis_%s_%s.xlsx"' % (
        talep.id, talep.olusturma.strftime('%Y%m%d'))
    return resp

def _kod_giris(request):
    ip = _istemci_ip(request)
    kilit, _ = KodKilit.objects.get_or_create(ip=ip)
    now = timezone.now()
    if kilit.kilit_bitis and now < kilit.kilit_bitis:
        kalan = int((kilit.kilit_bitis - now).total_seconds() // 60) + 1
        messages.error(request, f"Çok fazla hatalı deneme. Lütfen {kalan} dakika sonra tekrar deneyin.")
        return redirect('ana_sayfa')
    kod = request.POST.get('kod', '').strip()
    personel = Personel.objects.filter(giris_kodu=kod, rol__in=[Rol.PERSONEL, Rol.SEF, Rol.MAGAZA_MUDURU, Rol.MUTFAK_PERSONEL, Rol.MUTFAK_SORUMLUSU]).select_related('user').first()
    if personel and personel.user:
        kilit.hatali_deneme = 0
        kilit.kilit_bitis = None
        kilit.save()
        personel.user.backend = MODEL_BACKEND
        auth_login(request, personel.user)
        return redirect('ana_sayfa')
    kilit.hatali_deneme += 1
    if kilit.hatali_deneme >= MAX_DENEME:
        kilit.kilit_bitis = now + datetime.timedelta(minutes=KILIT_DK)
        kilit.hatali_deneme = 0
        messages.error(request, f"Çok fazla hatalı deneme. {KILIT_DK} dakika boyunca giriş kapatıldı.")
    else:
        messages.error(request, f"Kod hatalı. Kalan deneme hakkı: {MAX_DENEME - kilit.hatali_deneme}.")
    kilit.save()
    return redirect('ana_sayfa')

def _sifre_giris(request):
    user = authenticate(request, username=request.POST.get('kullanici_adi', '').strip(),
                        password=request.POST.get('sifre', ''))
    if user is not None:
        auth_login(request, user)
        return redirect('ana_sayfa')
    messages.error(request, 'Kullanıcı adı veya şifre hatalı.')
    return redirect('/?mod=yonetici')

def _hukuki(request, anahtar):
    return render(request, 'hukuki.html', HUKUKI_SAYFALAR[anahtar])

def kvkk(request):
    return _hukuki(request, 'kvkk')

def kullanim_kosullari(request):
    return _hukuki(request, 'kullanim_kosullari')

def gizlilik(request):
    return _hukuki(request, 'gizlilik')

_PWA_MANIFEST = {
    "name": "Geek Coffee & Eatery Panel",
    "short_name": "Geek Panel",
    "description": "Geek Coffee & Eatery personel ve sevkiyat yönetim paneli",
    "lang": "tr",
    "dir": "ltr",
    "start_url": "/",
    "scope": "/",
    "display": "standalone",
    "orientation": "portrait-primary",
    "background_color": "#ffffff",
    "theme_color": "#162AA3",
    "icons": [
        {"src": "/icons/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any"},
        {"src": "/icons/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any"},
        {"src": "/icons/icon-512-maskable.png", "sizes": "512x512", "type": "image/png", "purpose": "maskable"},
    ],
}

_PWA_SW = """
const STATIK = 'geek-statik-v3';
const KABUK = 'geek-kabuk-v3';
const KABUK_URL = '/';
self.addEventListener('install', function (e) {
  e.waitUntil(
    caches.open(KABUK).then(function (c) { return c.add(KABUK_URL); }).catch(function () {})
  );
  self.skipWaiting();
});
self.addEventListener('activate', function (e) {
  e.waitUntil(
    caches.keys().then(function (ks) {
      return Promise.all(ks.filter(function (k) { return k !== STATIK && k !== KABUK; })
                           .map(function (k) { return caches.delete(k); }));
    }).then(function () { return self.clients.claim(); })
  );
});
self.addEventListener('fetch', function (e) {
  var req = e.request;
  if (req.method !== 'GET') return;
  var url = new URL(req.url);
  if (url.origin !== location.origin) return;
  // Statik dosyalar: önce önbellek (hızlı + çevrimdışı çalışır).
  if (url.pathname.indexOf('/static/') === 0) {
    e.respondWith(
      caches.open(STATIK).then(function (c) {
        return c.match(req).then(function (hit) {
          return hit || fetch(req).then(function (res) {
            if (res && res.status === 200) c.put(req, res.clone());
            return res;
          });
        });
      })
    );
    return;
  }
  // Sayfa gezinmeleri: önce ağ (her zaman güncel), çevrimdışıysa kabuk.
  if (req.mode === 'navigate') {
    e.respondWith(
      fetch(req).catch(function () { return caches.match(KABUK_URL); })
    );
    return;
  }
});
self.addEventListener('push', function (e) {
  var d = {};
  try { d = e.data ? e.data.json() : {}; } catch (x) { d = {}; }
  var baslik = d.baslik || 'Geek Panel';
  var govde = d.mesaj || '';
  var link = d.link || '/bildirimler/';
  e.waitUntil(self.registration.showNotification(baslik, {
    body: govde, icon: '/icons/icon-192.png', badge: '/icons/icon-192.png',
    data: { link: link }, vibrate: [80, 40, 80]
  }));
});
self.addEventListener('notificationclick', function (e) {
  e.notification.close();
  var link = (e.notification.data && e.notification.data.link) || '/bildirimler/';
  e.waitUntil(clients.matchAll({ type: 'window', includeUncontrolled: true }).then(function (cs) {
    for (var i = 0; i < cs.length; i++) {
      if ('focus' in cs[i]) { try { cs[i].navigate(link); } catch (x) {} return cs[i].focus(); }
    }
    if (clients.openWindow) return clients.openWindow(link);
  }));
});
"""

def csrf_hata_sayfasi(request, reason=""):
    """
    Django'nun çıplak 'Forbidden (403) CSRF verification failed' sayfası yerine
    gösterilir. En sık sebep: eski bir sekme/PWA penceresinin uzun süredir açık
    kalıp bayat bir sayfa üzerinden form göndermesi (özellikle sunucu değişimi/
    deploy sonrası). Kullanıcıya net bir Türkçe talimat verir.
    """
    html = """<!DOCTYPE html><html lang="tr"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Sayfa yenilenmeli · Geek Panel</title>
<style>
  body{ font-family:-apple-system,Segoe UI,Roboto,sans-serif; background:#F5F7FC; color:#1c2333;
        display:flex; align-items:center; justify-content:center; min-height:100vh; margin:0; padding:1.5rem; }
  .kart{ max-width:420px; background:#fff; border-radius:16px; padding:2rem 1.6rem; text-align:center;
         box-shadow:0 12px 32px -18px rgba(20,30,60,.25); }
  h1{ font-size:1.25rem; margin:0 0 .6rem; }
  p{ font-size:.92rem; color:#5b6472; line-height:1.5; margin:0 0 1.4rem; }
  button{ background:#162AA3; color:#fff; border:none; border-radius:10px; padding:.8rem 1.4rem;
          font-size:.95rem; font-weight:600; cursor:pointer; width:100%; }
</style></head>
<body>
  <div class="kart">
    <h1>🔄 Sayfanın yenilenmesi gerekiyor</h1>
    <p>Bu sekme/uygulama uzun süredir açık kalmış olabilir. Devam etmek için sayfayı bir kez yenile.</p>
    <button onclick="window.location.href='/'">Sayfayı Yenile</button>
  </div>
  <script>
    if ('caches' in window) { caches.keys().then(function (ks) { ks.forEach(function (k) { caches.delete(k); }); }); }
  </script>
</body></html>"""
    from django.http import HttpResponseForbidden
    return HttpResponseForbidden(html)


def pwa_manifest(request):
    return HttpResponse(json.dumps(_PWA_MANIFEST, ensure_ascii=False),
                        content_type='application/manifest+json')

_PWA_IKON_IZIN = {'icon-192.png', 'icon-512.png', 'icon-512-maskable.png', 'icon-180.png'}

def pwa_icon(request, ad):
    import os
    from django.conf import settings
    from django.http import Http404
    if ad not in _PWA_IKON_IZIN:
        raise Http404("ikon yok")
    yol = os.path.join(settings.BASE_DIR, 'static', 'icons', ad)
    try:
        with open(yol, 'rb') as f:
            data = f.read()
    except OSError:
        raise Http404("ikon dosyası bulunamadı")
    resp = HttpResponse(data, content_type='image/png')
    resp['Cache-Control'] = 'public, max-age=604800'
    return resp

def pwa_service_worker(request):
    resp = HttpResponse(_PWA_SW, content_type='application/javascript')
    resp['Service-Worker-Allowed'] = '/'
    resp['Cache-Control'] = 'no-cache'
    return resp

SORU_ROLLERI = [Rol.PERSONEL, Rol.SEF]
SORU_SURE = 30
SORU_SURE_PAYLI = 38
CALISMAYAN_TIPLER = [VardiyaTipi.IZINLI, VardiyaTipi.YILLIK_IZIN, VardiyaTipi.RAPORLU, VardiyaTipi.DEVAMSIZ]

KARNE_ROLLERI = [Rol.EGITMEN, Rol.MUDUR, Rol.GENEL_MUDUR, Rol.OPERATOR, Rol.YATIRIMCI]

SORU_YONETIM_ROLLERI = [Rol.EGITMEN, Rol.GENEL_MUDUR]

def _soru_sistemi_aktif():
    return SoruAyar.get().aktif

def _egitmen_mi(personel):
    return bool(personel) and (getattr(personel, 'egitmen', False) or personel.rol == Rol.EGITMEN)

def _karne_gorebilir(personel):
    return bool(personel) and (_egitmen_mi(personel) or personel.rol in
                               [Rol.MUDUR, Rol.GENEL_MUDUR, Rol.OPERATOR, Rol.YATIRIMCI])

def _soru_yonetebilir(personel):
    return bool(personel) and (_egitmen_mi(personel) or personel.rol == Rol.GENEL_MUDUR)

def _bugun_calisiyor_mu(personel, gun):
    v = personel.vardiyalar.filter(tarih=gun).first()
    if v and v.vardiya_tipi in CALISMAYAN_TIPLER:
        return False
    return True

def _gunluk_soru_getir_veya_ata(personel, gun):
    gs = GunlukSoru.objects.filter(personel=personel, tarih=gun).first()
    if gs:
        return gs
    if not _bugun_calisiyor_mu(personel, gun):
        return None
    aktif_ids = list(KahveSoru.objects.filter(aktif=True).values_list('id', flat=True))
    if not aktif_ids:
        return None
    gorulen = set(GunlukSoru.objects.filter(personel=personel)
                  .values_list('soru_id', flat=True))
    havuz = [i for i in aktif_ids if i not in gorulen] or aktif_ids
    soru_id = random.choice(havuz)
    sube = personel.sube
    return GunlukSoru.objects.create(
        personel=personel, sube=sube, sube_ad=(sube.ad if sube else ''),
        soru_id=soru_id, tarih=gun)

def _gunluk_finalize(gs):
    if gs and gs.baslangic and not gs.cevaplandi:
        gecen = (timezone.now() - gs.baslangic).total_seconds()
        if gecen > SORU_SURE_PAYLI:
            gs.cevaplandi = True
            gs.sure_doldu = True
            gs.dogru_mu = False
            gs.secilen = ''
            gs.save(update_fields=['cevaplandi', 'sure_doldu', 'dogru_mu', 'secilen'])

def gunluk_soru(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in SORU_ROLLERI:
        return redirect('ana_sayfa')
    if not _soru_sistemi_aktif():
        return redirect('ana_sayfa')

    today = timezone.localdate()
    gs = _gunluk_soru_getir_veya_ata(personel, today)
    if gs is None:
        return redirect('ana_sayfa')
    _gunluk_finalize(gs)
    if gs.cevaplandi:
        return redirect('ana_sayfa')

    if request.method == 'POST':

        suresi_doldu = False
        if gs.baslangic:
            gecen = (timezone.now() - gs.baslangic).total_seconds()
            suresi_doldu = gecen > SORU_SURE_PAYLI
        secilen = (request.POST.get('secilen') or '').strip().upper()
        if secilen not in ('A', 'B', 'C', 'D'):
            secilen = ''
        if suresi_doldu or not secilen:
            gs.secilen = '' if suresi_doldu else secilen
            gs.sure_doldu = suresi_doldu
            gs.dogru_mu = False
        else:
            gs.secilen = secilen
            gs.dogru_mu = (gs.soru is not None and secilen == gs.soru.dogru)
        gs.cevaplandi = True
        gs.save()
        if gs.dogru_mu:
            messages.success(request, "Doğru cevap! 🎉")
        elif gs.sure_doldu:
            messages.error(request, "Süre doldu, soru boş işlendi.")
        else:
            dogru_txt = gs.soru.sik(gs.soru.dogru) if gs.soru else ''
            messages.error(request, f"Yanlış. Doğru cevap: {dogru_txt}")
        return redirect('ana_sayfa')

    if gs.baslangic is None:
        gs.baslangic = timezone.now()
        gs.save(update_fields=['baslangic'])
    gecen = (timezone.now() - gs.baslangic).total_seconds()
    kalan = max(1, int(SORU_SURE - gecen))
    return render(request, 'gunluk_soru.html', {
        'personel': personel, 'aktif': 'home', 'gs': gs, 'soru': gs.soru,
        'kalan': kalan, 'sure': SORU_SURE,
    })

def bilgi_karnesi(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or not _karne_gorebilir(personel):
        return redirect('ana_sayfa')

    subeler = _yon_subeler(personel)
    sel_sube = _yonetici_sube(request, subeler)
    ay_str = request.GET.get('ay') or timezone.localdate().strftime('%Y-%m')
    try:
        ay_ilk, sonraki = _ay_araligi(ay_str)
    except Exception:
        ay_ilk, sonraki = _ay_araligi(timezone.localdate().strftime('%Y-%m'))
        ay_str = ay_ilk.strftime('%Y-%m')

    satirlar = []
    yanlislar = []
    if sel_sube:
        kisiler = (Personel.objects.filter(sube=sel_sube, rol__in=SORU_ROLLERI)
                   .order_by('ad_soyad'))
        kayitlar = (GunlukSoru.objects.filter(personel__in=kisiler,
                                              tarih__gte=ay_ilk, tarih__lt=sonraki)
                    .select_related('soru', 'personel'))

        for gs in kayitlar:
            _gunluk_finalize(gs)
        per_map = {k.id: {'personel': k, 'toplam': 0, 'dogru': 0, 'yanlis': 0, 'bos': 0}
                   for k in kisiler}
        for gs in kayitlar:
            d = per_map.get(gs.personel_id)
            if not d:
                continue
            d['toplam'] += 1
            if not gs.cevaplandi:

                d['toplam'] -= 1
                continue
            if gs.dogru_mu:
                d['dogru'] += 1
            elif gs.sure_doldu or not gs.secilen:
                d['bos'] += 1
                d['yanlis'] += 1
            else:
                d['yanlis'] += 1
            if gs.cevaplandi and not gs.dogru_mu and gs.soru:
                yanlislar.append({
                    'ad': gs.personel.ad_soyad, 'tarih': gs.tarih, 'soru': gs.soru.metin,
                    'secilen': (gs.soru.sik(gs.secilen) if gs.secilen else '—'),
                    'dogru': gs.soru.sik(gs.soru.dogru),
                    'bos': gs.sure_doldu or not gs.secilen,
                })
        for d in per_map.values():
            t = d['toplam']
            d['basari'] = round(d['dogru'] * 100 / t) if t else 0
            satirlar.append(d)
        satirlar.sort(key=lambda x: (-x['basari'], x['personel'].ad_soyad))
        yanlislar.sort(key=lambda x: x['tarih'], reverse=True)

    return render(request, 'bilgi_karnesi.html', {
        'personel': personel, 'aktif': 'bilgi_karnesi',
        'subeler': subeler, 'sel_sube': sel_sube, 'selected_ay_str': ay_str,
        'satirlar': satirlar, 'yanlislar': yanlislar,
    })

def soru_yonetimi(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or not _soru_yonetebilir(personel):
        return redirect('ana_sayfa')

    ayar = SoruAyar.get()

    if request.method == 'POST':
        islem = request.POST.get('islem')

        if islem == 'sistem_durum':
            ayar.aktif = (request.POST.get('aktif') == '1')
            ayar.save(update_fields=['aktif', 'guncelleme'])
            messages.success(request, "Günlük soru sistemi " + ("aktif edildi." if ayar.aktif else "pasife alındı."))
            return redirect('soru_yonetimi')

        if islem == 'soru_ekle':
            metin = (request.POST.get('metin') or '').strip()
            a = (request.POST.get('sik_a') or '').strip()
            b = (request.POST.get('sik_b') or '').strip()
            c = (request.POST.get('sik_c') or '').strip()
            d = (request.POST.get('sik_d') or '').strip()
            dogru = (request.POST.get('dogru') or '').strip().upper()
            kategori = (request.POST.get('kategori') or '').strip()[:40]
            if metin and a and b and c and d and dogru in ('A', 'B', 'C', 'D'):
                KahveSoru.objects.create(metin=metin, sik_a=a[:200], sik_b=b[:200],
                                         sik_c=c[:200], sik_d=d[:200], dogru=dogru,
                                         kategori=kategori, aktif=True)
                messages.success(request, "Soru eklendi.")
            else:
                messages.error(request, "Tüm alanlar ve geçerli bir doğru şık (A/B/C/D) zorunlu.")
            return redirect('soru_yonetimi')

        if islem == 'soru_guncelle':
            s = KahveSoru.objects.filter(id=request.POST.get('soru_id')).first()
            if s:
                metin = (request.POST.get('metin') or '').strip()
                a = (request.POST.get('sik_a') or '').strip()
                b = (request.POST.get('sik_b') or '').strip()
                c = (request.POST.get('sik_c') or '').strip()
                d = (request.POST.get('sik_d') or '').strip()
                dogru = (request.POST.get('dogru') or '').strip().upper()
                if metin and a and b and c and d and dogru in ('A', 'B', 'C', 'D'):
                    s.metin = metin
                    s.sik_a, s.sik_b, s.sik_c, s.sik_d = a[:200], b[:200], c[:200], d[:200]
                    s.dogru = dogru
                    s.kategori = (request.POST.get('kategori') or '').strip()[:40]
                    s.save()
                    messages.success(request, "Soru güncellendi.")
                else:
                    messages.error(request, "Geçersiz soru bilgisi.")
            return redirect('soru_yonetimi')

        if islem == 'soru_durum':
            s = KahveSoru.objects.filter(id=request.POST.get('soru_id')).first()
            if s:
                s.aktif = not s.aktif
                s.save(update_fields=['aktif'])
                messages.success(request, ("Soru aktif edildi." if s.aktif else "Soru pasife alındı (çıkarıldı)."))
            return redirect('soru_yonetimi')

        return redirect('soru_yonetimi')

    arama = (request.GET.get('q') or '').strip()
    kategori = (request.GET.get('kategori') or '').strip()
    qs = KahveSoru.objects.all().order_by('-aktif', 'kategori', 'id')
    if arama:
        qs = qs.filter(metin__icontains=arama)
    if kategori:
        qs = qs.filter(kategori=kategori)
    sorular = list(qs)
    kategoriler = sorted(set(KahveSoru.objects.exclude(kategori='')
                             .values_list('kategori', flat=True)))
    return render(request, 'soru_yonetimi.html', {
        'personel': personel, 'aktif': 'soru_yonetimi',
        'ayar': ayar, 'sorular': sorular, 'kategoriler': kategoriler,
        'arama': arama, 'sel_kategori': kategori,
        'toplam': len(sorular),
        'aktif_sayi': KahveSoru.objects.filter(aktif=True).count(),
    })

STOK_DUZENLE_ROLLERI = [Rol.MUDUR, Rol.SATIN_ALMA, Rol.GENEL_MUDUR, Rol.OPERATOR]

def _dec(v, vars):
    try:
        d = Decimal(str(v).replace(',', '.'))
        return d if d > 0 else vars
    except Exception:
        return vars

def stok_duzenle(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in STOK_DUZENLE_ROLLERI:
        return redirect('ana_sayfa')

    if request.method == 'POST':
        islem = request.POST.get('islem')
        geri_grup = (request.POST.get('geri_grup') or '').strip()
        geri = f"{reverse('stok_duzenle')}?grup={geri_grup}" if geri_grup else reverse('stok_duzenle')

        if islem == 'urun_guncelle':
            u = StokUrun.objects.filter(id=request.POST.get('urun_id')).first()
            if u:
                ad = (request.POST.get('ad') or '').strip()
                if ad:
                    u.ad = ad[:200]
                u.kategori = (request.POST.get('kategori') or u.kategori).strip()[:80]
                u.kapali_icerik = _dec(request.POST.get('kapali_icerik'), u.kapali_icerik)
                u.acik_carpan = _dec(request.POST.get('acik_carpan'), u.acik_carpan)
                u.save(update_fields=['ad', 'kategori', 'kapali_icerik', 'acik_carpan'])
                messages.success(request, f"{u.ad} güncellendi.")
            return redirect(geri)

        if islem == 'urun_cikar':
            u = StokUrun.objects.filter(id=request.POST.get('urun_id')).first()
            if u:
                u.aktif = False
                u.save(update_fields=['aktif'])
                messages.success(request, f"{u.ad} listeden çıkarıldı.")
            return redirect(geri)

        if islem == 'urun_ekle':
            ad = (request.POST.get('ad') or '').strip()
            kategori = (request.POST.get('kategori') or '').strip() or geri_grup
            kapali = _dec(request.POST.get('kapali_icerik'), Decimal('1'))
            acik = _dec(request.POST.get('acik_carpan'), Decimal('1'))
            if ad and kategori:
                var = StokUrun.objects.filter(ad=ad).first()
                if var:
                    var.kategori = kategori[:80]
                    var.kapali_icerik = kapali
                    var.acik_carpan = acik
                    var.aktif = True
                    var.save()
                    messages.success(request, f"{ad} güncellendi (zaten vardı, yeniden eklendi).")
                else:
                    son = StokUrun.objects.order_by('-sira').first()
                    StokUrun.objects.create(kategori=kategori[:80], ad=ad[:200],
                                            kapali_icerik=kapali, acik_carpan=acik,
                                            sira=(son.sira + 1 if son else 0), aktif=True)
                    messages.success(request, f"{ad} eklendi.")
                geri = f"{reverse('stok_duzenle')}?grup={kategori}"
            else:
                messages.error(request, "Ürün adı ve grup zorunlu.")
            return redirect(geri)

        return redirect(geri)

    aktif_urunler = list(StokUrun.objects.filter(aktif=True).order_by('sira', 'ad'))
    gruplar = []
    for u in aktif_urunler:
        if u.kategori not in gruplar:
            gruplar.append(u.kategori)
    sel_grup = request.GET.get('grup') or (gruplar[0] if gruplar else '')
    urunler = [u for u in aktif_urunler if u.kategori == sel_grup]
    return render(request, 'stok_duzenle.html', {
        'personel': personel, 'aktif': 'stok_duzenle',
        'gruplar': gruplar, 'sel_grup': sel_grup, 'urunler': urunler,
    })

def sube_yeni(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in [Rol.GENEL_MUDUR, Rol.OPERATOR]:
        return redirect('ana_sayfa')
    if request.method == 'POST':
        ad = (request.POST.get('ad') or '').strip()
        depo_mu = request.POST.get('depo_mu') == 'on'
        if not ad:
            messages.error(request, "Şube adı boş olamaz.")
        elif Sube.objects.filter(ad__iexact=ad).exists():
            messages.error(request, "Bu isimde bir şube zaten var.")
        else:
            Sube.objects.create(ad=ad, depo_mu=depo_mu)
            messages.success(request, "Yeni şube eklendi: %s" % ad)
        return redirect('sube_yeni')
    subeler = Sube.objects.all().order_by('ad')
    veriler = [{'sube': s, 'sayi': Personel.objects.filter(sube=s).count()} for s in subeler]
    return render(request, 'sube_yonetim.html', {
        'personel': personel, 'aktif': 'sube_yeni', 'veriler': veriler,
    })


MOLA_QR_YETKI = [Rol.GENEL_MUDUR, Rol.OPERATOR]
MOLA_SURELER = [45, 15]
MOLA_QR_GECERLILIK_SN = 480  # 8 dakika (kamera açılışındaki gecikmelere pay bırakmak için 3 dk'dan yükseltildi)


def _mola_qr_acik():
    try:
        a = MolaQRAyar.objects.first()
        return bool(a and a.acik)
    except Exception:
        return False


def _sube_mola_token(sube):
    tok = SubeMolaToken.objects.filter(sube=sube).first()
    if tok is None:
        tok = SubeMolaToken.objects.create(sube=sube, token=secrets.token_urlsafe(12))
    return tok


def _aktif_mola(personel):
    return MolaOturum.objects.filter(personel=personel, bitis__isnull=True).order_by('-baslangic').first()


def mola_qr_yonetim(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in MOLA_QR_YETKI:
        return redirect('ana_sayfa')
    if request.method == 'POST':
        islem = request.POST.get('islem')
        if islem in ('sistem_ac', 'sistem_kapat'):
            a, _ = MolaQRAyar.objects.get_or_create(id=1)
            a.acik = (islem == 'sistem_ac')
            a.save()
            messages.success(request, "Mola QR sistemi %s." % ("açıldı" if a.acik else "kapatıldı"))
        elif islem == 'token_yenile':
            sb = Sube.objects.filter(id=request.POST.get('sube_id')).first()
            if sb:
                SubeMolaToken.objects.filter(sube=sb).delete()
                _sube_mola_token(sb)
                messages.success(request, "%s için yeni QR üretildi (eski QR geçersiz)." % sb.ad)
        return redirect('mola_qr_yonetim')
    veriler = []
    for s in Sube.objects.filter(depo_mu=False).order_by('ad'):
        tok = _sube_mola_token(s)
        veriler.append({'sube': s, 'url': request.build_absolute_uri('/mola/qr/%s/' % tok.token)})
    return render(request, 'mola_qr_yonetim.html', {
        'personel': personel, 'aktif': 'mola_qr', 'acik': _mola_qr_acik(),
        'veriler': veriler,
    })


def mola_qr_giris(request, token):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    if personel.rol == Rol.MUTFAK_SORUMLUSU:
        messages.info(request, "Mutfak Sorumlusu rolü için mola sistemi kullanılmaz.")
        return redirect('ana_sayfa')
    tok = SubeMolaToken.objects.filter(token=token).select_related('sube').first()
    if tok is None:
        return render(request, 'mola_qr_giris.html', {'personel': personel, 'hata': 'Geçersiz veya güncelliğini yitirmiş QR kodu.'})
    if not _mola_qr_acik():
        return render(request, 'mola_qr_giris.html', {'personel': personel, 'sube': tok.sube, 'kapali': True})
    aktif = _aktif_mola(personel)
    if request.method == 'POST':
        islem = request.POST.get('islem')
        tazelik = request.session.get('mola_qr_zaman_%s' % tok.sube_id)
        taze_mi = False
        if tazelik:
            try:
                taze_mi = (timezone.now().timestamp() - float(tazelik)) <= MOLA_QR_GECERLILIK_SN
            except (TypeError, ValueError):
                taze_mi = False
        if not taze_mi:
            messages.error(request, "QR kodunun üzerinden biraz zaman geçmiş. Lütfen QR'ı tekrar okut.")
            return redirect('mola_tara')
        if islem == 'baslat' and aktif is None:
            try:
                sure = int(request.POST.get('sure', '45'))
            except (TypeError, ValueError):
                sure = 45
            if sure not in MOLA_SURELER:
                sure = 45
            MolaOturum.objects.create(personel=personel, sube=tok.sube, sure_dk=sure, baslangic=timezone.now())
            messages.success(request, "Molan başladı (%d dk). Döndüğünde bu QR'ı tekrar okut." % sure)
        elif islem == 'baslat' and aktif is not None:
            messages.info(request, "Zaten aktif bir molan var.")
        elif islem == 'bitir' and aktif is None:
            messages.info(request, "Aktif bir molan görünmüyor (belki az önce bitirildi).")
        elif islem == 'bitir' and aktif is not None:
            aktif.bitis = timezone.now()
            gecen = max(0, int((aktif.bitis - aktif.baslangic).total_seconds() // 60))
            aktif.kullanilan_dk = gecen
            aktif.save(update_fields=['bitis', 'kullanilan_dk'])
            messages.success(request, "Molan bitti. Toplam %d dk mola kullandın." % gecen)
        return redirect('mola_qr_giris', token=token)
    aktif = _aktif_mola(personel)
    kalan = None
    if aktif is not None:
        bitecek = aktif.baslangic + datetime.timedelta(minutes=aktif.sure_dk)
        kalan = int((bitecek - timezone.now()).total_seconds() // 60)
    request.session['mola_qr_zaman_%s' % tok.sube_id] = timezone.now().timestamp()
    return render(request, 'mola_qr_giris.html', {
        'personel': personel, 'sube': tok.sube, 'token': token,
        'aktif': aktif, 'kalan': kalan, 'sureler': MOLA_SURELER,
    })


def mola_tara(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    if personel.rol == Rol.MUTFAK_SORUMLUSU:
        messages.info(request, "Mutfak Sorumlusu rolü için mola sistemi kullanılmaz.")
        return redirect('ana_sayfa')
    return render(request, 'mola_tara.html', {
        'personel': personel, 'aktif': 'mola_tara', 'acik': _mola_qr_acik(),
    })


MOLA_IZLEME_ROLLER = [Rol.GENEL_MUDUR, Rol.OPERATOR, Rol.YATIRIMCI, Rol.MUDUR, Rol.MAGAZA_MUDURU, Rol.SEF, Rol.MUTFAK_SORUMLUSU]


def _mola_izleme_subeler(personel):
    if personel.rol in (Rol.GENEL_MUDUR, Rol.OPERATOR, Rol.YATIRIMCI, Rol.MUTFAK_SORUMLUSU):
        return list(Sube.objects.filter(depo_mu=False).values_list('id', flat=True))
    if personel.rol == Rol.MUDUR:
        return list(personel.sorumlu_subeler.values_list('id', flat=True))
    if personel.rol == Rol.MAGAZA_MUDURU and personel.sorumlu_subeler.exists():
        return list(personel.sorumlu_subeler.values_list('id', flat=True))
    if personel.rol in (Rol.MAGAZA_MUDURU, Rol.SEF) and personel.sube_id:
        return [personel.sube_id]
    return []


def mola_izleme(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in MOLA_IZLEME_ROLLER:
        return redirect('ana_sayfa')
    sube_ids = _mola_izleme_subeler(personel)
    tekli_rol = personel.rol == Rol.SEF or (personel.rol == Rol.MAGAZA_MUDURU and len(sube_ids) <= 1)
    subeler = list(Sube.objects.filter(id__in=sube_ids).order_by('ad')) if (len(sube_ids) > 1 and not tekli_rol) else []
    return render(request, 'mola_izleme.html', {
        'personel': personel, 'aktif': 'mola_izleme', 'subeler': subeler,
    })


def mola_izleme_json(request):
    if not request.user.is_authenticated:
        return JsonResponse({'molalar': []}, status=403)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in MOLA_IZLEME_ROLLER:
        return JsonResponse({'molalar': []}, status=403)
    sube_ids = _mola_izleme_subeler(personel)
    tekli_rol = personel.rol == Rol.SEF or (personel.rol == Rol.MAGAZA_MUDURU and len(sube_ids) <= 1)
    if not tekli_rol:
        sec = request.GET.get('sube')
        if sec and sec.isdigit() and int(sec) in sube_ids:
            sube_ids = [int(sec)]
    now = timezone.now()
    out = []
    qs = (MolaOturum.objects.filter(bitis__isnull=True, sube_id__in=sube_ids)
          .select_related('personel', 'sube').order_by('baslangic'))
    if personel.rol == Rol.MUTFAK_SORUMLUSU:
        qs = qs.filter(personel__rol=Rol.MUTFAK_PERSONEL)
    for m in qs:
        gecen = max(0, int((now - m.baslangic).total_seconds() // 60))
        out.append({
            'ad': m.personel.ad_soyad if m.personel else '—',
            'sube': m.sube.ad if m.sube else '—',
            'sure_dk': m.sure_dk,
            'baslangic': m.baslangic.isoformat(),
            'gecen_dk': gecen,
            'kalan_dk': m.sure_dk - gecen,
            'manuel_mi': m.manuel_mi,
        })
    return JsonResponse({'molalar': out, 'zaman': now.isoformat()})


def mola_gecmis(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in MOLA_IZLEME_ROLLER:
        return redirect('ana_sayfa')
    sube_ids = _mola_izleme_subeler(personel)
    tekli_rol = personel.rol == Rol.SEF or (personel.rol == Rol.MAGAZA_MUDURU and len(sube_ids) <= 1)
    subeler = list(Sube.objects.filter(id__in=sube_ids).order_by('ad')) if (len(sube_ids) > 1 and not tekli_rol) else []
    sec = request.GET.get('sube')
    if tekli_rol:
        gecmis_ids = sube_ids
        secili = personel.sube_id or 0
    elif sec and sec.isdigit() and int(sec) in sube_ids:
        gecmis_ids = [int(sec)]
        secili = int(sec)
    else:
        gecmis_ids = sube_ids
        secili = 0
    aralik = _gun_araligi(request, 'mg_bas', 'mg_bit')
    hizli = request.GET.get('hizli', '14')
    if aralik:
        bas, bit_ust, bas_str, bit_str = aralik
        sinir, ust = bas, bit_ust
        hizli = ''
    else:
        bugun = timezone.localdate()
        if hizli == 'bugun':
            bas_g = bugun
        elif hizli == 'hafta':
            bas_g = bugun - datetime.timedelta(days=bugun.weekday())
        elif hizli == 'ay':
            bas_g = bugun.replace(day=1)
        else:
            hizli = '14'
            bas_g = bugun - datetime.timedelta(days=13)
        sinir = datetime.datetime.combine(bas_g, datetime.time.min, tzinfo=timezone.get_current_timezone())
        ust = timezone.now() + datetime.timedelta(days=1)
        bas_str = bas_g.strftime('%Y-%m-%d')
        bit_str = bugun.strftime('%Y-%m-%d')
    kayitlar = (MolaOturum.objects.filter(bitis__isnull=False, sube_id__in=gecmis_ids,
                                          baslangic__gte=sinir, baslangic__lt=ust)
                .select_related('personel', 'sube').order_by('-baslangic'))
    if personel.rol == Rol.MUTFAK_SORUMLUSU:
        kayitlar = kayitlar.filter(personel__rol=Rol.MUTFAK_PERSONEL)
    kayitlar = kayitlar[:500]
    liste = []
    for m in kayitlar:
        if m.kullanilan_dk is not None:
            dk = m.kullanilan_dk
        else:
            dk = max(0, int((m.bitis - m.baslangic).total_seconds() // 60))
        liste.append({
            'ad': m.personel.ad_soyad if m.personel else '—',
            'sube': m.sube.ad if m.sube else '—',
            'giris': timezone.localtime(m.baslangic),
            'cikis': timezone.localtime(m.bitis),
            'sure_dk': dk,
            'limit': m.sure_dk,
            'asti': dk > m.sure_dk,
            'manuel_mi': m.manuel_mi,
        })
    toplam_kisi = len(set(k['ad'] for k in liste))
    asan_sayisi = sum(1 for k in liste if k['asti'])
    return render(request, 'mola_gecmis.html', {
        'personel': personel, 'aktif': 'mola_gecmis',
        'subeler': subeler, 'secili': secili, 'kayitlar': liste,
        'hizli': hizli, 'mg_bas': bas_str, 'mg_bit': bit_str,
        'toplam_kayit': len(liste), 'toplam_kisi': toplam_kisi, 'asan_sayisi': asan_sayisi,
    })


INSAAT_ATAMA_ROLLER = [Rol.GENEL_MUDUR, Rol.YATIRIMCI, Rol.OPERATOR]
INSAAT_GORUNTULE_ROLLER = [Rol.GENEL_MUDUR, Rol.YATIRIMCI, Rol.OPERATOR, Rol.MUDUR]


def _insaat_yonetebilir(personel, proje):
    if personel.rol in (Rol.GENEL_MUDUR, Rol.OPERATOR):
        return True
    if personel.rol == Rol.MUDUR and proje.sorumlu_id == personel.id:
        return True
    return False


def _insaat_ilerleme(maddeler):
    toplam = len(maddeler)
    tamam = sum(1 for m in maddeler if m.durum == InsaatMaddeDurum.TAMAM)
    return toplam, tamam, (round(tamam * 100 / toplam) if toplam else 0)


def insaat_liste(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in INSAAT_GORUNTULE_ROLLER:
        return redirect('ana_sayfa')
    atayabilir = personel.rol in INSAAT_ATAMA_ROLLER
    if request.method == 'POST' and atayabilir and request.POST.get('islem') == 'proje_ekle':
        ad = (request.POST.get('ad') or '').strip()
        mudur = Personel.objects.filter(id=request.POST.get('sorumlu_id'), rol=Rol.MUDUR).first() if (request.POST.get('sorumlu_id') or '').isdigit() else None
        if ad:
            proje = InsaatProje.objects.create(ad=ad[:160], sorumlu=mudur, olusturan=personel)
            sablonlar = list(InsaatSablonMadde.objects.all())
            if sablonlar:
                InsaatMadde.objects.bulk_create([
                    InsaatMadde(proje=proje, kategori=s.kategori, metin=s.metin, sira=s.sira)
                    for s in sablonlar
                ])
            messages.success(request, "Proje oluşturuldu: %s%s" % (ad, (" (%d şablon madde eklendi)" % len(sablonlar)) if sablonlar else ""))
        else:
            messages.error(request, "Proje adı boş olamaz.")
        return redirect('insaat_liste')
    veriler = []
    for p in InsaatProje.objects.select_related('sorumlu').all():
        toplam, tamam, yuzde = _insaat_ilerleme(list(p.maddeler.all()))
        veriler.append({'proje': p, 'toplam': toplam, 'tamam': tamam, 'yuzde': yuzde,
                        'yonet': _insaat_yonetebilir(personel, p)})
    mudurler = Personel.objects.filter(rol=Rol.MUDUR).order_by('ad_soyad') if atayabilir else []
    return render(request, 'insaat_liste.html', {
        'personel': personel, 'aktif': 'insaat', 'veriler': veriler,
        'atayabilir': atayabilir, 'mudurler': mudurler,
    })


def insaat_detay(request, pid):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in INSAAT_GORUNTULE_ROLLER:
        return redirect('ana_sayfa')
    proje = InsaatProje.objects.filter(id=pid).select_related('sorumlu').first()
    if proje is None:
        return redirect('insaat_liste')
    yonet = _insaat_yonetebilir(personel, proje)
    atayabilir = personel.rol in INSAAT_ATAMA_ROLLER

    def _geri(anchor=''):
        from django.urls import reverse
        url = reverse('insaat_detay', args=[pid])
        return redirect(url + ('#' + anchor if anchor else ''))

    if request.method == 'POST':
        islem = request.POST.get('islem')
        if islem == 'proje_sil' and atayabilir:
            proje.delete()
            messages.success(request, "Proje silindi.")
            return redirect('insaat_liste')
        if islem == 'ata' and atayabilir:
            m = Personel.objects.filter(id=request.POST.get('sorumlu_id'), rol=Rol.MUDUR).first() if (request.POST.get('sorumlu_id') or '').isdigit() else None
            proje.sorumlu = m
            proje.save(update_fields=['sorumlu'])
            messages.success(request, "Sorumlu müdür güncellendi.")
            return _geri()
        if islem == 'sablon_yap' and atayabilir:
            InsaatSablonMadde.objects.all().delete()
            InsaatSablonMadde.objects.bulk_create([
                InsaatSablonMadde(kategori=m.kategori, metin=m.metin, sira=m.sira)
                for m in proje.maddeler.all()
            ])
            messages.success(request, "Bu projedeki maddeler yeni projeler için varsayılan şablon yapıldı.")
            return _geri()
        if yonet:
            if islem == 'madde_ekle':
                metin = (request.POST.get('metin') or '').strip()
                kategori = request.POST.get('kategori')
                if kategori not in dict(InsaatKategori.choices):
                    kategori = InsaatKategori.URUN
                if metin:
                    yeni = InsaatMadde.objects.create(proje=proje, metin=metin[:300],
                                                      kategori=kategori, sira=proje.maddeler.count())
                    return _geri('madde-%d' % yeni.id)
                return _geri('ekle')
            elif islem == 'madde_sil':
                InsaatMadde.objects.filter(id=request.POST.get('madde_id'), proje=proje).delete()
                return _geri('liste')
            elif islem == 'durum':
                md = InsaatMadde.objects.filter(id=request.POST.get('madde_id'), proje=proje).first()
                yd = request.POST.get('durum')
                if md and yd in dict(InsaatMaddeDurum.choices):
                    md.durum = yd
                    md.save(update_fields=['durum'])
                    return _geri('madde-%d' % md.id)
            elif islem == 'not_kaydet':
                md = InsaatMadde.objects.filter(id=request.POST.get('madde_id'), proje=proje).first()
                if md:
                    md.aciklama = (request.POST.get('aciklama') or '')[:2000]
                    md.save(update_fields=['aciklama'])
                    return _geri('madde-%d' % md.id)
            elif islem == 'proje_tamamla':
                proje.tamamlandi = not proje.tamamlandi
                proje.save(update_fields=['tamamlandi'])
        return _geri()
    maddeler = list(proje.maddeler.all())
    toplam, tamam, yuzde = _insaat_ilerleme(maddeler)
    gruplar = []
    for kod, ad in InsaatKategori.choices:
        gm = [m for m in maddeler if m.kategori == kod]
        gt, gtm, gy = _insaat_ilerleme(gm)
        gruplar.append({'kod': kod, 'ad': ad, 'maddeler': gm, 'toplam': gt, 'tamam': gtm, 'yuzde': gy})
    mudurler = Personel.objects.filter(rol=Rol.MUDUR).order_by('ad_soyad') if atayabilir else []
    return render(request, 'insaat_detay.html', {
        'personel': personel, 'aktif': 'insaat', 'proje': proje,
        'gruplar': gruplar, 'kategoriler': InsaatKategori.choices,
        'toplam': toplam, 'tamam': tamam, 'yuzde': yuzde, 'yonet': yonet,
        'atayabilir': atayabilir, 'mudurler': mudurler,
    })


def insaat_pdf(request, pid):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in INSAAT_GORUNTULE_ROLLER:
        return redirect('ana_sayfa')
    proje = InsaatProje.objects.filter(id=pid).select_related('sorumlu').first()
    if proje is None:
        return redirect('insaat_liste')
    from .insaat_pdf import insaat_pdf_uret
    icerik = insaat_pdf_uret(proje, list(proje.maddeler.all()))
    resp = HttpResponse(icerik, content_type='application/pdf')
    resp['Content-Disposition'] = 'inline; filename="insaat_denetim_%d.pdf"' % proje.id
    return resp


def sube_sec(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    sid = request.GET.get('sube_id', '')
    if sid.isdigit():
        request.session['sel_sube_id'] = sid
    geri = request.GET.get('next', '') or '/'
    if not geri.startswith('/') or geri.startswith('//'):
        geri = '/'
    return redirect(geri)


def profil_sayfa(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')

    if request.method == 'POST':
        islem = request.POST.get('islem')
        if islem == 'profil_kaydet':
            dt = request.POST.get('dogum_tarihi', '').strip()
            if dt:
                try:
                    personel.dogum_tarihi = datetime.datetime.strptime(dt, '%Y-%m-%d').date()
                except ValueError:
                    pass
            else:
                personel.dogum_tarihi = None
            cinsiyet = request.POST.get('cinsiyet', '')
            if cinsiyet in ('E', 'K', ''):
                personel.cinsiyet = cinsiyet
            foto = request.FILES.get('profil_foto')
            if foto:
                gecerli = False
                if foto.size <= 5 * 1024 * 1024 and (foto.content_type or '').startswith('image/'):
                    try:
                        from PIL import Image as _PILImage
                        _img = _PILImage.open(foto)
                        _img.verify()
                        foto.seek(0)
                        gecerli = True
                    except Exception:
                        gecerli = False
                if gecerli:
                    personel.profil_foto = foto
                else:
                    messages.error(request, "Geçersiz fotoğraf dosyası (en fazla 5MB, resim formatı olmalı).")
                    return redirect('profil')
            personel.save(update_fields=['dogum_tarihi', 'cinsiyet', 'profil_foto'])
            messages.success(request, "Profilin güncellendi.")
        elif islem == 'foto_kaldir':
            if personel.profil_foto:
                personel.profil_foto.delete(save=False)
            personel.profil_foto = None
            personel.save(update_fields=['profil_foto'])
            messages.success(request, "Profil fotoğrafı kaldırıldı.")
        return redirect('profil')

    return render(request, 'profil.html', {'personel': personel, 'aktif': 'profil'})


MESAI_QR_YETKI = [Rol.GENEL_MUDUR, Rol.OPERATOR]
MESAI_QR_GECERLILIK_SN = 480  # 8 dakika (kamera açılışındaki gecikmelere pay bırakmak için 3 dk'dan yükseltildi)


def _sube_mesai_token(sube):
    tok = SubeMesaiToken.objects.filter(sube=sube).first()
    if tok is None:
        tok = SubeMesaiToken.objects.create(sube=sube, token=secrets.token_urlsafe(12))
    return tok


def _aktif_mesai(personel):
    return MesaiKayit.objects.filter(personel=personel, cikis__isnull=True).order_by('-giris').first()


MANUEL_YETKI_DELEGE_ROLLER = [Rol.SEF, Rol.MAGAZA_MUDURU]


def manuel_giris(request):
    """Kamerası çalışmayan personel için: yetkili kişi (kendisi ya da Şef/Mağaza Müdürü ise
    kendi şubesindeki personel adına) QR olmadan mola/mesai başlatıp bitirebilir. Kayıtlar
    QR sistemiyle AYNI tablolara (MolaOturum/MesaiKayit) yazılır, sadece 'manuel_mi' ile
    işaretlenir — böylece Mola Takibi, Mesai Kayıtları ve Puantaj'a otomatik yansır."""
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or not personel.manuel_giris_yetkisi:
        return redirect('ana_sayfa')

    delege_mi = personel.rol in MANUEL_YETKI_DELEGE_ROLLER

    if request.method == 'POST':
        islem = request.POST.get('islem')
        if delege_mi:
            hedef = Personel.objects.filter(id=request.POST.get('personel_id'), sube=personel.sube).first()
        else:
            hedef = personel
        if hedef is None:
            messages.error(request, "Geçersiz kişi.")
            return redirect('manuel_giris')
        not_metni = (request.POST.get('not', '') or '').strip()[:200]

        if islem == 'mola_baslat':
            if _aktif_mola(hedef) is None:
                try:
                    sure = int(request.POST.get('sure', '45'))
                except (TypeError, ValueError):
                    sure = 45
                if sure not in MOLA_SURELER:
                    sure = 45
                MolaOturum.objects.create(personel=hedef, sube=hedef.sube, sure_dk=sure,
                                          baslangic=timezone.now(), manuel_mi=True,
                                          manuel_giren=personel, manuel_not=not_metni)
                messages.success(request, f"{hedef.ad_soyad} için mola manuel başlatıldı.")
            else:
                messages.info(request, f"{hedef.ad_soyad} için zaten aktif bir mola var.")
        elif islem == 'mola_bitir':
            aktif = _aktif_mola(hedef)
            if aktif:
                aktif.bitis = timezone.now()
                aktif.kullanilan_dk = max(0, int((aktif.bitis - aktif.baslangic).total_seconds() // 60))
                aktif.manuel_mi = True
                aktif.manuel_giren = personel
                if not_metni:
                    aktif.manuel_not = (aktif.manuel_not + ' / ' if aktif.manuel_not else '') + not_metni
                aktif.save()
                messages.success(request, f"{hedef.ad_soyad} için mola manuel bitirildi.")
            else:
                messages.info(request, f"{hedef.ad_soyad} için aktif bir mola yok.")
        elif islem == 'mesai_giris':
            if _aktif_mesai(hedef) is None:
                MesaiKayit.objects.create(personel=hedef, sube=hedef.sube, giris=timezone.now(),
                                          personel_ad_arsiv=hedef.ad_soyad, manuel_mi=True,
                                          manuel_giren=personel, manuel_not=not_metni)
                messages.success(request, f"{hedef.ad_soyad} için mesai girişi manuel yapıldı.")
            else:
                messages.info(request, f"{hedef.ad_soyad} için zaten açık bir mesai var.")
        elif islem == 'mesai_cikis':
            aktif = _aktif_mesai(hedef)
            if aktif:
                aktif.cikis = timezone.now()
                aktif.manuel_mi = True
                aktif.manuel_giren = personel
                if not_metni:
                    aktif.manuel_not = (aktif.manuel_not + ' / ' if aktif.manuel_not else '') + not_metni
                aktif.save()
                messages.success(request, f"{hedef.ad_soyad} için mesai çıkışı manuel yapıldı.")
            else:
                messages.info(request, f"{hedef.ad_soyad} için açık bir mesai yok.")
        return redirect('manuel_giris')

    if delege_mi:
        hedefler = list(Personel.objects.filter(sube=personel.sube,
                                                rol__in=[Rol.PERSONEL, Rol.SEF, Rol.MAGAZA_MUDURU])
                        .exclude(id=personel.id).order_by('ad_soyad'))
    else:
        hedefler = [personel]

    for h in hedefler:
        h.aktif_mola_obj = _aktif_mola(h)
        h.aktif_mesai_obj = _aktif_mesai(h)

    return render(request, 'manuel_giris.html', {
        'personel': personel, 'aktif': 'manuel_giris',
        'delege_mi': delege_mi, 'hedefler': hedefler,
        'mola_sureler': MOLA_SURELER,
    })


def mesai_qr_yonetim(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in MESAI_QR_YETKI:
        return redirect('ana_sayfa')
    if request.method == 'POST':
        if request.POST.get('islem') == 'token_yenile':
            sb = Sube.objects.filter(id=request.POST.get('sube_id')).first()
            if sb:
                SubeMesaiToken.objects.filter(sube=sb).delete()
                _sube_mesai_token(sb)
                messages.success(request, "%s için yeni QR üretildi (eski QR geçersiz)." % sb.ad)
        return redirect('mesai_qr_yonetim')
    veriler = []
    for s in Sube.objects.filter(depo_mu=False).order_by('ad'):
        tok = _sube_mesai_token(s)
        veriler.append({'sube': s, 'url': request.build_absolute_uri('/mesai/qr/%s/' % tok.token)})
    return render(request, 'mesai_qr_yonetim.html', {
        'personel': personel, 'aktif': 'mesai_qr', 'veriler': veriler,
    })


def mesai_tara(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    return render(request, 'mesai_tara.html', {'personel': personel, 'aktif': 'mesai_tara'})


def mesai_qr_giris(request, token):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    tok = SubeMesaiToken.objects.filter(token=token).select_related('sube').first()
    if tok is None:
        return render(request, 'mesai_qr_giris.html', {'personel': personel, 'hata': 'Geçersiz veya güncelliğini yitirmiş QR kodu.'})
    aktif = _aktif_mesai(personel)
    if request.method == 'POST':
        islem = request.POST.get('islem')
        tazelik = request.session.get('mesai_qr_zaman_%s' % tok.sube_id)
        taze_mi = False
        if tazelik:
            try:
                taze_mi = (timezone.now().timestamp() - float(tazelik)) <= MESAI_QR_GECERLILIK_SN
            except (TypeError, ValueError):
                taze_mi = False
        if not taze_mi:
            messages.error(request, "QR kodunun üzerinden biraz zaman geçmiş. Lütfen QR'ı tekrar okut.")
            return redirect('mesai_tara')
        if islem == 'mesai_giris' and aktif is None:
            MesaiKayit.objects.create(personel=personel, sube=tok.sube, giris=timezone.now(),
                                      personel_ad_arsiv=personel.ad_soyad)
            messages.success(request, "Giriş kaydedildi. İyi çalışmalar!")
        elif islem == 'mesai_giris' and aktif is not None:
            messages.info(request, "Zaten açık bir mesain var.")
        elif islem == 'mesai_cikis' and aktif is None:
            messages.info(request, "Açık bir mesain görünmüyor (belki az önce çıkış yapıldı).")
        elif islem == 'mesai_cikis' and aktif is not None:
            aktif.cikis = timezone.now()
            aktif.save(update_fields=['cikis'])
            messages.success(request, "Çıkış kaydedildi. İyi günler!")
        return redirect('mesai_qr_giris', token=token)
    request.session['mesai_qr_zaman_%s' % tok.sube_id] = timezone.now().timestamp()
    return render(request, 'mesai_qr_giris.html', {
        'personel': personel, 'sube': tok.sube, 'token': token, 'aktif': aktif,
    })


def mesai_kayitlari(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or (personel.rol not in UST_YONETIM and personel.rol not in (Rol.SEF, Rol.MAGAZA_MUDURU, Rol.MUTFAK_SORUMLUSU)):
        return redirect('ana_sayfa')
    magaza_coklu = personel.rol == Rol.MAGAZA_MUDURU and personel.sorumlu_subeler.exists()
    is_yon = personel.rol in UST_YONETIM or personel.rol == Rol.MUTFAK_SORUMLUSU or magaza_coklu
    if is_yon:
        if personel.rol in UST_YONETIM:
            subeler = _yon_subeler(personel)
        elif magaza_coklu:
            subeler = list(personel.sorumlu_subeler.order_by('ad'))
        else:
            subeler = list(Sube.objects.filter(depo_mu=False).order_by('ad'))
        sube_ids = [s.id for s in subeler]
        sec = request.GET.get('sube')
        if sec and sec.isdigit() and int(sec) in sube_ids:
            gecmis_ids = [int(sec)]
            secili = int(sec)
        else:
            gecmis_ids = sube_ids
            secili = 0
        subeler_secim = list(Sube.objects.filter(id__in=sube_ids).order_by('ad')) if len(sube_ids) > 1 else []
    else:
        gecmis_ids = [personel.sube_id] if personel.sube_id else []
        secili = personel.sube_id or 0
        subeler_secim = []

    today = timezone.localdate()
    try:
        secili_tarih = datetime.datetime.strptime(request.GET.get('mesai_tarih', ''), '%Y-%m-%d').date()
    except ValueError:
        secili_tarih = today
    secili_tarih = min(secili_tarih, today)
    gun_bas = datetime.datetime.combine(secili_tarih, datetime.time.min, tzinfo=timezone.get_current_timezone())
    gun_son = gun_bas + datetime.timedelta(days=1)

    kayitlar = (MesaiKayit.objects.filter(sube_id__in=gecmis_ids, giris__gte=gun_bas, giris__lt=gun_son)
                .select_related('personel', 'sube').order_by('-giris'))
    if personel.rol == Rol.MUTFAK_SORUMLUSU:
        kayitlar = kayitlar.filter(personel__rol=Rol.MUTFAK_PERSONEL)
    liste = []
    toplam_dk = {}
    for m in kayitlar:
        ad = m.personel.ad_soyad if m.personel else (m.personel_ad_arsiv or '—')
        anahtar = m.personel_id or ('arsiv_%s' % ad)
        dk = None
        if m.cikis:
            dk = max(0, int((m.cikis - m.giris).total_seconds() // 60))
            toplam_dk[anahtar] = toplam_dk.get(anahtar, 0) + dk
        liste.append({
            'anahtar': anahtar, 'ad': ad, 'sube': m.sube.ad if m.sube else '—',
            'giris': timezone.localtime(m.giris),
            'cikis': timezone.localtime(m.cikis) if m.cikis else None,
            'sure_dk': dk,
            'manuel_mi': m.manuel_mi,
        })
    for k in liste:
        toplam = toplam_dk.get(k['anahtar'], 0)
        k['toplam_saat'] = toplam // 60
        k['toplam_dk'] = toplam % 60
    return render(request, 'mesai_kayitlari.html', {
        'personel': personel, 'aktif': 'mesai_kayitlari',
        'subeler': subeler_secim, 'secili': secili, 'kayitlar': liste,
        'secili_tarih': secili_tarih.strftime('%Y-%m-%d'), 'bugun': today.strftime('%Y-%m-%d'),
    })


def gosterge(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in UST_YONETIM:
        return redirect('ana_sayfa')

    subeler = list(_yon_subeler(personel))
    sube_ids = [s.id for s in subeler]
    sel_sube_ozet = _yonetici_sube(request, subeler)
    hesap_ids = [sel_sube_ozet.id] if sel_sube_ozet else sube_ids
    today = timezone.localdate()
    ay_ilk, sonraki = _ay_araligi(today.strftime('%Y-%m'))

    calisan_tipler = [VardiyaTipi.SABAHCI, VardiyaTipi.ARACI, VardiyaTipi.AKSAMCI]
    v_today = Vardiya.objects.filter(tarih=today, personel__sube_id__in=hesap_ids)

    bugun = {
        'toplam': Personel.objects.filter(sube_id__in=hesap_ids, rol__in=[Rol.PERSONEL, Rol.SEF]).count(),
        'calisan': v_today.filter(vardiya_tipi__in=calisan_tipler).count(),
        'izinli': v_today.filter(vardiya_tipi__in=[VardiyaTipi.IZINLI, VardiyaTipi.YILLIK_IZIN]).count(),
        'raporlu': v_today.filter(vardiya_tipi=VardiyaTipi.RAPORLU).count(),
        'devamsiz': v_today.filter(vardiya_tipi=VardiyaTipi.DEVAMSIZ).count(),
    }

    onay_bekleyen = Vardiya.objects.filter(
        personel__sube_id__in=hesap_ids, durum=OnayDurumu.ONAY_BEKLIYOR
    ).values('personel__sube_id', 'tarih').distinct().count()
    acik_sevkiyat = SevkiyatTalep.objects.filter(
        sube_id__in=hesap_ids,
        durum__in=[SevkiyatDurumu.TALEP, SevkiyatDurumu.SEVKIYATTA, SevkiyatDurumu.ONAY_BEKLIYOR]
    ).count()

    sayim_yapan = (StokSayim.objects.filter(sube_id__in=hesap_ids, ay=ay_ilk)
                   .values('sube_id').distinct().count())
    sube_sayisi = len(hesap_ids)

    ayar = SoruAyar.get()
    soru = {'aktif': ayar.aktif, 'cevap': 0, 'dogru': 0, 'basari': 0}
    if ayar.aktif:
        gs = GunlukSoru.objects.filter(tarih=today, personel__sube_id__in=hesap_ids, cevaplandi=True)
        soru['cevap'] = gs.count()
        soru['dogru'] = gs.filter(dogru_mu=True).count()
        soru['basari'] = round(soru['dogru'] * 100 / soru['cevap']) if soru['cevap'] else 0

    _durum_kalemler = [
        ('Çalışıyor', bugun['calisan'], '#162AA3'),
        ('İzinli', bugun['izinli'], '#E8A33D'),
        ('Raporlu', bugun['raporlu'], '#8FA0E8'),
        ('Devamsız', bugun['devamsiz'], '#D7263D'),
    ]
    durum_top = sum(v for _, v, _ in _durum_kalemler)
    durum = [{'ad': a, 'sayi': v, 'renk': r,
              'yuzde': round(v * 100 / durum_top, 1) if durum_top else 0}
             for a, v, r in _durum_kalemler]

    gun7 = [today - datetime.timedelta(days=i) for i in range(6, -1, -1)]
    sevk_say = {g: 0 for g in gun7}
    for dt in (SevkiyatTalep.objects.filter(sube_id__in=hesap_ids, olusturma__date__gte=gun7[0])
               .values_list('olusturma', flat=True)):
        g = timezone.localtime(dt).date()
        if g in sevk_say:
            sevk_say[g] += 1
    sevk_max = max(sevk_say.values()) or 1
    _gun_kisa = ['Pzt', 'Sal', 'Çar', 'Per', 'Cum', 'Cmt', 'Paz']
    trend = []
    for g in gun7:
        trend.append({
            'gun': _gun_kisa[g.weekday()], 'tarih': g,
            'sevk': sevk_say[g], 'sevk_h': round(sevk_say[g] * 100 / sevk_max),
        })

    stok_oran = round(sayim_yapan * 100 / sube_sayisi) if sube_sayisi else 0

    ozet_personel = (list(Personel.objects.filter(sube=sel_sube_ozet).order_by('ad_soyad'))
                     if sel_sube_ozet else [])

    return render(request, 'gosterge.html', {
        'personel': personel, 'aktif': 'gosterge',
        'bugun': bugun, 'onay_bekleyen': onay_bekleyen, 'acik_sevkiyat': acik_sevkiyat,
        'sayim_yapan': sayim_yapan, 'sube_sayisi': sube_sayisi,
        'soru': soru, 'bugun_tarih': today,
        'durum': durum, 'durum_top': durum_top,
        'trend': trend, 'stok_oran': stok_oran,
        'sel_sube_ozet': sel_sube_ozet, 'ozet_personel': ozet_personel,
    })

def sevkiyat_kalem_hazirla(request):
    if request.method != 'POST':
        return JsonResponse({'ok': False}, status=405)
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False}, status=403)
    personel = _aktif_personel(request)
    if personel is None or personel.rol != Rol.SEVKIYAT:
        return JsonResponse({'ok': False}, status=403)
    k = SevkiyatKalem.objects.filter(id=request.POST.get('kalem_id')).first()
    if not k:
        return JsonResponse({'ok': False}, status=404)
    k.hazirlandi = (request.POST.get('hazir') == '1')
    k.save(update_fields=['hazirlandi'])
    return JsonResponse({'ok': True, 'hazirlandi': k.hazirlandi})

import os as _os
import re as _re
import glob as _glob
import datetime as _dt
from panel.management.commands.yedek_al import yedek_dizin as _yedek_dizin

_YEDEK_AD_DESEN = _re.compile(r'^yedek-\d{8}-\d{6}\.json\.gz$')

def yedekler(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol != Rol.GENEL_MUDUR:
        return redirect('ana_sayfa')

    if request.method == 'POST' and request.POST.get('islem') == 'al':
        try:
            from django.core.management import call_command
            call_command('yedek_al')
            messages.success(request, "Yedek başarıyla alındı.")
        except Exception as e:
            messages.error(request, f"Yedek alınamadı: {e}")
        return redirect('yedekler')

    d = _yedek_dizin()
    kayitlar = []
    for yol in sorted(_glob.glob(_os.path.join(d, 'yedek-*.json.gz')), reverse=True):
        ad = _os.path.basename(yol)
        try:
            boyut = _os.path.getsize(yol)
            mtime = _dt.datetime.fromtimestamp(_os.path.getmtime(yol))
        except OSError:
            continue
        kayitlar.append({'ad': ad, 'kb': round(boyut / 1024, 1), 'tarih': mtime})
    return render(request, 'yedekler.html', {
        'personel': personel, 'aktif': 'yedekler',
        'kayitlar': kayitlar, 'dizin': d,
    })

def yedek_indir(request, ad):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    personel = _aktif_personel(request)
    if personel is None or personel.rol != Rol.GENEL_MUDUR:
        raise Http404()
    if not _YEDEK_AD_DESEN.match(ad or ''):
        raise Http404()
    yol = _os.path.join(_yedek_dizin(), ad)
    if not _os.path.isfile(yol):
        raise Http404()
    return FileResponse(open(yol, 'rb'), as_attachment=True, filename=ad)

def bildirimler(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    if request.method == 'POST' and request.POST.get('islem') == 'hepsi_oku':
        Bildirim.objects.filter(alici=personel, okundu=False).update(okundu=True)
        return redirect('bildirimler')
    if request.method == 'POST' and request.POST.get('islem') == 'hepsi_sil':
        Bildirim.objects.filter(alici=personel).delete()
        return redirect('bildirimler')
    kayitlar = list(Bildirim.objects.filter(alici=personel)[:100])
    return render(request, 'bildirimler.html', {
        'personel': personel, 'aktif': 'bildirimler', 'kayitlar': kayitlar})

def bildirim_oku(request, bid):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    b = Bildirim.objects.filter(id=bid, alici=personel).first()
    if not b:
        return redirect('bildirimler')
    if not b.okundu:
        b.okundu = True
        b.save(update_fields=['okundu'])
    return redirect(b.link or 'bildirimler')

DUYURU_YAYIN = [Rol.GENEL_MUDUR, Rol.MUDUR, Rol.OPERATOR, Rol.YATIRIMCI]

def duyurular(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    yayinci = personel.rol in DUYURU_YAYIN

    if request.method == 'POST' and yayinci:
        islem = request.POST.get('islem')
        if islem == 'duyuru_ekle':
            baslik = (request.POST.get('baslik') or '').strip()[:150]
            icerik = (request.POST.get('icerik') or '').strip()
            h_rol = request.POST.get('hedef_rol') or ''
            h_rol = h_rol if h_rol in Rol.values else ''
            h_sube = Sube.objects.filter(id=request.POST.get('hedef_sube') or 0).first()
            if baslik and icerik:
                d = Duyuru.objects.create(baslik=baslik, icerik=icerik,
                                          yayinlayan_ad=personel.ad_soyad,
                                          hedef_rol=h_rol, hedef_sube=h_sube)
                hedef = Personel.objects.all()
                if d.hedef_rol:
                    hedef = hedef.filter(rol=d.hedef_rol)
                if d.hedef_sube:
                    hedef = hedef.filter(sube=d.hedef_sube)
                _bildir(list(hedef), "Yeni duyuru: %s" % d.baslik, '/duyurular/', 'duyuru')
                messages.success(request, "Duyuru yayınlandı.")
            else:
                messages.error(request, "Başlık ve içerik gerekli.")
            return redirect('duyurular')
        if islem == 'duyuru_sil':
            Duyuru.objects.filter(id=request.POST.get('duyuru_id')).update(aktif=False)
            messages.success(request, "Duyuru kaldırıldı.")
            return redirect('duyurular')

    if yayinci:
        kayitlar = list(Duyuru.objects.filter(aktif=True)[:100])
    else:
        from django.db.models import Q
        kayitlar = list(Duyuru.objects.filter(aktif=True)
                        .filter(Q(hedef_rol='') | Q(hedef_rol=personel.rol))
                        .filter(Q(hedef_sube__isnull=True) | Q(hedef_sube=personel.sube_id))[:100])

    subeler = list(Sube.objects.order_by('ad'))
    return render(request, 'duyurular.html', {
        'personel': personel, 'aktif': 'duyurular', 'yayinci': yayinci,
        'kayitlar': kayitlar, 'subeler': subeler, 'roller': Rol.choices})

import glob as _glob2
import re as _re2
import datetime as _dt2
from panel.management.commands.aylik_rapor import rapor_dizin as _rapor_dizin

_RAPOR_DESEN = _re2.compile(r'^rapor-\d{4}-\d{2}\.pdf$')

def raporlar(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in UST_YONETIM:
        return redirect('ana_sayfa')

    if request.method == 'POST' and request.POST.get('islem') == 'olustur':
        try:
            from django.core.management import call_command
            call_command('aylik_rapor', '--force')
            messages.success(request, "Geçen ayın raporu oluşturuldu.")
        except Exception as e:
            messages.error(request, "Rapor oluşturulamadı: %s" % e)
        return redirect('raporlar')

    d = _rapor_dizin()
    kayitlar = []
    for yol in sorted(_glob2.glob(_os.path.join(d, 'rapor-*.pdf')), reverse=True):
        ad = _os.path.basename(yol)
        try:
            kb = round(_os.path.getsize(yol) / 1024, 1)
            mt = _dt2.datetime.fromtimestamp(_os.path.getmtime(yol))
        except OSError:
            continue
        kayitlar.append({'ad': ad, 'kb': kb, 'tarih': mt})
    return render(request, 'raporlar.html', {
        'personel': personel, 'aktif': 'raporlar', 'kayitlar': kayitlar})

def rapor_indir(request, ad):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in UST_YONETIM:
        raise Http404()
    if not _RAPOR_DESEN.match(ad or ''):
        raise Http404()
    yol = _os.path.join(_rapor_dizin(), ad)
    if not _os.path.isfile(yol):
        raise Http404()
    return FileResponse(open(yol, 'rb'), as_attachment=True, filename=ad)


GSOSYAL_EMOJILER = ['👍', '❤️', '😂', '😮', '😢', '👏']


def g_sosyal(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    paylasabilir = personel.rol in UST_YONETIM

    if request.method == 'POST' and request.POST.get('islem') == 'tepki':
        g = GSosyalGonderi.objects.filter(id=request.POST.get('gonderi_id')).first()
        emoji = request.POST.get('emoji', '')
        if not g or emoji not in GSOSYAL_EMOJILER:
            return JsonResponse({'ok': False})
        mevcut = GSosyalTepki.objects.filter(gonderi=g, personel=personel).first()
        if mevcut and mevcut.emoji == emoji:
            mevcut.delete()
            benim = ''
        elif mevcut:
            mevcut.emoji = emoji
            mevcut.save()
            benim = emoji
        else:
            GSosyalTepki.objects.create(gonderi=g, personel=personel, emoji=emoji)
            benim = emoji
        sayim = {}
        for t in g.tepkiler.all():
            sayim[t.emoji] = sayim.get(t.emoji, 0) + 1
        return JsonResponse({'ok': True, 'benim': benim,
                             'sayim': {e: sayim.get(e, 0) for e in GSOSYAL_EMOJILER},
                             'toplam': sum(sayim.values())})

    if request.method == 'POST' and request.POST.get('islem') == 'gonderi_ekle' and paylasabilir:
        metin = (request.POST.get('metin') or '').strip()
        gorsel = request.FILES.get('gorsel')
        if gorsel:
            gecerli = False
            if gorsel.size <= 8 * 1024 * 1024 and (gorsel.content_type or '').startswith('image/'):
                try:
                    from PIL import Image as _PILImage
                    _img = _PILImage.open(gorsel)
                    _img.verify()
                    gorsel.seek(0)
                    gecerli = True
                except Exception:
                    gecerli = False
            if not gecerli:
                messages.error(request, "Geçersiz görsel dosyası. Lütfen bir fotoğraf seçin (en fazla 8MB).")
                return redirect('g_sosyal')
        if metin or gorsel:
            GSosyalGonderi.objects.create(yazan=personel, yazan_ad=personel.ad_soyad,
                                          metin=metin[:4000], gorsel=gorsel)
            _bildir(list(Personel.objects.exclude(id=personel.id)),
                    "%s yeni bir Geek Crew paylaşımı yaptı" % personel.ad_soyad,
                    '/g-sosyal/', 'gsosyal')
            messages.success(request, "Paylaşıldı.")
        else:
            messages.error(request, "Bir şeyler yazın ya da görsel ekleyin.")
        return redirect('g_sosyal')

    if request.method == 'POST' and request.POST.get('islem') == 'gonderi_sil':
        g = GSosyalGonderi.objects.filter(id=request.POST.get('gonderi_id')).first()
        if g and (g.yazan_id == personel.id or personel.rol in UST_YONETIM):
            g.delete()
            messages.success(request, "Gönderi silindi.")
        return redirect('g_sosyal')

    gonderiler = list(GSosyalGonderi.objects.select_related('yazan').prefetch_related('tepkiler__personel').order_by('-olusturma')[:80])
    for g in gonderiler:
        sayim = {}
        benim = ''
        sahibi = (g.yazan_id == personel.id)
        verenler = []
        for t in g.tepkiler.all():
            sayim[t.emoji] = sayim.get(t.emoji, 0) + 1
            if t.personel_id == personel.id:
                benim = t.emoji
            if sahibi:
                verenler.append((t.personel.ad_soyad if t.personel else '—', t.emoji))
        g.tepki_listesi = [(e, sayim.get(e, 0)) for e in GSOSYAL_EMOJILER]
        g.tepki_toplam = sum(sayim.values())
        g.benim_tepki = benim
        g.sahibi = sahibi
        g.tepki_verenler = verenler
        g.silebilir = (g.yazan_id == personel.id) or (personel.rol in UST_YONETIM)

    sampiyonlar, sampiyon_sube = _egitim_sampiyon_verisi()
    return render(request, 'g_sosyal.html', {
        'personel': personel,
        'aktif': 'g_sosyal',
        'gonderiler': gonderiler,
        'paylasabilir': paylasabilir,
        'emojiler': GSOSYAL_EMOJILER,
        'sampiyonlar': sampiyonlar[:12],
        'sampiyon_sube': sampiyon_sube,
        'egitim_soru_sayisi': _egitim_ayar_getir().soru_sayisi,
    })


def _pdf_sikistir(pdf_bytes, max_kenar=1600, jpeg_kalite=55):
    """
    PDF içindeki gömülü görselleri (telefonla taranmış reçete/oryantasyon belgelerinde
    sık görülür) yeniden JPEG olarak sıkıştırır. Piksel boyutu zaten küçük olsa bile
    yüksek kalitede (düşük sıkıştırma oranıyla) kaydedilmiş görseller de hedef kalitede
    yeniden kodlanır — asıl büyüklük genelde çözünürlükten değil, orijinal tarama
    uygulamasının kullandığı yüksek JPEG kalitesinden kaynaklanır. Metin/vektör içerik
    olduğu gibi kalır. PyMuPDF kurulu değilse veya herhangi bir sorun olursa orijinal
    bytes olduğu gibi döner — yükleme asla bu yüzden başarısız olmaz.
    """
    try:
        import fitz
    except ImportError:
        return pdf_bytes
    try:
        doc = fitz.open(stream=pdf_bytes, filetype='pdf')
        degisti = False
        for page in doc:
            for img in page.get_images(full=True):
                xref = img[0]
                try:
                    pix = fitz.Pixmap(doc, xref)
                    if pix.colorspace and pix.colorspace.n > 3:
                        pix = fitz.Pixmap(fitz.csRGB, pix)
                    if pix.alpha:
                        pix = fitz.Pixmap(pix, 0)
                    if pix.width > max_kenar or pix.height > max_kenar:
                        oran = max_kenar / max(pix.width, pix.height)
                        pix = fitz.Pixmap(pix, max(1, int(pix.width * oran)), max(1, int(pix.height * oran)))
                    onceki_bytes = doc.xref_stream(xref)
                    jpeg_bytes = pix.tobytes('jpg', jpg_quality=jpeg_kalite)
                    if onceki_bytes is None or len(jpeg_bytes) < len(onceki_bytes):
                        page.replace_image(xref, stream=jpeg_bytes)
                        degisti = True
                except Exception:
                    continue
        if not degisti:
            doc.close()
            return pdf_bytes
        cikti = doc.tobytes(garbage=4, deflate=True)
        doc.close()
        return cikti if len(cikti) < len(pdf_bytes) else pdf_bytes
    except Exception:
        return pdf_bytes


EGITIM_SORU_SAYISI_VARSAYILAN = 10
EGITIM_GECME_VARSAYILAN = 6
EGITIM_SURE_VARSAYILAN = 20
EGITIM_HEDEF_ROLLER = [Rol.PERSONEL, Rol.SEF]
EGITIM_GORUNTULE_ROLLER = [Rol.GENEL_MUDUR, Rol.MUDUR, Rol.OPERATOR, Rol.YATIRIMCI, Rol.EGITMEN]
EGITIM_DUZENLE_ROLLER = [Rol.EGITMEN, Rol.MUDUR]
EGITIM_ACMA_ROLLER = [Rol.GENEL_MUDUR, Rol.OPERATOR, Rol.MUDUR]
# Soru sayısı/süre/geçme puanını ayarlayabilecek roller (Eğitmen + Operatör).
EGITIM_AYAR_ROLLER = [Rol.EGITMEN, Rol.OPERATOR, Rol.MUDUR]
# Açık uçlu (yazılı) cevapları görüp puanlayabilecek roller.
EGITIM_ACIK_PUANLA_ROLLER = [Rol.EGITMEN, Rol.OPERATOR, Rol.MUDUR]


def _egitim_ayar_getir():
    """Tek satırlık EgitimAyar kaydını (yoksa varsayılan değerlerle) döner."""
    ayar, _ = EgitimAyar.objects.get_or_create(id=1, defaults={
        'soru_sayisi': EGITIM_SORU_SAYISI_VARSAYILAN,
        'sure_sn': EGITIM_SURE_VARSAYILAN,
        'gecme_puan': EGITIM_GECME_VARSAYILAN,
    })
    # Eskiden oluşturulmuş kayıtlarda bu alanlar 0/boş kalmışsa (ilk migration) güvenli tarafta kal.
    if not ayar.soru_sayisi:
        ayar.soru_sayisi = EGITIM_SORU_SAYISI_VARSAYILAN
    if not ayar.sure_sn:
        ayar.sure_sn = EGITIM_SURE_VARSAYILAN
    if not ayar.gecme_puan:
        ayar.gecme_puan = EGITIM_GECME_VARSAYILAN
    return ayar


def _egitim_acik(sube=None):
    """Eğitim sistemi açık mı? sube verilirse o şube için, verilmezse en az bir şube için."""
    try:
        a = EgitimAyar.objects.first()
        if not a or not a.acik:
            return False
        acik_ids = set(a.acik_subeler.values_list('id', flat=True))
        if not acik_ids:
            return True  # geriye dönük uyum: hiç şube seçilmemişse tüm şubeler açık
        if sube is None:
            return True
        sid = getattr(sube, 'id', sube)
        return sid in acik_ids
    except Exception:
        return False


def _egitim_soru_havuzu(personel):
    """Kişinin şubesine ait sorular + tüm şubeler için genel sorular."""
    qs = EgitimSoru.objects.filter(aktif=True)
    sube = getattr(personel, 'sube', None)
    if sube is not None:
        return list(qs.filter(Q(sube__isnull=True) | Q(sube=sube)))
    return list(qs.filter(sube__isnull=True))


def _egitim_durum(personel):
    d, _ = EgitimDurum.objects.get_or_create(personel=personel)
    return d


def _egitim_sampiyon_esik(soru_sayisi):
    """Bireysel 'şampiyon' sayılmak için gereken doğru sayısı: sabit 18 doğru."""
    return 18


def _egitim_sampiyon_verisi():
    """Şampiyonlar: soru sayısının ~%80'i veya üzerini yapanlar. Şampiyon şube: DOĞRU CEVAP ORANI
    (denenen sınavlardaki ortalama doğru/soru sayısı) en yüksek olan şube — tamamlama oranı değil."""
    ayar = _egitim_ayar_getir()
    esik = _egitim_sampiyon_esik(ayar.soru_sayisi)
    sampiyonlar = []
    try:
        durumlar = list(EgitimDurum.objects.filter(son_puan__gte=esik)
                        .select_related('personel', 'personel__sube'))
        for d in durumlar:
            p = d.personel
            if p is None or p.rol not in EGITIM_HEDEF_ROLLER:
                continue
            sampiyonlar.append({'ad': p.ad_soyad, 'puan': d.son_puan,
                                'sube': (p.sube.ad if p.sube else '—')})
        sampiyonlar.sort(key=lambda x: (-x['puan'], x['ad']))
    except Exception:
        sampiyonlar = []
    sampiyon_sube = None
    try:
        for s in Sube.objects.filter(depo_mu=False):
            hedef = Personel.objects.filter(sube=s, rol__in=EGITIM_HEDEF_ROLLER)
            toplam = hedef.count()
            if not toplam:
                continue
            denemis = list(EgitimDurum.objects.filter(personel__in=hedef, deneme__gt=0))
            if not denemis:
                continue
            tamam = sum(1 for d in denemis if d.tamamlandi)
            ort_dogru_orani = sum(d.son_puan for d in denemis) / (len(denemis) * ayar.soru_sayisi)
            if (sampiyon_sube is None or ort_dogru_orani > sampiyon_sube['oran'] / 100
                    or (round(ort_dogru_orani * 100) == sampiyon_sube['oran'] and tamam > sampiyon_sube['tamam'])):
                sampiyon_sube = {'ad': s.ad, 'oran': round(ort_dogru_orani * 100), 'tamam': tamam, 'toplam': toplam}
    except Exception:
        sampiyon_sube = None
    return sampiyonlar, sampiyon_sube


def egitim(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    ayar = _egitim_ayar_getir()
    acik = _egitim_acik(personel.sube)
    hedef = personel.rol in EGITIM_HEDEF_ROLLER
    yonetebilir = personel.rol in EGITIM_GORUNTULE_ROLLER
    durum = _egitim_durum(personel) if hedef else None
    if hedef:
        dok_qs = EgitimDokuman.objects.filter(aktif=True).filter(Q(sube__isnull=True) | Q(sube=personel.sube))
        soru_var = len(_egitim_soru_havuzu(personel)) >= ayar.soru_sayisi
    else:
        dok_qs = EgitimDokuman.objects.filter(aktif=True)
        soru_var = EgitimSoru.objects.filter(aktif=True).count() >= ayar.soru_sayisi
    return render(request, 'egitim.html', {
        'personel': personel,
        'aktif': 'egitim',
        'dokumanlar': list(dok_qs),
        'durum': durum,
        'hedef': hedef,
        'yonetebilir': yonetebilir,
        'soru_var': soru_var,
        'acik': acik,
        'gecme': ayar.gecme_puan,
        'soru_sayisi': ayar.soru_sayisi,
    })


def egitim_test(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in EGITIM_HEDEF_ROLLER:
        return redirect('egitim')
    if not _egitim_acik(personel.sube):
        return redirect('egitim')
    durum = _egitim_durum(personel)
    if durum.tamamlandi:
        return redirect('egitim')
    if durum.inceleme_bekliyor:
        messages.info(request, "Önceki sınavının yazılı soruları hâlâ inceleniyor. Sonuç çıkınca tekrar deneyebilirsin.")
        return redirect('egitim')
    if durum.gecti:
        messages.info(request, "Sınavı zaten geçtin — devam etmek için sözleşmeyi onayla.")
        return redirect('egitim_sozlesme')
    ayar = _egitim_ayar_getir()

    if request.method == 'POST':
        ids = [x for x in (request.POST.get('sorular', '') or '').split(',') if x.isdigit()]
        durum.son_sorular = ','.join(ids)
        durum.deneme += 1
        sorular = list(EgitimSoru.objects.filter(id__in=ids))
        durum.son_cevaplar = json.dumps({str(s.id): request.POST.get('soru_%d' % s.id, '') for s in sorular})
        if request.POST.get('cikis') == '1':
            durum.gecti = False
            durum.inceleme_bekliyor = False
            durum.son_puan = 0
            durum.save()
            messages.error(request, "Sınav sırasında uygulamadan ayrıldın — sınav başarısız sayıldı. Farklı sorularla tekrar dene.")
            return redirect('egitim')

        coktan_secmeli = [s for s in sorular if s.tur == 'coktan_secmeli']
        acik_uclu = [s for s in sorular if s.tur == 'acik_uclu']
        dogru_sayi = sum(1 for s in coktan_secmeli if request.POST.get('soru_%d' % s.id, '') == s.dogru)
        durum.son_puan = dogru_sayi

        if acik_uclu:
            # Yazılı sorular var — eğitmen/operatör okuyup puanlayana kadar sonuç kesinleşmez.
            for s in acik_uclu:
                cevap_metni = (request.POST.get('soru_%d' % s.id, '') or '').strip()
                EgitimAcikCevap.objects.create(personel=personel, soru=s, deneme_no=durum.deneme,
                                               cevap_metni=cevap_metni[:4000])
            durum.gecti = False
            durum.inceleme_bekliyor = True
            durum.save()
            _bildir(_rol_personelleri(*EGITIM_ACIK_PUANLA_ROLLER),
                    "%s eğitim sınavını tamamladı, yazılı sorular inceleme bekliyor." % personel.ad_soyad,
                    '/egitim/acik-cevaplar/', 'egitim_acik')
            messages.info(request, "Çoktan seçmeli kısmı değerlendirildi (%d/%d doğru). Yazılı soruların "
                                    "eğitmen/operatör tarafından okunmasının ardından sonucun kesinleşecek."
                                    % (dogru_sayi, len(coktan_secmeli)))
            return redirect('egitim')

        durum.inceleme_bekliyor = False
        if dogru_sayi >= ayar.gecme_puan:
            durum.gecti = True
            durum.save()
            return redirect('egitim_sozlesme')
        durum.gecti = False
        durum.save()
        messages.error(request, "%d/%d doğru — başarısız. Bilgileri tekrar oku, farklı sorularla yeniden dene."
                       % (dogru_sayi, ayar.soru_sayisi))
        return redirect('egitim')

    havuz = _egitim_soru_havuzu(personel)
    if len(havuz) < ayar.soru_sayisi:
        messages.error(request, "Test için yeterli soru tanımlı değil.")
        return redirect('egitim')
    random.shuffle(havuz)
    onceki = set(x for x in (durum.son_sorular or '').split(',') if x.isdigit())
    yeni = [s for s in havuz if str(s.id) not in onceki]
    aday = yeni + [s for s in havuz if str(s.id) in onceki]
    secili = aday[:ayar.soru_sayisi]
    random.shuffle(secili)
    return render(request, 'egitim_test.html', {
        'personel': personel,
        'sorular': secili,
        'sure': ayar.sure_sn,
        'soru_sayisi': ayar.soru_sayisi,
        'gecme': ayar.gecme_puan,
        'id_listesi': ','.join(str(s.id) for s in secili),
    })


def egitim_sozlesme(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in EGITIM_HEDEF_ROLLER:
        return redirect('egitim')
    durum = _egitim_durum(personel)
    if durum.tamamlandi:
        return redirect('egitim')
    if not durum.gecti:
        return redirect('egitim')
    if request.method == 'POST':
        if request.POST.get('onay') == 'on':
            durum.sozlesme_onayli = True
            durum.tamamlandi = True
            durum.save()
            messages.success(request, "Eğitimi tamamladın. Teşekkürler.")
            return redirect('egitim')
        messages.error(request, "Devam etmek için sözleşmeyi okuyup onaylamalısın.")
    return render(request, 'egitim_sozlesme.html', {
        'personel': personel,
        'puan': durum.son_puan,
        'soru_sayisi': _egitim_ayar_getir().soru_sayisi,
    })


def egitim_yonetim(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in EGITIM_GORUNTULE_ROLLER:
        return redirect('egitim')
    duzenleyebilir = personel.rol in EGITIM_DUZENLE_ROLLER
    acabilir = personel.rol in EGITIM_ACMA_ROLLER
    ayar_duzenleyebilir = personel.rol in EGITIM_AYAR_ROLLER

    if request.method == 'POST' and ayar_duzenleyebilir and request.POST.get('islem') == 'karneleri_sifirla':
        hedef_kisiler = Personel.objects.filter(rol__in=EGITIM_HEDEF_ROLLER)
        durumlar = EgitimDurum.objects.filter(personel__in=hedef_kisiler, deneme__gt=0)
        etkilenen = durumlar.count()
        EgitimAcikCevap.objects.filter(personel__in=hedef_kisiler).delete()
        durumlar.update(tamamlandi=False, gecti=False, inceleme_bekliyor=False, son_puan=0,
                        deneme=0, son_sorular='', son_cevaplar='', sozlesme_onayli=False)
        messages.success(request, "%d kişinin sınav karnesi sıfırlandı — herkes tekrar sınava girebilir." % etkilenen)
        return redirect('egitim_yonetim')

    if request.method == 'POST' and acabilir and request.POST.get('islem') in ('sistem_ac', 'sistem_kapat', 'sube_ac', 'sube_kapat'):
        ayar, _ = EgitimAyar.objects.get_or_create(id=1)
        islem = request.POST.get('islem')
        if islem == 'sistem_ac':
            ayar.acik = True
            ayar.save()
            messages.success(request, "Eğitim sistemi açıldı.")
        elif islem == 'sistem_kapat':
            ayar.acik = False
            ayar.save()
            messages.success(request, "Eğitim sistemi kapatıldı.")
        elif islem in ('sube_ac', 'sube_kapat'):
            sb = Sube.objects.filter(id=request.POST.get('sube_id')).first()
            if sb and personel.rol == Rol.MUDUR and not personel.sorumlu_subeler.filter(id=sb.id).exists():
                sb = None
            if sb:
                if islem == 'sube_ac':
                    ayar.acik = True
                    ayar.save()
                    ayar.acik_subeler.add(sb)
                    messages.success(request, "%s şubesi için eğitim açıldı." % sb.ad)
                else:
                    ayar.acik_subeler.remove(sb)
                    messages.success(request, "%s şubesi için eğitim kapatıldı." % sb.ad)
        return redirect('egitim_yonetim')

    if request.method == 'POST' and ayar_duzenleyebilir and request.POST.get('islem') == 'sinav_ayar_kaydet':
        ayar = _egitim_ayar_getir()
        try:
            soru_sayisi = int(request.POST.get('soru_sayisi', ''))
            sure_sn = int(request.POST.get('sure_sn', ''))
            gecme_puan = int(request.POST.get('gecme_puan', ''))
        except (TypeError, ValueError):
            messages.error(request, "Geçerli sayılar girmelisin.")
            return redirect('egitim_yonetim')
        if not (1 <= soru_sayisi <= 100):
            messages.error(request, "Soru sayısı 1-100 arasında olmalı.")
        elif not (5 <= sure_sn <= 600):
            messages.error(request, "Soru başına süre 5-600 saniye arasında olmalı.")
        elif not (1 <= gecme_puan <= soru_sayisi):
            messages.error(request, "Geçme puanı 1 ile soru sayısı arasında olmalı.")
        else:
            ayar.soru_sayisi = soru_sayisi
            ayar.sure_sn = sure_sn
            ayar.gecme_puan = gecme_puan
            ayar.save()
            messages.success(request, "Sınav ayarları güncellendi.")
        return redirect('egitim_yonetim')

    if request.method == 'POST' and duzenleyebilir:
        islem = request.POST.get('islem')
        if islem == 'dokuman_ekle':
            dosya = request.FILES.get('dosya')
            baslik = (request.POST.get('baslik') or '').strip()
            kategori = request.POST.get('kategori') if request.POST.get('kategori') in ('RECETE', 'ORYANTASYON', 'ICECEK') else 'RECETE'
            gecerli_uzanti = ('.pdf', '.mp4', '.webm', '.mov', '.m4v', '.ogv')
            if dosya and baslik and dosya.name.lower().endswith(gecerli_uzanti) and dosya.size <= 200 * 1024 * 1024:
                sb = Sube.objects.filter(id=request.POST.get('sube_id')).first() if (request.POST.get('sube_id') or '').isdigit() else None
                if sb and personel.rol == Rol.MUDUR and not personel.sorumlu_subeler.filter(id=sb.id).exists():
                    sb = None
                orijinal_boyut = dosya.size
                if dosya.name.lower().endswith('.pdf'):
                    try:
                        sikisik = _pdf_sikistir(dosya.read())
                        dosya = ContentFile(sikisik, name=dosya.name)
                    except Exception:
                        dosya.seek(0)
                d = EgitimDokuman.objects.create(kategori=kategori, baslik=baslik[:160], dosya=dosya, sube=sb)
                try:
                    yeni_boyut = d.dosya.size
                    if dosya.name.lower().endswith('.pdf') and yeni_boyut < orijinal_boyut:
                        oran = round(100 - (yeni_boyut / orijinal_boyut * 100))
                        messages.success(request, "Doküman eklendi (PDF %%%d küçültüldü)." % oran)
                    else:
                        messages.success(request, "Doküman eklendi.")
                except Exception:
                    messages.success(request, "Doküman eklendi.")
            else:
                messages.error(request, "Başlık ve geçerli bir dosya (PDF/MP4/WEBM/MOV, en fazla 200MB) gerekli.")
            return redirect('egitim_yonetim')
        if islem == 'dokuman_sil':
            d = EgitimDokuman.objects.filter(id=request.POST.get('id')).first()
            if d:
                if d.dosya:
                    try:
                        d.dosya.delete(save=False)
                    except Exception:
                        pass
                d.delete()
                messages.success(request, "Doküman silindi.")
            return redirect('egitim_yonetim')
        if islem == 'soru_ekle':
            metin = (request.POST.get('metin') or '').strip()
            tur = request.POST.get('tur') if request.POST.get('tur') in ('coktan_secmeli', 'acik_uclu') else 'coktan_secmeli'
            a = (request.POST.get('sik_a') or '').strip()
            b = (request.POST.get('sik_b') or '').strip()
            c = (request.POST.get('sik_c') or '').strip()
            d_ = (request.POST.get('sik_d') or '').strip()
            dogru = request.POST.get('dogru') if request.POST.get('dogru') in ('A', 'B', 'C', 'D') else 'A'
            kategori = request.POST.get('kategori') if request.POST.get('kategori') in ('RECETE', 'ORYANTASYON', 'ICECEK') else 'RECETE'
            gecerli = metin and (tur == 'acik_uclu' or (a and b))
            if gecerli:
                sb = Sube.objects.filter(id=request.POST.get('sube_id')).first() if (request.POST.get('sube_id') or '').isdigit() else None
                if sb and personel.rol == Rol.MUDUR and not personel.sorumlu_subeler.filter(id=sb.id).exists():
                    sb = None
                EgitimSoru.objects.create(kategori=kategori, tur=tur, metin=metin,
                                          sik_a=(a[:300] if tur == 'coktan_secmeli' else ''),
                                          sik_b=(b[:300] if tur == 'coktan_secmeli' else ''),
                                          sik_c=(c[:300] if tur == 'coktan_secmeli' else ''),
                                          sik_d=(d_[:300] if tur == 'coktan_secmeli' else ''),
                                          dogru=(dogru if tur == 'coktan_secmeli' else ''), sube=sb)
                messages.success(request, "Soru eklendi.")
            else:
                messages.error(request, "Soru metni gerekli (çoktan seçmeli için en az A/B şıkları da gerekli).")
            return redirect('egitim_yonetim')
        if islem == 'soru_sil':
            EgitimSoru.objects.filter(id=request.POST.get('id')).delete()
            messages.success(request, "Soru silindi.")
            return redirect('egitim_yonetim')
        if islem == 'soru_aktif':
            s = EgitimSoru.objects.filter(id=request.POST.get('id')).first()
            if s:
                s.aktif = not s.aktif
                s.save()
            return redirect('egitim_yonetim')

    if personel.rol == Rol.MUDUR:
        sube_secenek = list(personel.sorumlu_subeler.all().order_by('ad'))
    else:
        sube_secenek = list(Sube.objects.all().order_by('ad'))
    secili_sube = request.GET.get('sube') or ''
    ayar = _egitim_ayar_getir()

    kisiler_qs = Personel.objects.filter(rol__in=EGITIM_HEDEF_ROLLER).select_related('sube')
    if personel.rol == Rol.MUDUR:
        kisiler_qs = kisiler_qs.filter(sube__in=personel.sorumlu_subeler.all())
    if secili_sube.isdigit():
        kisiler_qs = kisiler_qs.filter(sube_id=int(secili_sube))
    kisiler = list(kisiler_qs.order_by('sube__ad', 'ad_soyad'))
    durum_map = {d.personel_id: d for d in EgitimDurum.objects.filter(personel__in=kisiler)}
    for k in kisiler:
        d = durum_map.get(k.id)
        k.durum_obj = d
        denedi = bool(d and d.deneme)
        k.denedi = denedi
        k.dogru = d.son_puan if denedi else None
        k.yanlis = (ayar.soru_sayisi - d.son_puan) if denedi else None
        k.giris = d.deneme if d else 0
    tamamlayan = sum(1 for k in kisiler if k.durum_obj and k.durum_obj.tamamlandi)

    grup = {}
    for k in kisiler:
        if not k.denedi:
            continue
        g = grup.setdefault(k.sube_id, {'ad': (k.sube.ad if k.sube else '—'), 'dogru': 0, 'yanlis': 0, 'deneme': 0, 'n': 0})
        g['dogru'] += k.dogru
        g['yanlis'] += k.yanlis
        g['deneme'] += k.giris
        g['n'] += 1
    grafik = []
    for _sid, g in grup.items():
        n = g['n'] or 1
        grafik.append({'ad': g['ad'], 'dogru': round(g['dogru'] / n, 1),
                       'yanlis': round(g['yanlis'] / n, 1), 'deneme': round(g['deneme'] / n, 1)})
    grafik.sort(key=lambda x: x['ad'])
    deneme_max = max([g['deneme'] for g in grafik] + [1])
    ayar_obj = EgitimAyar.objects.first()
    acik_sube_ids = list(ayar_obj.acik_subeler.values_list('id', flat=True)) if ayar_obj else []
    sistem_acik = bool(ayar_obj and ayar_obj.acik)
    for s in sube_secenek:
        s.acik_mi = (s.id in acik_sube_ids)
    sampiyonlar, sampiyon_sube = _egitim_sampiyon_verisi()
    acik_cevap_bekleyen = EgitimAcikCevap.objects.filter(puanlandi=False).count()
    return render(request, 'egitim_yonetim.html', {
        'personel': personel,
        'aktif': 'egitim',
        'dokumanlar': list(EgitimDokuman.objects.all().select_related('sube')),
        'sorular': list(EgitimSoru.objects.all().select_related('sube')),
        'aktif_soru': EgitimSoru.objects.filter(aktif=True).count(),
        'kisiler': kisiler,
        'tamamlayan': tamamlayan,
        'toplam_kisi': len(kisiler),
        'subeler': sube_secenek,
        'secili_sube': secili_sube,
        'grafik': grafik,
        'deneme_max': deneme_max,
        'gecme': ayar.gecme_puan,
        'soru_sayisi': ayar.soru_sayisi,
        'sure_sn': ayar.sure_sn,
        'duzenleyebilir': duzenleyebilir,
        'acabilir': acabilir,
        'ayar_duzenleyebilir': ayar_duzenleyebilir,
        'egitim_acik': _egitim_acik(),
        'sistem_acik': sistem_acik,
        'acik_sube_ids': acik_sube_ids,
        'sampiyonlar': sampiyonlar,
        'sampiyon_sube': sampiyon_sube,
        'sampiyon_esik': _egitim_sampiyon_esik(ayar.soru_sayisi),
        'acik_cevap_bekleyen': acik_cevap_bekleyen,
        'acik_puanlayabilir': personel.rol in EGITIM_ACIK_PUANLA_ROLLER,
    })


def egitim_sonucum(request):
    """Personel/Şef kendi son sınav denemesindeki soruları ve doğru/yanlışlarını görür."""
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in EGITIM_HEDEF_ROLLER:
        return redirect('egitim')
    kisi = personel
    durum = EgitimDurum.objects.filter(personel=kisi).first()
    cevap_var = bool(durum and durum.son_cevaplar and durum.son_sorular)
    detay = []
    acik_cevap_map = {}
    if cevap_var and durum.deneme:
        acik_cevap_map = {ac.soru_id: ac for ac in
                          EgitimAcikCevap.objects.filter(personel=kisi, deneme_no=durum.deneme)}
    if cevap_var:
        ids = [x for x in durum.son_sorular.split(',') if x.isdigit()]
        soru_map = {str(s.id): s for s in EgitimSoru.objects.filter(id__in=ids)}
        try:
            cevaplar = json.loads(durum.son_cevaplar or '{}')
        except Exception:
            cevaplar = {}
        for sid in ids:
            s = soru_map.get(sid)
            if not s:
                continue
            verilen = cevaplar.get(sid, '')
            if s.tur == 'acik_uclu':
                ac = acik_cevap_map.get(s.id)
                detay.append({'metin': s.metin, 'acik_uclu': True, 'verilen': verilen,
                              'puanlandi': bool(ac and ac.puanlandi),
                              'dogru_mu': (ac.dogru_mu if ac else None),
                              'puanlama_notu': (ac.puanlama_notu if ac else '')})
                continue
            siklar = [('A', s.sik_a), ('B', s.sik_b)]
            if s.sik_c:
                siklar.append(('C', s.sik_c))
            if s.sik_d:
                siklar.append(('D', s.sik_d))
            detay.append({'metin': s.metin, 'acik_uclu': False, 'siklar': siklar, 'dogru': s.dogru,
                          'verilen': verilen, 'dogru_mu': (verilen == s.dogru)})
    yanlis_sayi = sum(1 for d in detay if not d.get('acik_uclu') and not d['dogru_mu'])
    return render(request, 'egitim_kisi_detay.html', {
        'personel': personel,
        'kisi': kisi,
        'durum': durum,
        'detay': detay,
        'cevap_var': cevap_var,
        'yanlis_sayi': yanlis_sayi,
        'soru_sayisi': _egitim_ayar_getir().soru_sayisi,
        'kendi_sonucu': True,
    })


def egitim_kisi_detay(request, pid):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in EGITIM_GORUNTULE_ROLLER:
        return redirect('egitim')
    kisi = Personel.objects.filter(id=pid).select_related('sube').first()
    if kisi is None:
        return redirect('egitim_yonetim')
    if personel.rol == Rol.MUDUR and kisi.sube_id not in [s.id for s in personel.sorumlu_subeler.all()]:
        return redirect('egitim_yonetim')
    durum = EgitimDurum.objects.filter(personel=kisi).first()
    cevap_var = bool(durum and durum.son_cevaplar and durum.son_sorular)
    detay = []
    acik_cevap_map = {}
    if cevap_var and durum.deneme:
        acik_cevap_map = {ac.soru_id: ac for ac in
                          EgitimAcikCevap.objects.filter(personel=kisi, deneme_no=durum.deneme)}
    if cevap_var:
        ids = [x for x in durum.son_sorular.split(',') if x.isdigit()]
        soru_map = {str(s.id): s for s in EgitimSoru.objects.filter(id__in=ids)}
        try:
            cevaplar = json.loads(durum.son_cevaplar or '{}')
        except Exception:
            cevaplar = {}
        for sid in ids:
            s = soru_map.get(sid)
            if not s:
                continue
            verilen = cevaplar.get(sid, '')
            if s.tur == 'acik_uclu':
                ac = acik_cevap_map.get(s.id)
                detay.append({'metin': s.metin, 'acik_uclu': True, 'verilen': verilen,
                              'puanlandi': bool(ac and ac.puanlandi),
                              'dogru_mu': (ac.dogru_mu if ac else None),
                              'puanlama_notu': (ac.puanlama_notu if ac else '')})
                continue
            siklar = [('A', s.sik_a), ('B', s.sik_b)]
            if s.sik_c:
                siklar.append(('C', s.sik_c))
            if s.sik_d:
                siklar.append(('D', s.sik_d))
            detay.append({'metin': s.metin, 'acik_uclu': False, 'siklar': siklar, 'dogru': s.dogru,
                          'verilen': verilen, 'dogru_mu': (verilen == s.dogru)})
    yanlis_sayi = sum(1 for d in detay if not d.get('acik_uclu') and not d['dogru_mu'])
    return render(request, 'egitim_kisi_detay.html', {
        'personel': personel,
        'kisi': kisi,
        'durum': durum,
        'detay': detay,
        'cevap_var': cevap_var,
        'yanlis_sayi': yanlis_sayi,
        'soru_sayisi': _egitim_ayar_getir().soru_sayisi,
    })


def _egitim_acik_sonucu_kesinlestir(kisi):
    """Bir kişinin güncel denemesindeki tüm açık uçlu cevaplar puanlandıysa sınav sonucunu kesinleştirir."""
    durum = EgitimDurum.objects.filter(personel=kisi).first()
    if durum is None or not durum.inceleme_bekliyor or not durum.deneme:
        return
    bu_deneme = EgitimAcikCevap.objects.filter(personel=kisi, deneme_no=durum.deneme)
    if not bu_deneme.exists() or bu_deneme.filter(puanlandi=False).exists():
        return  # hâlâ puanlanmamış cevap var, bekle
    acik_dogru = bu_deneme.filter(dogru_mu=True).count()
    ayar = _egitim_ayar_getir()
    toplam_dogru = durum.son_puan + acik_dogru
    durum.son_puan = toplam_dogru
    durum.inceleme_bekliyor = False
    if toplam_dogru >= ayar.gecme_puan:
        durum.gecti = True
        durum.save()
        _bildir([kisi], "Eğitim sınavın değerlendirildi: %d/%d doğru — geçtin! Sözleşmeyi onaylamak için Eğitim sayfasına gir."
                % (toplam_dogru, ayar.soru_sayisi), '/egitim/', 'egitim_sonuc')
    else:
        durum.gecti = False
        durum.save()
        _bildir([kisi], "Eğitim sınavın değerlendirildi: %d/%d doğru — başarısız. Farklı sorularla tekrar deneyebilirsin."
                % (toplam_dogru, ayar.soru_sayisi), '/egitim/', 'egitim_sonuc')


def egitim_acik_degerlendir(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in EGITIM_ACIK_PUANLA_ROLLER:
        return redirect('egitim')

    if request.method == 'POST':
        cevap = EgitimAcikCevap.objects.filter(id=request.POST.get('cevap_id')).select_related('personel').first()
        sonuc = request.POST.get('sonuc')
        if cevap and sonuc in ('dogru', 'yanlis'):
            cevap.puanlandi = True
            cevap.dogru_mu = (sonuc == 'dogru')
            cevap.puanlayan = personel
            cevap.puanlama_notu = (request.POST.get('not', '') or '').strip()[:300]
            cevap.puanlama_tarihi = timezone.now()
            cevap.save()
            if cevap.personel:
                _egitim_acik_sonucu_kesinlestir(cevap.personel)
            messages.success(request, "Cevap puanlandı.")
        return redirect('egitim_acik_degerlendir')

    bekleyenler = list(EgitimAcikCevap.objects.filter(puanlandi=False)
                       .select_related('personel', 'personel__sube', 'soru').order_by('olusturma'))
    gruplar = {}
    for c in bekleyenler:
        anahtar = (c.personel_id, c.deneme_no)
        gruplar.setdefault(anahtar, {'kisi': c.personel, 'cevaplar': []})
        gruplar[anahtar]['cevaplar'].append(c)

    gecmis = list(EgitimAcikCevap.objects.filter(puanlandi=True)
                  .select_related('personel', 'puanlayan', 'soru').order_by('-puanlama_tarihi')[:50])

    return render(request, 'egitim_acik_degerlendir.html', {
        'personel': personel,
        'aktif': 'egitim_acik',
        'gruplar': list(gruplar.values()),
        'gecmis': gecmis,
    })


# ------------------------------------------------------------------
# GEEK QR MENÜ — müşteriye açık, giriş gerektirmeyen kalori destekli menü.
# ------------------------------------------------------------------
# NOT: Aşağıdaki ürün/fiyat/kalori verileri ÖRNEK/TASLAKTIR (Menulux'taki canlı
# menüye otomatik erişilemediği için elle, tahmini olarak dolduruldu). Gerçek
# ürün listesi ve kesin kalori bilgisiyle güncellenmeden yayına alınmamalıdır.
GEEK_MENU_KATEGORILER = [
    {
        'kod': 'kahvaltilar',
        'ad': 'Kahvaltılar',
        'urunler': [
            {'ad': 'Serpme Kahvaltı Tabağı', 'aciklama': 'Peynir çeşitleri, zeytin, bal, kaymak, yumurta, taze sebze, reçel',
             'fiyat': 285, 'kalori': 650},
            {'ad': 'Menemen', 'aciklama': 'Domates, biber, yumurta, tereyağı',
             'fiyat': 165, 'kalori': 320},
            {'ad': 'Avokado Tost', 'aciklama': 'Avokado, ekşi maya ekmek, cherry domates, roka',
             'fiyat': 195, 'kalori': 380},
            {'ad': 'Granola Kase', 'aciklama': 'Yoğurt, ev yapımı granola, mevsim meyveleri, bal',
             'fiyat': 175, 'kalori': 410},
            {'ad': 'Sahanda Sucuklu Yumurta', 'aciklama': 'Sucuk, yumurta, tereyağı',
             'fiyat': 155, 'kalori': 340},
        ],
    },
    {
        'kod': 'icecekler',
        'ad': 'İçecekler',
        'urunler': [
            {'ad': 'Filtre Kahve', 'aciklama': 'Günün demlemesi',
             'fiyat': 110, 'kalori': 5},
            {'ad': 'Türk Kahvesi', 'aciklama': 'Şekersiz / az şekerli / şekerli',
             'fiyat': 95, 'kalori': 5},
            {'ad': 'Cappuccino', 'aciklama': 'Espresso, buharda ısıtılmış süt, süt köpüğü',
             'fiyat': 145, 'kalori': 120},
            {'ad': 'Latte', 'aciklama': 'Espresso, bol sütlü, ince köpük',
             'fiyat': 150, 'kalori': 190},
            {'ad': 'Karamel Machiatto', 'aciklama': 'Vanilya, süt, espresso, karamel sos',
             'fiyat': 165, 'kalori': 250},
            {'ad': 'Ice Latte', 'aciklama': 'Soğuk espresso, süt, buz',
             'fiyat': 160, 'kalori': 170},
            {'ad': 'Taze Sıkılmış Portakal Suyu', 'aciklama': 'Günlük taze sıkım',
             'fiyat': 120, 'kalori': 110},
        ],
    },
]


def geek_menu(request):
    """Müşteriye açık QR menü sayfası — giriş gerektirmez."""
    return render(request, 'geek_menu.html', {
        'kategoriler': GEEK_MENU_KATEGORILER,
    })


# ------------------------------------------------------------------
# Şube Denetim Sistemi
# ------------------------------------------------------------------

DENETIM_YAPABILEN_ROLLER = [Rol.MUDUR, Rol.GENEL_MUDUR, Rol.OPERATOR]
DENETIM_GORUNTULEME_ROLLER = [Rol.GENEL_MUDUR, Rol.OPERATOR, Rol.YATIRIMCI]
DENETIM_YONETIM_ROLLER = [Rol.GENEL_MUDUR, Rol.OPERATOR]


def denetim_baslat(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in DENETIM_YAPABILEN_ROLLER:
        return redirect('ana_sayfa')

    if request.method == 'POST' and request.POST.get('islem') == 'baslat':
        sube = Sube.objects.filter(id=request.POST.get('sube_id'), depo_mu=False).first()
        if not sube:
            messages.error(request, "Geçerli bir şube seçmelisin.")
            return redirect('denetim_baslat')
        maddeler = list(DenetimMadde.objects.filter(aktif=True, bolum__aktif=True).select_related('bolum'))
        if not maddeler:
            messages.error(request, "Sistemde henüz denetim maddesi tanımlı değil.")
            return redirect('denetim_baslat')
        denetim = Denetim.objects.create(sube=sube, denetleyen=personel)
        DenetimCevap.objects.bulk_create([DenetimCevap(denetim=denetim, madde=m) for m in maddeler])
        etkilenenler = list(Personel.objects.filter(sube=sube, rol__in=(Rol.MAGAZA_MUDURU, Rol.SEF)))
        if etkilenenler:
            _bildir(etkilenenler, "Şubeniz şu anda denetleniyor.", '', 'denetim')
        messages.success(request, "Denetim başlatıldı: %s" % sube.ad)
        return redirect('denetim_doldur', denetim_id=denetim.id)

    if personel.rol == Rol.MUDUR:
        subeler = list(personel.sorumlu_subeler.filter(depo_mu=False).order_by('ad'))
    else:
        subeler = list(Sube.objects.filter(depo_mu=False).order_by('ad'))
    devam_eden = list(Denetim.objects.filter(denetleyen=personel, tamamlandi=False)
                      .select_related('sube').order_by('-baslangic'))
    return render(request, 'denetim_baslat.html', {
        'personel': personel, 'aktif': 'denetim',
        'subeler': subeler, 'devam_eden': devam_eden,
    })


def denetim_doldur(request, denetim_id):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in DENETIM_YAPABILEN_ROLLER:
        return redirect('ana_sayfa')
    denetim = Denetim.objects.filter(id=denetim_id).select_related('sube').first()
    if denetim is None:
        return redirect('denetim_baslat')
    if denetim.denetleyen_id != personel.id and personel.rol not in (Rol.GENEL_MUDUR, Rol.OPERATOR):
        messages.error(request, "Bu denetimi sadece başlatan kişi ya da yönetim düzenleyebilir.")
        return redirect('denetim_baslat')
    if denetim.tamamlandi:
        return redirect('denetim_detay', denetim_id=denetim.id)

    if request.method == 'POST':
        islem = request.POST.get('islem')
        cevaplar = list(denetim.cevaplar.select_related('madde'))
        for c in cevaplar:
            degisti = False
            puan_raw = request.POST.get('puan_%d' % c.id, '')
            if puan_raw.isdigit() and 0 <= int(puan_raw) <= 5 and c.puan != int(puan_raw):
                c.puan = int(puan_raw)
                degisti = True
            not_raw = (request.POST.get('not_%d' % c.id, '') or '').strip()[:500]
            if not_raw != c.not_metni:
                c.not_metni = not_raw
                degisti = True
            foto_b64 = request.POST.get('foto_b64_%d' % c.id, '')
            if foto_b64.startswith('data:image'):
                raw = None
                try:
                    raw = base64.b64decode(foto_b64.split(',', 1)[1])
                except (ValueError, IndexError):
                    raw = None
                if raw and 100 < len(raw) <= 8 * 1024 * 1024:
                    fname = "denetim_%d_%s.jpg" % (c.id, timezone.now().strftime('%Y%m%d_%H%M%S'))
                    c.foto.save(fname, ContentFile(raw), save=False)
                    degisti = True
            if degisti:
                c.save()

        if islem == 'bitir':
            cevaplanan = [c for c in cevaplar if c.puan is not None]
            if cevaplanan:
                denetim.toplam_puan = round(sum(c.puan for c in cevaplanan) / (len(cevaplanan) * 5) * 100, 1)
            denetim.tamamlandi = True
            denetim.bitis = timezone.now()
            denetim.save()
            messages.success(request, "Denetim tamamlandı. Rapor hazır.")
            return redirect('denetim_detay', denetim_id=denetim.id)

        messages.success(request, "İlerleme kaydedildi.")
        return redirect('denetim_doldur', denetim_id=denetim.id)

    bolumler = DenetimBolum.objects.filter(aktif=True).prefetch_related(
        Prefetch('maddeler', queryset=DenetimMadde.objects.filter(aktif=True)))
    cevap_map = {c.madde_id: c for c in denetim.cevaplar.select_related('madde')}
    bolum_listesi = []
    for b in bolumler:
        satirlar = []
        for m in b.maddeler.all():
            c = cevap_map.get(m.id)
            if c is not None:
                satirlar.append({'madde': m, 'cevap': c})
        if satirlar:
            bolum_listesi.append({'bolum': b, 'maddeler': satirlar})

    cevaplanan_sayi = sum(1 for c in cevap_map.values() if c.puan is not None)
    return render(request, 'denetim_doldur.html', {
        'personel': personel, 'aktif': 'denetim',
        'denetim': denetim, 'bolum_listesi': bolum_listesi,
        'toplam_madde': len(cevap_map), 'cevaplanan_sayi': cevaplanan_sayi,
    })


def denetim_sonuclar(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in DENETIM_GORUNTULEME_ROLLER:
        return redirect('ana_sayfa')
    denetimler = list(Denetim.objects.filter(tamamlandi=True)
                      .select_related('sube', 'denetleyen').order_by('-bitis')[:200])
    return render(request, 'denetim_sonuclar.html', {
        'personel': personel, 'aktif': 'denetim_sonuclar',
        'denetimler': denetimler,
    })


def denetim_detay(request, denetim_id):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    denetim = Denetim.objects.filter(id=denetim_id).select_related('sube', 'denetleyen').first()
    if denetim is None:
        return redirect('ana_sayfa')
    yetkili = personel.rol in DENETIM_GORUNTULEME_ROLLER or denetim.denetleyen_id == personel.id
    if not yetkili:
        return redirect('ana_sayfa')
    cevaplar = list(denetim.cevaplar.select_related('madde', 'madde__bolum')
                    .order_by('madde__bolum__sira', 'madde__sira'))
    gruplar, gruplar_map = [], {}
    for c in cevaplar:
        bolum = c.madde.bolum
        if bolum.id not in gruplar_map:
            grup = {'bolum': bolum, 'cevaplar': []}
            gruplar_map[bolum.id] = grup
            gruplar.append(grup)
        gruplar_map[bolum.id]['cevaplar'].append(c)
    return render(request, 'denetim_detay.html', {
        'personel': personel, 'aktif': 'denetim_sonuclar',
        'denetim': denetim, 'gruplar': gruplar,
    })


def denetim_pdf_indir(request, denetim_id):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    personel = _aktif_personel(request)
    if personel is None:
        return redirect('ana_sayfa')
    denetim = Denetim.objects.filter(id=denetim_id).select_related('sube', 'denetleyen').first()
    if denetim is None:
        return redirect('ana_sayfa')
    yetkili = personel.rol in DENETIM_GORUNTULEME_ROLLER or denetim.denetleyen_id == personel.id
    if not yetkili:
        return redirect('ana_sayfa')
    cevaplar = list(denetim.cevaplar.select_related('madde', 'madde__bolum')
                    .order_by('madde__bolum__sira', 'madde__sira'))
    from .denetim_pdf import denetim_pdf_uret
    icerik = denetim_pdf_uret(denetim, cevaplar)
    resp = HttpResponse(icerik, content_type='application/pdf')
    resp['Content-Disposition'] = 'inline; filename="denetim_%d.pdf"' % denetim.id
    return resp


def denetim_yonetim(request):
    if not request.user.is_authenticated:
        return redirect('ana_sayfa')
    if _cikis_mi(request):
        return _logout(request)
    personel = _aktif_personel(request)
    if personel is None or personel.rol not in DENETIM_YONETIM_ROLLER:
        return redirect('ana_sayfa')

    if request.method == 'POST':
        islem = request.POST.get('islem')
        if islem == 'bolum_ekle':
            ad = (request.POST.get('ad') or '').strip()
            if ad:
                sira = (DenetimBolum.objects.aggregate(m=Max('sira'))['m'] or 0) + 1
                DenetimBolum.objects.create(ad=ad[:200], sira=sira)
                messages.success(request, "Bölüm eklendi.")
            else:
                messages.error(request, "Bölüm adı boş olamaz.")
            return redirect('denetim_yonetim')
        if islem == 'bolum_aktif':
            b = DenetimBolum.objects.filter(id=request.POST.get('id')).first()
            if b:
                b.aktif = not b.aktif
                b.save(update_fields=['aktif'])
            return redirect('denetim_yonetim')
        if islem == 'bolum_sil':
            DenetimBolum.objects.filter(id=request.POST.get('id')).delete()
            messages.success(request, "Bölüm ve içindeki maddeler silindi.")
            return redirect('denetim_yonetim')
        if islem == 'madde_ekle':
            bolum = DenetimBolum.objects.filter(id=request.POST.get('bolum_id')).first()
            metin = (request.POST.get('metin') or '').strip()
            if bolum and metin:
                sira = (bolum.maddeler.aggregate(m=Max('sira'))['m'] or 0) + 1
                DenetimMadde.objects.create(bolum=bolum, metin=metin[:500], sira=sira)
                messages.success(request, "Madde eklendi.")
            else:
                messages.error(request, "Bölüm ve madde metni gerekli.")
            return redirect('denetim_yonetim')
        if islem == 'madde_aktif':
            m = DenetimMadde.objects.filter(id=request.POST.get('id')).first()
            if m:
                m.aktif = not m.aktif
                m.save(update_fields=['aktif'])
            return redirect('denetim_yonetim')
        if islem == 'madde_sil':
            DenetimMadde.objects.filter(id=request.POST.get('id')).delete()
            messages.success(request, "Madde silindi.")
            return redirect('denetim_yonetim')

    bolumler = list(DenetimBolum.objects.all().prefetch_related('maddeler'))
    return render(request, 'denetim_yonetim.html', {
        'personel': personel, 'aktif': 'denetim_yonetim',
        'bolumler': bolumler,
    })

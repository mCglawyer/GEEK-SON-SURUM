import os
from django.conf import settings
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries, coordinate_to_tuple


def excel_logo(ws, band=2):
    try:
        path = os.path.join(settings.BASE_DIR, 'static', 'images', 'geek_logo_blue.png')
        if not os.path.exists(path):
            return
        fp = ws.freeze_panes
        eski = [str(r) for r in list(ws.merged_cells.ranges)]
        for m in eski:
            ws.unmerge_cells(m)
        ws.insert_rows(1, band)
        for m in eski:
            a, b, c, d = range_boundaries(m)
            ws.merge_cells(start_row=b + band, start_column=a, end_row=d + band, end_column=c)
        if fp:
            r, c = coordinate_to_tuple(fp)
            ws.freeze_panes = '%s%d' % (get_column_letter(c), r + band)
        img = XLImage(path)
        if img.height:
            ratio = 44.0 / float(img.height)
            img.width = int(img.width * ratio)
            img.height = 44
        ws.row_dimensions[1].height = 34
        ws.add_image(img, 'A1')
    except Exception:
        pass

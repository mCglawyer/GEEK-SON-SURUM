"""Onaylanan siparişi, fatura sistemine yüklenecek sade bir Excel'e dönüştürür.

Sütunlar: Ürün Adı | Miktar | Birim | Birim Fiyat
- Miktar/Birim: siparişin SON (çıkış) hâli (sevkiyat > satın alma > istenen).
- Birim Fiyat: ürün ne olursa olsun her zaman 0.
"""
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

BASLIKLAR = ['Ürün Adı', 'Miktar', 'Birim', 'Birim Fiyat']


def _say(d):
    if d is None:
        return 0
    d = float(d)
    return int(d) if d == int(d) else round(d, 2)


def siparis_excel_bytes(talep):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Sipariş'
    ws.append(BASLIKLAR)

    kafa_font = Font(bold=True, color='FFFFFF')
    kafa_dolgu = PatternFill('solid', fgColor='162AA3')
    ince = Side(style='thin', color='D0D5E4')
    kenar = Border(left=ince, right=ince, top=ince, bottom=ince)
    orta = Alignment(horizontal='center', vertical='center')

    for c in range(1, 5):
        h = ws.cell(1, c)
        h.font = kafa_font
        h.fill = kafa_dolgu
        h.alignment = orta
        h.border = kenar

    r = 2
    for k in talep.kalemler.all():
        miktar = (k.sevkiyat_miktar if k.sevkiyat_miktar is not None
                  else (k.satinalma_miktar if k.satinalma_miktar is not None else k.istenen_miktar))
        birim = k.sevkiyat_birim or k.satinalma_birim or k.istenen_birim
        ws.cell(r, 1, k.urun_ad)
        ws.cell(r, 2, _say(miktar))
        ws.cell(r, 3, birim)
        ws.cell(r, 4, 0)
        ws.cell(r, 2).alignment = orta
        ws.cell(r, 3).alignment = orta
        ws.cell(r, 4).alignment = orta
        for c in range(1, 5):
            ws.cell(r, c).border = kenar
        r += 1

    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    data = buf.getvalue()
    buf.close()
    return data
